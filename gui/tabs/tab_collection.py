"""数据采集Tab：关键词输入 + 参数配置 + 横向漏斗 + 排序列表 + 开始/停止 + 导出 + 飞轮状态"""
import json
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path

from gui.theme import SURF, SURF2, FG, FG_M, ACC, SUCC, DANGER, BRD, FONTS
from gui.widgets.funnel_progress import FunnelProgress


class CollectionTab:
    def __init__(self, parent: ttk.Frame, app):
        self.parent = parent
        self.app = app
        self._engine = None
        self._sort_state = {}
        self._market_sort_mode = {}  # "uv" or "trend"
        self._last_kw_results = None
        self._last_pd_results = None
        self._last_supply_results = None
        self._build_ui()

    def set_engine(self, engine):
        self._engine = engine

    def _build_ui(self):
        # 上方：关键词输入 + 参数
        top = tk.Frame(self.parent, bg=SURF)
        top.pack(fill=tk.X, padx=8, pady=(8, 4))

        tk.Label(top, text="关键词:", font=FONTS["ui_bold"]).pack(side=tk.LEFT)
        self._kw_text = tk.Text(top, height=3, width=40, font=FONTS["ui"], wrap=tk.WORD,
                                relief=tk.SOLID, borderwidth=1)
        self._kw_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        self._kw_text.insert(tk.END, "手机壳\n蓝牙耳机\n机械键盘")

        ttk.Button(top, text="导入", command=self._import_keywords, width=6).pack(side=tk.LEFT, padx=(4, 0))

        # 参数快速设置
        params = tk.Frame(top, bg=SURF)
        params.pack(side=tk.RIGHT, padx=4)

        ttk.Label(params, text="搜索页:", font=FONTS["ui"]).pack(side=tk.LEFT)
        self._pages_var = tk.IntVar(value=10)
        ttk.Spinbox(params, from_=1, to=30, textvariable=self._pages_var, width=4).pack(side=tk.LEFT, padx=(2, 8))

        ttk.Label(params, text="详情数:", font=FONTS["ui"]).pack(side=tk.LEFT)
        self._detail_var = tk.IntVar(value=5)
        ttk.Spinbox(params, from_=0, to=20, textvariable=self._detail_var, width=4).pack(side=tk.LEFT, padx=(2, 8))

        ttk.Label(params, text="评论数:", font=FONTS["ui"]).pack(side=tk.LEFT)
        self._comment_var = tk.IntVar(value=3)
        ttk.Spinbox(params, from_=0, to=10, textvariable=self._comment_var, width=4).pack(side=tk.LEFT, padx=2)

        ttk.Separator(params, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        # 飞轮状态指示灯
        self._fw_status_canvas = tk.Canvas(params, width=10, height=10, bg=SURF, highlightthickness=0)
        self._fw_status_canvas.pack(side=tk.LEFT, padx=(2, 4))
        self._fw_dot = self._fw_status_canvas.create_oval(1, 1, 9, 9, fill="#6B7280", outline="")
        ttk.Label(params, text="飞轮:", font=FONTS["ui"]).pack(side=tk.LEFT)
        self._fw_status_label = ttk.Label(params, text="待启动", font=FONTS["ui"], foreground=FG_M)
        self._fw_status_label.pack(side=tk.LEFT, padx=(2, 0))

        # 控制按钮
        ctrl = tk.Frame(self.parent, bg=SURF)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        self._start_btn = ttk.Button(ctrl, text="开始采集", command=self._start)
        self._start_btn.pack(side=tk.LEFT, padx=(0, 8))

        self._stop_btn = ttk.Button(ctrl, text="停止", command=self._stop, state=tk.DISABLED)
        self._stop_btn.pack(side=tk.LEFT)

        # 导出按钮 + 下拉菜单
        self._export_btn = ttk.Menubutton(ctrl, text="导出 ▼", direction="above")
        self._export_menu = tk.Menu(self._export_btn, tearoff=0)
        self._export_menu.add_command(label="导出采集结果", command=self._export_collection)
        self._export_menu.add_command(label="导出词库(Pass/Watch)", command=self._export_word_library)
        self._export_btn["menu"] = self._export_menu
        self._export_btn.pack(side=tk.LEFT, padx=(8, 0))

        self._progress_label = ttk.Label(ctrl, text="就绪", font=FONTS["ui"], foreground=FG_M)
        self._progress_label.pack(side=tk.LEFT, padx=15)

        # 下方：左右并排 — 漏斗(左) + 关键词列表(右)
        bottom = tk.Frame(self.parent, bg=SURF)
        bottom.pack(fill=tk.BOTH, expand=True, padx=8, pady=2)

        # 左：漏斗
        left = tk.Frame(bottom, bg=SURF, width=320)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))
        left.pack_propagate(False)

        self._funnel = FunnelProgress(left)
        self._funnel.pack(fill=tk.X, pady=2)

        # 右：关键词列表
        list_frame = tk.LabelFrame(bottom, text="采集详情", font=FONTS["ui_bold"], padx=4, pady=4)
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        columns = [
            ("关键词", 120, "w"),
            ("搜索数", 55, "center"),
            ("行情(UV/涨跌)", 115, "center"),
            ("详情", 45, "center"),
            ("评论", 45, "center"),
            ("状态", 95, "center"),
        ]
        from gui.widgets.tree_helpers import make_columns
        self._tree = ttk.Treeview(list_frame, columns=[c[0] for c in columns], show="headings")
        make_columns(self._tree, columns)

        # 绑定列头点击排序
        self._setup_sortable()

        vsb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._tree.yview)
        hsb = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)

        # 词库面板（底部）
        wl_frame = tk.LabelFrame(self.parent, text="词库", font=FONTS["ui_bold"], padx=4, pady=4)
        wl_frame.pack(fill=tk.BOTH, expand=False, padx=8, pady=(2, 4))
        wl_frame.configure(height=160)

        wl_ctrl = tk.Frame(wl_frame, bg=SURF)
        wl_ctrl.pack(fill=tk.X)
        tk.Label(wl_ctrl, text="", font=FONTS["ui"]).pack(side=tk.LEFT)  # spacer
        self._wl_stats = tk.Label(wl_ctrl, text="", font=FONTS["ui"], fg=FG_M)
        self._wl_stats.pack(side=tk.LEFT, padx=8)
        ttk.Button(wl_ctrl, text="刷新词库", command=self._refresh_word_library, width=10).pack(side=tk.RIGHT)

        self._wl_notebook = ttk.Notebook(wl_frame, height=130)
        self._wl_notebook.pack(fill=tk.BOTH, expand=True, pady=(2, 0))

        # Tab A: 搜索词
        sw_frame = tk.Frame(self._wl_notebook, bg=SURF)
        self._wl_notebook.add(sw_frame, text="搜索词")
        sw_cols = [("关键词", 120, "w"), ("状态", 70, "center"), ("品类方向", 80, "center"),
                   ("来源", 100, "center"), ("行情?", 50, "center")]
        self._wl_sw_tree = ttk.Treeview(sw_frame, columns=[c[0] for c in sw_cols], show="headings", height=5)
        make_columns(self._wl_sw_tree, sw_cols)
        sw_vsb = ttk.Scrollbar(sw_frame, orient=tk.VERTICAL, command=self._wl_sw_tree.yview)
        self._wl_sw_tree.configure(yscrollcommand=sw_vsb.set)
        self._wl_sw_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sw_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Tab B: 品类拓展词
        cs_frame = tk.Frame(self._wl_notebook, bg=SURF)
        self._wl_notebook.add(cs_frame, text="品类拓展词")
        cs_cols = [("拓展词", 100, "w"), ("品类方向", 100, "center"), ("seed_for", 180, "w"),
                   ("状态", 60, "center"), ("来源", 100, "center")]
        self._wl_cs_tree = ttk.Treeview(cs_frame, columns=[c[0] for c in cs_cols], show="headings", height=5)
        make_columns(self._wl_cs_tree, cs_cols)
        cs_vsb = ttk.Scrollbar(cs_frame, orient=tk.VERTICAL, command=self._wl_cs_tree.yview)
        self._wl_cs_tree.configure(yscrollcommand=cs_vsb.set)
        self._wl_cs_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cs_vsb.pack(side=tk.RIGHT, fill=tk.Y)

    # ════════════════════════════════════════════════
    # 表头排序
    # ════════════════════════════════════════════════
    def _setup_sortable(self):
        numeric_cols = {"搜索数", "详情", "评论"}
        self._sort_state = {}
        self._market_sort_mode = {}

        def _parse_market(text, mode="uv"):
            """解析行情单元格: '6634UV ↓1' → UV=6634, trend=-1"""
            if not text or text in ("-", "无行情", ""):
                return 0
            uv_match = re.search(r'(\d+)UV', text)
            if mode == "uv":
                return int(uv_match.group(1)) if uv_match else 0
            else:
                trend_match = re.search(r'([↓↑→])\s*(\d+)', text)
                if trend_match:
                    sign = -1 if trend_match.group(1) == "↓" else (1 if trend_match.group(1) == "↑" else 0)
                    return sign * int(trend_match.group(2))
                return 0

        def sort_col(col):
            # 切换行情列的排序模式
            if col == "行情(UV/涨跌)":
                cur = self._market_sort_mode.get(col, "uv")
                self._market_sort_mode[col] = "trend" if cur == "uv" else "uv"

            reverse = self._sort_state.get(col, False)
            self._sort_state[col] = not reverse

            # 读取所有行
            rows = []
            for k in self._tree.get_children(""):
                vals = self._tree.item(k, "values")
                rows.append((k, vals))

            # 获取列索引
            col_names = [c[0] for c in [
                ("关键词",), ("搜索数",), ("行情(UV/涨跌)",), ("详情",), ("评论",), ("状态",)
            ]]
            try:
                col_idx = col_names.index(col)
            except ValueError:
                col_idx = 0

            # 排序键
            def key_fn(row):
                text = str(row[1][col_idx]) if col_idx < len(row[1]) else ""
                if col in numeric_cols:
                    try:
                        return float(text)
                    except ValueError:
                        return -99999
                elif col == "行情(UV/涨跌)":
                    mode = self._market_sort_mode.get(col, "uv")
                    return _parse_market(text, mode)
                else:
                    return text

            rows.sort(key=key_fn, reverse=reverse)

            # 重新排列
            for idx, (k, _) in enumerate(rows):
                self._tree.move(k, "", idx)

            # 更新表头箭头
            for c_name in col_names:
                if c_name == col:
                    arrow = " ▲" if reverse else " ▼"
                    if c_name == "行情(UV/涨跌)":
                        mode = self._market_sort_mode.get(col, "uv")
                        sub = "[UV]" if mode == "uv" else "[趋势]"
                        self._tree.heading(c_name, text=c_name + arrow + sub)
                    else:
                        self._tree.heading(c_name, text=c_name + arrow)
                else:
                    self._tree.heading(c_name, text=c_name)

            # 重新绑定（lambda 闭包需要用默认参数捕获当前值）
            for c_name in col_names:
                self._tree.heading(c_name, command=lambda cn=c_name: sort_col(cn))

        for c_name in [c[0] for c in [
            ("关键词",), ("搜索数",), ("行情(UV/涨跌)",), ("详情",), ("评论",), ("状态",)
        ]]:
            self._tree.heading(c_name, command=lambda cn=c_name: sort_col(cn))

    # ════════════════════════════════════════════════
    # 数据更新
    # ════════════════════════════════════════════════
    def _start(self):
        if not self._engine:
            return
        text = self._kw_text.get("1.0", tk.END).strip()
        keywords = [k.strip() for k in text.split("\n") if k.strip()]
        if not keywords:
            return

        self._start_btn.configure(state=tk.DISABLED)
        self._stop_btn.configure(state=tk.NORMAL)
        self._window = None
        self._funnel.reset()

        # 清空列表
        self._tree.delete(*self._tree.get_children())
        self._sort_state.clear()
        self._market_sort_mode.clear()
        # 恢复表头（清除箭头）
        for c_name in [c[0] for c in [
            ("关键词",), ("搜索数",), ("行情(UV/涨跌)",), ("详情",), ("评论",), ("状态",)
        ]]:
            self._tree.heading(c_name, text=c_name)

        for kw in keywords:
            self._tree.insert("", tk.END, values=(kw, "-", "-", "-", "-", "排队中"))

        # 设置回调
        self._engine.set_callbacks(
            on_stage=self._on_stage_update,
            on_keyword=self._on_kw_update,
            on_product=None,
            on_complete=self._on_done,
        )

        output_dir = Path.home() / ".xianyu_tool" / "collected_data"
        # 更新飞轮状态
        self._set_flywheel_status("running")
        # 使用并行流水线（如果有的话）
        if hasattr(self._engine, 'start_parallel'):
            self._engine.start_parallel(keywords, output_dir)
        else:
            self._engine.start(keywords, output_dir)

    def _stop(self):
        if self._engine:
            self._engine.stop()
        self._on_done(self._last_kw_results, self._last_pd_results, self._last_supply_results)

    def _import_keywords(self):
        path = filedialog.askopenfilename(
            title="导入关键词文件",
            filetypes=[
                ("Excel文件", "*.xlsx *.xls"),
                ("CSV文件", "*.csv"),
                ("文本文件", "*.txt"),
                ("所有文件", "*.*"),
            ],
        )
        if not path:
            return

        keywords = []
        try:
            ext = Path(path).suffix.lower()
            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                        val = str(row[0]).strip() if row[0] is not None else ""
                        if val and val not in keywords:
                            keywords.append(val)
                    wb.close()
                except ImportError:
                    with open(path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            line = line.strip()
                            if line and line not in keywords:
                                keywords.append(line)
            elif ext == ".csv":
                import csv
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        val = row[0].strip() if row else ""
                        if val and val not in keywords:
                            keywords.append(val)
            else:
                with open(path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line and line not in keywords:
                            keywords.append(line)

            if not keywords:
                messagebox.showwarning("导入结果", "文件中未找到有效关键词")
                return

            self._kw_text.delete("1.0", tk.END)
            self._kw_text.insert("1.0", "\n".join(keywords))
            self.app.logger.info(f"[采集] 已导入 {len(keywords)} 个关键词: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("导入失败", f"读取文件出错:\n{e}")

    def _on_stage_update(self, stage: str, info: str, done: int = 1, total: int = 0):
        self.app.root.after(0, lambda t=info: self._progress_label.configure(text=t))
        # 飞轮状态联动
        if stage == "flywheel":
            self._set_flywheel_status("running")
        stage_map = {
            "flywheel": 0,        # 词·飞轮
            "keyword_scoring": 1, # 词·海选
            "keyword_full": 2,    # 词·精选
            "product_prefilter": 3, # 品·预筛选
            "product_detail": 4,  # 品·详情
            "supply": 5,          # 品·货源
            # 兼容旧stage名
            "market": 0,
            "product_search": 3,
            "product_scoring": 5,
        }
        sidx = stage_map.get(stage)
        if sidx is not None:
            self.app.root.after(0, lambda s=sidx, d=done, t=total: self._funnel.update_stage(s, d, t))

    def _on_kw_update(self, kw: str, idx: int, total: int, status: str,
                       search_cnt: int = 0, has_market: bool = False,
                       detail_cnt: int = 0, comment_cnt: int = 0,
                       market_uv: int = 0, market_price_inc: float = 0):
        def update():
            self._progress_label.configure(text=f"进度: {idx}/{total} — {kw} ({status})")
            for child in self._tree.get_children():
                vals = self._tree.item(child, "values")
                if vals and vals[0] == kw:
                    new_vals = list(vals)
                    if search_cnt > 0:
                        new_vals[1] = str(search_cnt)
                    if has_market and market_uv > 0:
                        trend = f"↑{market_price_inc:.0f}" if market_price_inc > 0 else (f"↓{abs(market_price_inc):.0f}" if market_price_inc < 0 else "→0")
                        new_vals[2] = f"{market_uv}UV {trend}"
                    elif not has_market and new_vals[2] in ("-", ""):
                        new_vals[2] = "无行情"
                    if detail_cnt > 0:
                        new_vals[3] = str(detail_cnt)
                    if comment_cnt > 0:
                        new_vals[4] = str(comment_cnt)
                    new_vals[5] = status
                    self._tree.item(child, values=tuple(new_vals))
                    break
        self.app.root.after(0, update)

    def _on_product_update(self, pd_result: dict):
        def update():
            title = pd_result.get("title", "")[:30]
            grade = pd_result.get("grade", "?")
            total = pd_result.get("total_100", 0)
            self._tree.insert("", tk.END, values=(f"  -> {title}", "-", "-", "-", "-", f"{total}分 {grade}级"))
        self.app.root.after(0, update)

    def _set_flywheel_status(self, status: str):
        """更新飞轮状态指示灯: running/paused/stopped"""
        colors = {"running": "#10B981", "paused": "#F59E0B", "stopped": "#6B7280"}
        labels = {"running": "运行中", "paused": "已暂停", "stopped": "已停止"}
        def update():
            c = colors.get(status, "#6B7280")
            self._fw_status_canvas.itemconfig(self._fw_dot, fill=c)
            self._fw_status_label.configure(text=labels.get(status, status))
        self.app.root.after(0, update)

    def _on_done(self, kw_results=None, pd_results=None, supply_pushed=None):
        self._last_kw_results = kw_results
        self._last_pd_results = pd_results
        self._last_supply_results = supply_pushed
        self._set_flywheel_status("stopped")
        def update():
            self._start_btn.configure(state=tk.NORMAL)
            self._stop_btn.configure(state=tk.DISABLED)
            a_plus = sum(1 for r in (kw_results or []) if r.get("grade") in ("S", "A"))
            s_a = sum(1 for r in (pd_results or []) if r.get("grade") in ("S", "A"))
            self._progress_label.configure(
                text=f"完成: {len(kw_results or [])}词 {len(pd_results or [])}商品 | A+词:{a_plus} S/A品:{s_a} | 货源:{len(supply_pushed or [])}"
            )
            dash = self.app.get_tab("分析看板")
            if dash and hasattr(dash, 'load_results'):
                dash.load_results(kw_results or [], pd_results or [])
            charts = self.app.get_tab("图表分析")
            if charts and hasattr(charts, 'load_data'):
                charts.load_data(kw_results or [], pd_results or [])
            self._refresh_word_library()
        self.app.root.after(0, update)

    # ════════════════════════════════════════════════
    # 导出
    # ════════════════════════════════════════════════

    # ════════════════════════════════════════════════
    # 词库面板
    # ════════════════════════════════════════════════
    def _refresh_word_library(self):
        """加载 word_library.json，刷新搜索词和品类拓展词两个列表"""
        wl_path = Path.home() / ".xianyu_tool" / "collected_data" / "word_library.json"
        if not wl_path.exists():
            wl_path = Path("collected_data/word_library.json")
        if not wl_path.exists():
            self._wl_stats.config(text="词库文件不存在")
            return

        try:
            with open(wl_path, "r", encoding="utf-8") as f:
                lib = json.load(f)
            words = lib.get("words", {})

            search_words = {}
            category_seeds = {}
            for w, info in words.items():
                if info.get("word_type") == "category_seed":
                    category_seeds[w] = info
                else:
                    search_words[w] = info

            # 统计
            pass_n = sum(1 for info in search_words.values() if info.get("status") == "pass")
            watch_n = sum(1 for info in search_words.values() if info.get("status") == "watch")
            pending_n = sum(1 for info in search_words.values() if info.get("status") == "pending_verify")
            no_mkt_n = sum(1 for info in search_words.values()
                          if info.get("scores", {}).get("method") == "no_market"
                          and info.get("status") == "pending_verify")
            self._wl_stats.config(
                text=f"搜索词 {len(search_words)} (pass {pass_n} | watch {watch_n} | pending {pending_n})  |  品类拓展词 {len(category_seeds)}")

            # 清空树
            for tree in [self._wl_sw_tree, self._wl_cs_tree]:
                for item in tree.get_children(""):
                    tree.delete(item)

            # 搜索词列表
            sorted_sw = sorted(search_words.items(),
                              key=lambda x: (x[1].get("status") != "pass", -(float(x[1].get("composite", 0)) if x[1].get("composite") is not None else 0)))
            for word, info in sorted_sw:
                sc = info.get("scores", {})
                tags = []
                if info.get("status") == "pass":
                    tags.append("pass")
                if sc.get("method") == "no_market":
                    tags.append("no_market")

                self._wl_sw_tree.insert("", tk.END, values=(
                    word,
                    info.get("status", ""),
                    info.get("category_direction", sc.get("category_direction", "")),
                    info.get("source", "").replace("phase_b:", "").replace("phase_b_seed:", "[拓展]"),
                    "否(直通)" if sc.get("method") == "no_market" else ("是" if sc.get("has_market") else "否"),
                ), tags=tags)

            # 品类拓展词列表
            sorted_cs = sorted(category_seeds.items(),
                              key=lambda x: len(x[1].get("seed_for", [])), reverse=True)
            for word, info in sorted_cs:
                self._wl_cs_tree.insert("", tk.END, values=(
                    word,
                    info.get("category_direction", ""),
                    ", ".join(info.get("seed_for", [])[:4]),
                    info.get("status", ""),
                    info.get("source", "").replace("phase_b_seed:", ""),
                ))

            # 标签颜色
            for tree, tag, bg, fg in [
                (self._wl_sw_tree, "pass", "#C6EFCE", "#006100"),
                (self._wl_sw_tree, "no_market", "#FFF2CC", "#9C6500"),
            ]:
                tree.tag_configure(tag, background=bg, foreground=fg)

        except Exception as e:
            self._wl_stats.config(text=f"加载失败: {e}")

    def _export_collection(self):
        """导出采集结果到 Excel（关键词评分 + 商品评分 + 货源结果）"""
        kw_results = self._last_kw_results
        pd_results = self._last_pd_results
        supply_results = self._last_supply_results

        if not kw_results and not pd_results:
            # 没有缓存结果，尝试从树中导出
            rows = []
            for child in self._tree.get_children():
                vals = self._tree.item(child, "values")
                if vals:
                    rows.append(vals)
            if not rows:
                messagebox.showinfo("导出", "暂无采集结果可导出")
                return
            # 简单导出树中的数据
            path = filedialog.asksaveasfilename(
                title="导出采集结果",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")],
                initialfile=f"采集结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            )
            if not path:
                return
            self._export_tree_to_file(path, rows)
            return

        path = filedialog.asksaveasfilename(
            title="导出采集结果",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")],
            initialfile=f"采集结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        )
        if not path:
            return

        try:
            from exporter.excel_exporter import ExcelExporter

            # 使用已有导出器
            ExcelExporter.export_dashboard(path, kw_results or [], pd_results or [])

            # 追加货源 Sheet
            if supply_results:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                wb = openpyxl.load_workbook(path)
                ws = wb.create_sheet("货源结果")
                headers = ["商品标题", "来源", "货源标题", "价格", "销量",
                          "佣金", "优惠券", "货源链接", "价值分"]
                header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
                body_font = Font(name="Microsoft YaHei", size=10)

                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                for ri, s in enumerate(supply_results, 2):
                    vals = [
                        s.get("title", s.get("闲鱼标题", "")),
                        s.get("source", s.get("来源", "")),
                        s.get("goods_title", s.get("货源标题", "")),
                        s.get("price", s.get("价格", "")),
                        s.get("sales", s.get("销量", "")),
                        s.get("commission", s.get("佣金", "")),
                        s.get("coupon", s.get("优惠券", "")),
                        s.get("goods_url", s.get("货源链接", "")),
                        s.get("value_score", s.get("价值分", "")),
                    ]
                    for ci, v in enumerate(vals, 1):
                        cell = ws.cell(row=ri, column=ci, value=v)
                        cell.font = body_font
                        cell.alignment = Alignment(horizontal="left" if ci in (1, 3, 8) else "center")

                wb.save(path)
            self.app.logger.info(f"[导出] 采集结果已导出: {path}")
            messagebox.showinfo("导出成功", f"已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_tree_to_file(self, path, rows):
        """简单导出树形列表数据"""
        try:
            if path.endswith(".csv"):
                import csv
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["关键词", "搜索数", "行情(UV/涨跌)", "详情", "评论", "状态"])
                    w.writerows(rows)
            else:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "采集结果"
                headers = ["关键词", "搜索数", "行情(UV/涨跌)", "详情", "评论", "状态"]
                header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for ri, row in enumerate(rows, 2):
                    for ci, v in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=v).font = Font(name="Microsoft YaHei", size=10)
                wb.save(path)
            self.app.logger.info(f"[导出] 采集结果已导出: {path}")
            messagebox.showinfo("导出成功", f"已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_word_library(self):
        """导出词库 — 搜索词 + 品类拓展词 + 标题素材"""
        word_lib_path = Path("collected_data/word_library.json")
        if not word_lib_path.exists():
            word_lib_path = Path.home() / ".xianyu_tool" / "collected_data" / "word_library.json"
        if not word_lib_path.exists():
            messagebox.showinfo("导出", "词库文件不存在，请先运行飞轮")
            return

        path = filedialog.asksaveasfilename(
            title="导出词库",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")],
            initialfile=f"词库_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            with open(word_lib_path, "r", encoding="utf-8") as f:
                lib = json.load(f)
            words = lib.get("words", {})

            hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hf_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            hfont = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
            bfont = Font(name="Microsoft YaHei", size=10)
            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            no_mkt_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

            def write_header(ws, headers, fill=None):
                f = fill or hf
                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.fill = f
                    cell.font = hfont
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

            def write_row(ws, row, vals, fill=None):
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=v if v is not None else "")
                    cell.font = bfont
                    cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill

            def auto_width(ws, headers, max_rows, max_w=40):
                for ci in range(1, len(headers) + 1):
                    w = len(str(headers[ci - 1])) + 4
                    for ri in range(2, min(max_rows + 2, 500)):
                        v = ws.cell(row=ri, column=ci).value
                        if v:
                            w = max(w, len(str(v)) + 2)
                    ws.column_dimensions[get_column_letter(ci)].width = min(w, max_w)

            wb = openpyxl.Workbook()

            # ═══ Sheet 1: 搜索词 ═══
            search_words = {
                w: info for w, info in words.items()
                if info.get("word_type") != "category_seed"
                   and info.get("status") in ("pass", "watch", "pending_verify", "signal_insufficient")
            }

            ws1 = wb.active
            ws1.title = "搜索词"
            sw_headers = ["关键词", "状态", "词类型", "composite",
                         "需求规模", "成交效率", "成交质量", "利润确定性", "竞争格局", "趋势信号",
                         "均价", "涨跌", "行情?", "评分方法", "品类方向", "来源", "添加时间"]
            write_header(ws1, sw_headers)

            sorted_sw = sorted(search_words.items(),
                              key=lambda x: float(x[1].get("composite", 0)) if x[1].get("composite") is not None else 0,
                              reverse=True)
            for ri, (word, info) in enumerate(sorted_sw, 2):
                sc = info.get("scores", {})
                vals = [
                    word,
                    info.get("status", ""),
                    info.get("word_type", "search_word"),
                    info.get("composite", ""),
                    sc.get("demand_scale", sc.get("no_mkt_demand", "")),
                    sc.get("deal_efficiency", ""),
                    sc.get("deal_quality", ""),
                    sc.get("profit_certainty", ""),
                    sc.get("competition", ""),
                    sc.get("trend_signal", sc.get("no_mkt_blue_ocean", "")),
                    sc.get("avg_price", ""),
                    sc.get("avg_price_inc", ""),
                    "是" if sc.get("has_market") else ("否(直通)" if sc.get("method") == "no_market" else "否"),
                    sc.get("method", "standard"),
                    info.get("category_direction", ""),
                    info.get("source", "").replace("phase_b:", "").replace("phase_b_seed:", "[拓展]"),
                    info.get("added_at", ""),
                ]
                row_fill = None
                if info.get("status") == "pass":
                    row_fill = pass_fill
                elif sc.get("method") == "no_market":
                    row_fill = no_mkt_fill
                write_row(ws1, ri, vals, fill=row_fill)
            auto_width(ws1, sw_headers, len(sorted_sw))

            # ═══ Sheet 2: 品类拓展词 ═══
            category_seeds = {
                w: info for w, info in words.items()
                if info.get("word_type") == "category_seed"
            }
            ws2 = wb.create_sheet("品类拓展词")
            cs_headers = ["拓展词", "品类方向", "seed_for (可组合品类)", "状态", "来源", "添加时间"]
            write_header(ws2, cs_headers, fill=hf_green)

            sorted_cs = sorted(category_seeds.items(),
                              key=lambda x: len(x[1].get("seed_for", [])),
                              reverse=True)
            for ri, (word, info) in enumerate(sorted_cs, 2):
                vals = [
                    word,
                    info.get("category_direction", ""),
                    ", ".join(info.get("seed_for", [])[:6]),
                    info.get("status", ""),
                    info.get("source", "").replace("phase_b_seed:", ""),
                    info.get("added_at", ""),
                ]
                write_row(ws2, ri, vals)
            auto_width(ws2, cs_headers, len(sorted_cs), max_w=50)

            # ═══ Sheet 3: 标题素材 ═══
            materials = lib.get("title_materials", {})
            if materials:
                ws3 = wb.create_sheet("标题素材")
                mh = ["分类", "素材", "适用说明"]
                write_header(ws3, mh)
                ri = 2
                for cat, items in materials.items():
                    for item in items:
                        write_row(ws3, ri, [
                            cat,
                            item.get("text", ""),
                            item.get("usage", item.get("conflict_with", "")),
                        ])
                        ri += 1
                auto_width(ws3, mh, ri, max_w=50)

            wb.save(path)
            total = len(sorted_sw) + len(sorted_cs)
            self.app.logger.info(f"[导出] 词库已导出: {path} (搜索词{len(sorted_sw)}+品类拓展词{len(sorted_cs)}={total}词)")
            messagebox.showinfo("导出成功", f"已导出 {total} 词到:\n{path}\n(搜索词 {len(sorted_sw)} + 品类拓展词 {len(sorted_cs)})")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
