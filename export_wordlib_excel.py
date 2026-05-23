"""词库导出 Excel 报告 — 飞轮运转全景图版"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

lib_path = Path.home() / ".xianyu_tool" / "collected_data" / "word_library.json"
data = json.loads(lib_path.read_text(encoding="utf-8"))
words = data.get("words", {})

wb = Workbook()

# ── 样式 ──
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
seed_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
recycle_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
title_font = Font(name="微软雅黑", bold=True, size=14)
normal_font = Font(name="微软雅黑", size=10)
bold_font = Font(name="微软雅黑", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
wrap_align = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, cols, fill=None):
    f = fill or header_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = wrap_align

# ═══════════════════════════════════════════════════════
# 预处理: 按来源分组，重建轮次链条
# ═══════════════════════════════════════════════════════

# key: parent (e.g. "微波炉"), value: {search_words: [...], cat_seeds: [...]}
by_parent = defaultdict(lambda: {"search_words": [], "cat_seeds": [], "earliest_ts": "z"})

for w, info in words.items():
    src = info.get("source", "unknown")
    # 统一 parent: phase_b:X 和 phase_b_seed:X 都归到 X
    if src.startswith("phase_b_seed:"):
        parent = src.replace("phase_b_seed:", "")
        is_from_seed = True
    elif src.startswith("phase_b:"):
        parent = src.replace("phase_b:", "")
        is_from_seed = False
    else:
        parent = src
        is_from_seed = False

    ts = info.get("added_at", "z")
    wt = info.get("word_type", "search_word")

    entry = {"word": w, "status": info.get("status", ""),
             "direction": info.get("category_direction", ""),
             "seed_for": info.get("seed_for", []),
             "composite": info.get("composite", ""),
             "type": wt, "added_at": ts}

    if wt == "category_seed":
        by_parent[parent]["cat_seeds"].append(entry)
    else:
        by_parent[parent]["search_words"].append(entry)

    if ts < by_parent[parent]["earliest_ts"]:
        by_parent[parent]["earliest_ts"] = ts

# 初始种子: 不来自 phase_b_seed 的 parent（且本身是搜索结果产生的）
# 简化: 按时间排序即可
sorted_parents = sorted(by_parent.items(), key=lambda x: x[1]["earliest_ts"])

# 识别哪些 parent 本身就是 category_seed（即它被回种了）
# 在 by_parent 中，如果一个 parent 在别处的 cat_seeds 里出现过，说明它是被回种的
all_cat_seed_names = set()
for parent, items in by_parent.items():
    for cs in items["cat_seeds"]:
        all_cat_seed_names.add(cs["word"])

# 初始种子 = parent 不在 all_cat_seed_names 中
initial_seeds = []
recycled_seeds = []
for parent, items in sorted_parents:
    if parent in all_cat_seed_names:
        recycled_seeds.append((parent, items))
    else:
        initial_seeds.append((parent, items))

# ═══════════════════════════════════════════════════════
# Sheet 1: 飞轮运转全景图
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "飞轮运转全景图"

ws1.merge_cells("A1:J1")
ws1.cell(row=1, column=1, value="闲鱼飞轮 — 运转全景图（因果链路）").font = title_font

# 说明行
ws1.merge_cells("A2:J2")
ws1.cell(row=2, column=1,
         value="绿色行 = 初始种子词搜索  橙色行 = 品类拓展词回种搜索  →  每一行展示该词搜索后产出的 search_word 和 category_seed").font = Font(name="微软雅黑", size=9, color="666666")

# === 初始种子部分 ===
n_init = len(initial_seeds)
ws1.merge_cells(f"A3:J3")
ws1.cell(row=3, column=1, value=f"▌初始种子 ({n_init} 个) — 搜索种子词，从标题中 AI 提取候选词").font = Font(name="微软雅黑", bold=True, size=11, color="2E7D32")

headers1 = ["轮次", "种子词", "搜索词数", "pass", "watch", "pending", "signal_不足", "discard",
            "→ 品类拓展词 (回种下次搜索)", "→ 主要搜索词"]
row = 4
for j, h in enumerate(headers1):
    ws1.cell(row=row, column=j+1, value=h)
style_header(ws1, row, len(headers1), fill=seed_fill)
row += 1

round_num = 0
for parent, items in initial_seeds:
    if not items["search_words"] and not items["cat_seeds"]:
        continue
    round_num += 1
    sw = items["search_words"]
    status_cnt = defaultdict(int)
    for s in sw:
        status_cnt[s["status"]] += 1

    # 品类拓展词
    cs_list = [f"{c['word']}({c.get('direction','')})" for c in items["cat_seeds"]]
    cs_str = "\n".join(cs_list[:8]) if cs_list else "—"

    # 主要搜索词 (pass 优先)
    pass_sw = [s["word"] for s in sw if s["status"] == "pass"][:5]
    other_sw = [s["word"] for s in sw if s["status"] != "pass"][:5]
    sw_show = pass_sw + other_sw
    sw_str = "\n".join(sw_show[:10]) if sw_show else "—"

    ws1.cell(row=row, column=1, value=round_num)
    ws1.cell(row=row, column=2, value=parent)
    ws1.cell(row=row, column=3, value=len(sw))
    ws1.cell(row=row, column=4, value=status_cnt.get("pass", 0))
    ws1.cell(row=row, column=5, value=status_cnt.get("watch", 0))
    ws1.cell(row=row, column=6, value=status_cnt.get("pending_verify", 0))
    ws1.cell(row=row, column=7, value=status_cnt.get("signal_insufficient", 0))
    ws1.cell(row=row, column=8, value=status_cnt.get("discard", 0))
    ws1.cell(row=row, column=9, value=cs_str)
    ws1.cell(row=row, column=10, value=sw_str)
    style_data(ws1, row, len(headers1))
    # 绿色标记初始种子行
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=row, column=c).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row += 1

# === 品类拓展回种部分 ===
n_recycled = len(recycled_seeds)
row += 1
ws1.merge_cells(f"A{row}:J{row}")
ws1.cell(row=row, column=1, value=f"▌品类拓展回种 ({n_recycled} 个) — 上面产出的品类拓展词 → 回种到飞轮 → 单独搜索 → 产出新搜索词").font = Font(name="微软雅黑", bold=True, size=11, color="BF5700")
row += 1

for j, h in enumerate(headers1):
    ws1.cell(row=row, column=j+1, value=h)
style_header(ws1, row, len(headers1), fill=recycle_fill)
row += 1

for parent, items in recycled_seeds:
    if not items["search_words"] and not items["cat_seeds"]:
        continue
    round_num += 1
    sw = items["search_words"]
    status_cnt = defaultdict(int)
    for s in sw:
        status_cnt[s["status"]] += 1

    cs_list = [f"{c['word']}({c.get('direction','')})" for c in items["cat_seeds"]]
    cs_str = "\n".join(cs_list[:8]) if cs_list else "—"

    pass_sw = [s["word"] for s in sw if s["status"] == "pass"][:5]
    other_sw = [s["word"] for s in sw if s["status"] != "pass"][:5]
    sw_show = pass_sw + other_sw
    sw_str = "\n".join(sw_show[:10]) if sw_show else "—"

    ws1.cell(row=row, column=1, value=round_num)
    ws1.cell(row=row, column=2, value=parent)
    ws1.cell(row=row, column=3, value=len(sw))
    ws1.cell(row=row, column=4, value=status_cnt.get("pass", 0))
    ws1.cell(row=row, column=5, value=status_cnt.get("watch", 0))
    ws1.cell(row=row, column=6, value=status_cnt.get("pending_verify", 0))
    ws1.cell(row=row, column=7, value=status_cnt.get("signal_insufficient", 0))
    ws1.cell(row=row, column=8, value=status_cnt.get("discard", 0))
    ws1.cell(row=row, column=9, value=cs_str)
    ws1.cell(row=row, column=10, value=sw_str)
    style_data(ws1, row, len(headers1))
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=row, column=c).fill = PatternFill(start_color="FDE4D0", end_color="FDE4D0", fill_type="solid")
    row += 1

# 汇总
row += 1
ws1.merge_cells(f"A{row}:J{row}")
total_sw = sum(len(v["search_words"]) for _, v in sorted_parents)
total_cs = sum(len(v["cat_seeds"]) for _, v in sorted_parents)
total_pass = sum(1 for w, info in words.items() if info.get("status") == "pass" and info.get("word_type") != "category_seed")
ws1.cell(row=row, column=1,
         value=f"总计: {len(words)}词入库 | search_word {total_sw}个 (✓pass {total_pass}) | category_seed {total_cs}个 | 初始种子{n_init}个 → 品类拓展回种{n_recycled}个 | 品类方向{len(set(info.get('category_direction','') for info in words.values() if info.get('category_direction')))}个").font = bold_font

# 列宽
widths1 = [6, 14, 10, 6, 6, 8, 10, 6, 38, 40]
for i, w in enumerate(widths1):
    ws1.column_dimensions[chr(65 + i)].width = w

# ═══════════════════════════════════════════════════════
# Sheet 2: 搜索词详情 (所有 search_word)
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("搜索词详情")
ws2.merge_cells("A1:H1")
ws2.cell(row=1, column=1, value="全部搜索词 — 按来源分组，含状态和评分").font = title_font

headers2 = ["来源种子", "搜索词", "状态", "词类型", "品类方向", "composite", "evidence"]
for j, h in enumerate(headers2):
    ws2.cell(row=3, column=j+1, value=h)
style_header(ws2, 3, len(headers2))

row = 4
for parent, items in sorted_parents:
    for s in sorted(items["search_words"], key=lambda x: x["word"]):
        ws2.cell(row=row, column=1, value=parent)
        ws2.cell(row=row, column=2, value=s["word"])
        ws2.cell(row=row, column=3, value=s["status"])
        ws2.cell(row=row, column=4, value=s.get("type", "search_word"))
        ws2.cell(row=row, column=5, value=s.get("direction", ""))
        ws2.cell(row=row, column=6, value=s.get("composite", ""))
        ws2.cell(row=row, column=7, value="")
        st = s["status"]
        if st == "pass":
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=c).fill = pass_fill
        style_data(ws2, row, len(headers2))
        row += 1

widths2 = [14, 20, 14, 12, 14, 10, 50]
for i, w in enumerate(widths2):
    ws2.column_dimensions[chr(65 + i)].width = w
ws2.auto_filter.ref = f"A3:G{row - 1}"

# ═══════════════════════════════════════════════════════
# Sheet 3: 品类拓展词详情
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("品类拓展词详情")
ws3.merge_cells("A1:G1")
ws3.cell(row=1, column=1, value="品类拓展词 (category_seed) — 回种到飞轮继续探索").font = title_font

headers3 = ["来源", "拓展词", "品类方向", "seed_for (可组合品类)", "状态", "已在后续搜索?", "回种后产出搜索词数"]
for j, h in enumerate(headers3):
    ws3.cell(row=3, column=j+1, value=h)
style_header(ws3, 3, len(headers3))

# 构建查找: 这个 category_seed 是否被回种过（即出现在 by_parent 的 key 中）
all_parents_set = set(p for p, _ in sorted_parents)

row = 4
for parent, items in sorted_parents:
    for cs in sorted(items["cat_seeds"], key=lambda x: x["word"]):
        cs_word = cs["word"]
        was_recycled = cs_word in all_parents_set
        # 如果回种了，产出多少搜索词
        recycled_output = len(by_parent.get(cs_word, {}).get("search_words", [])) if was_recycled else 0

        ws3.cell(row=row, column=1, value=parent)
        ws3.cell(row=row, column=2, value=cs_word)
        ws3.cell(row=row, column=3, value=cs.get("direction", ""))
        ws3.cell(row=row, column=4, value=", ".join(cs.get("seed_for", [])[:5]))
        ws3.cell(row=row, column=5, value=cs.get("status", ""))
        ws3.cell(row=row, column=6, value="是" if was_recycled else "否")
        ws3.cell(row=row, column=7, value=recycled_output)
        style_data(ws3, row, len(headers3))
        if was_recycled:
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=c).fill = PatternFill(start_color="FDE4D0", end_color="FDE4D0", fill_type="solid")
        row += 1

widths3 = [14, 16, 16, 35, 10, 14, 16]
for i, w in enumerate(widths3):
    ws3.column_dimensions[chr(65 + i)].width = w

# ═══════════════════════════════════════════════════════
# Sheet 4: 品类方向分布
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("品类方向分布")
ws4.merge_cells("A1:E1")
ws4.cell(row=1, column=1, value="品类方向分布 — 按词数排名").font = title_font

directions = defaultdict(list)
for parent, items in sorted_parents:
    for cs in items["cat_seeds"]:
        d = cs.get("direction", "未分类")
        directions[d].append({"word": cs["word"], "from": parent,
                              "seed_for": cs.get("seed_for", [])})

headers4 = ["排名", "品类方向", "词数", "占比", "包含词（来源）"]
for j, h in enumerate(headers4):
    ws4.cell(row=3, column=j+1, value=h)
style_header(ws4, 3, len(headers4))

total_cs = sum(len(v) for v in directions.values())
row = 4
for rank, (d, items) in enumerate(sorted(directions.items(), key=lambda x: -len(x[1])), 1):
    n = len(items)
    pct = n / total_cs * 100 if total_cs else 0
    names = ", ".join(f"{it['word']}(←{it['from']})" for it in items[:8])
    ws4.cell(row=row, column=1, value=rank)
    ws4.cell(row=row, column=2, value=d)
    ws4.cell(row=row, column=3, value=n)
    ws4.cell(row=row, column=4, value=f"{pct:.1f}%")
    ws4.cell(row=row, column=5, value=names)
    style_data(ws4, row, len(headers4))
    row += 1

widths4 = [6, 20, 8, 8, 65]
for i, w in enumerate(widths4):
    ws4.column_dimensions[chr(65 + i)].width = w

# ═══════════════════════════════════════════════════════
# Sheet 5: 跨品类连接
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("跨品类连接")
ws5.merge_cells("A1:E1")
ws5.cell(row=1, column=1, value="跨品类连接 — seed_for 网络：品类拓展词 → 可组合的目标品类").font = title_font

headers5 = ["源词(拓展词)", "来源种子", "目标品类词", "源品类方向", "目标词已在词库?"]
for j, h in enumerate(headers5):
    ws5.cell(row=3, column=j+1, value=h)
style_header(ws5, 3, len(headers5))

row = 4
all_seed_for = []
for parent, items in sorted_parents:
    for cs in items["cat_seeds"]:
        for sf in cs.get("seed_for", []):
            all_seed_for.append((cs["word"], parent, sf, cs.get("direction", "")))

for src_word, src_parent, dst, direction in sorted(all_seed_for, key=lambda x: x[2]):
    exists = "是" if dst in words else "否(新发现)"
    ws5.cell(row=row, column=1, value=src_word)
    ws5.cell(row=row, column=2, value=src_parent)
    ws5.cell(row=row, column=3, value=dst)
    ws5.cell(row=row, column=4, value=direction)
    ws5.cell(row=row, column=5, value=exists)
    style_data(ws5, row, len(headers5))
    row += 1

widths5 = [16, 14, 18, 18, 16]
for i, w in enumerate(widths5):
    ws5.column_dimensions[chr(65 + i)].width = w

# ═══════════════════════════════════════════════════════
# Sheet 6: 回种效率分析
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet("回种效率分析")
ws6.merge_cells("A1:F1")
ws6.cell(row=1, column=1, value="回种效率 — 品类拓展词回种后产出多少有效搜索词").font = title_font

headers6 = ["品类拓展词", "来源", "品类方向", "是否已回种", "回种后产出搜索词数", "产出 pass 词"]
for j, h in enumerate(headers6):
    ws6.cell(row=3, column=j+1, value=h)
style_header(ws6, 3, len(headers6))

row = 4
recycled_stats = []  # for summary
for parent, items in sorted_parents:
    for cs in sorted(items["cat_seeds"], key=lambda x: x["word"]):
        cs_word = cs["word"]
        was_recycled = cs_word in all_parents_set
        if was_recycled:
            output = by_parent.get(cs_word, {}).get("search_words", [])
            output_n = len(output)
            output_pass = [s["word"] for s in output if s["status"] == "pass"]
            recycled_stats.append((cs_word, output_n, len(output_pass)))
        else:
            output_n = 0
            output_pass = []

        ws6.cell(row=row, column=1, value=cs_word)
        ws6.cell(row=row, column=2, value=parent)
        ws6.cell(row=row, column=3, value=cs.get("direction", ""))
        ws6.cell(row=row, column=4, value="是" if was_recycled else "否(待回种)")
        ws6.cell(row=row, column=5, value=output_n)
        ws6.cell(row=row, column=6, value=", ".join(output_pass[:5]) if output_pass else "—")
        style_data(ws6, row, len(headers6))
        if was_recycled and output_pass:
            for c in range(1, len(headers6) + 1):
                ws6.cell(row=row, column=c).fill = pass_fill
        row += 1

row += 1
ws6.merge_cells(f"A{row}:F{row}")
total_recycled = len(recycled_stats)
total_r_output = sum(s[1] for s in recycled_stats)
total_r_pass = sum(s[2] for s in recycled_stats)
if total_recycled:
    ws6.cell(row=row, column=1,
             value=f"回种效率汇总: {total_recycled}个品类拓展词已回种 → 产出{total_r_output}个搜索词 → 其中{total_r_pass}个 pass (转化率 {total_r_pass/total_r_output*100:.1f}%)" if total_r_output else
                   f"回种效率汇总: {total_recycled}个品类拓展词已回种 → 产出{total_r_output}个搜索词").font = bold_font

widths6 = [16, 14, 16, 14, 18, 45]
for i, w in enumerate(widths6):
    ws6.column_dimensions[chr(65 + i)].width = w

# ═══════════════════════════════════════════════════════
# Sheet 7: 总体概览
# ═══════════════════════════════════════════════════════
ws7 = wb.create_sheet("总体概览")
ws7.merge_cells("A1:C1")
ws7.cell(row=1, column=1, value="闲鱼飞轮词库 — 总体概览").font = title_font

statuses = defaultdict(int)
types = defaultdict(int)
for w, info in words.items():
    statuses[info.get("status", "?")] += 1
    types[info.get("word_type", "search_word")] += 1

overview = [
    ["指标", "数值", "说明"],
    ["词库总词数", len(words), "所有入库词"],
    ["search_word (可搜索词)", types.get("search_word", 0), "在闲鱼直接搜索 → 进入选词评分线"],
    ["category_seed (品类拓展词)", types.get("category_seed", 0), "回种飞轮 → 继续探索新品类方向"],
    ["品类方向数", len(set(info.get("category_direction", "") for info in words.values() if info.get("category_direction"))), "category_seed 覆盖的品类方向"],
    ["初始种子数", n_init, "用户输入的种子词"],
    ["品类拓展回种数", n_recycled, "category_seed 被回种搜索的次数"],
    ["pass (可用)", statuses.get("pass", 0), "通过验证，可进入选品"],
    ["watch (观察)", statuses.get("watch", 0), "信号不够强，积累信号后验证"],
    ["pending_verify (待验证)", statuses.get("pending_verify", 0), "AI提取后等待 Phase C 验证"],
    ["signal_insufficient", statuses.get("signal_insufficient", 0), "闲鱼搜索结果太少"],
    ["discard (淘汰)", statuses.get("discard", 0), "明确无价值"],
    ["seed_for 连接数", sum(len(info.get("seed_for", [])) for info in words.values()), "品类拓展词的跨品类连接总数"],
]

for i, row_data in enumerate(overview):
    for j, val in enumerate(row_data):
        ws7.cell(row=i + 3, column=j + 1, value=val)
style_header(ws7, 3, 3)
for i in range(4, 4 + len(overview) - 1):
    style_data(ws7, i, 3)

ws7.column_dimensions["A"].width = 26
ws7.column_dimensions["B"].width = 12
ws7.column_dimensions["C"].width = 45

# ── 保存 ──
output_path = Path.home() / ".xianyu_tool" / "飞轮词库分析报告.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(output_path))
print(f"Excel 已保存: {output_path}")
print(f"共 7 个 Sheet:")
print(f"  1. 飞轮运转全景图 — 因果链路，每轮种子→产出→回种")
print(f"  2. 搜索词详情 — 所有 search_word 按来源分组")
print(f"  3. 品类拓展词详情 — category_seed 回种状态")
print(f"  4. 品类方向分布 — 按词数排名")
print(f"  5. 跨品类连接 — seed_for 网络")
print(f"  6. 回种效率分析 — 品类拓展→有效搜索词转化率")
print(f"  7. 总体概览")

# 打印关键统计
print(f"\n数据摘要:")
print(f"  初始种子 {n_init} 个 → 品类拓展回种 {n_recycled} 个")
print(f"  总词数 {len(words)} → search_word {types.get('search_word', 0)} → pass {statuses.get('pass', 0)}")
if recycled_stats:
    print(f"  回种总产出: {sum(s[1] for s in recycled_stats)} 搜索词, {sum(s[2] for s in recycled_stats)} pass")
