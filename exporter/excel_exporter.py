"""Excel导出：关键词评分/商品评分/货源结果"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List


class ExcelExporter:
    @staticmethod
    def export_dashboard(file_path: str, kw_results: List[dict], pd_results: List[dict]):
        wb = openpyxl.Workbook()

        # Sheet 1: 选词评分
        ws1 = wb.active
        ws1.title = "选词评分"
        kw_headers = ["关键词", "总分", "等级", "需求规模", "成交效率", "成交质量", "利润确定性", "竞争格局", "趋势信号", "均价"]
        ExcelExporter._write_sheet(ws1, kw_headers, kw_results, [
            "keyword", "total_100", "grade",
            lambda r: r.get("scores", {}).get("demand_scale", 0),
            lambda r: r.get("scores", {}).get("deal_efficiency", 0),
            lambda r: r.get("scores", {}).get("deal_quality", 0),
            lambda r: r.get("scores", {}).get("profit_certainty", 0),
            lambda r: r.get("scores", {}).get("competition", 0),
            lambda r: r.get("scores", {}).get("trend_signal", 0),
            "avg_price",
        ])

        # Sheet 2: 商品评分
        ws2 = wb.create_sheet("商品评分")
        pd_headers = ["商品标题", "总分", "等级", "需求信号", "价格优势", "卖家验证", "时效性", "货源属性", "商品质量", "价格"]
        ExcelExporter._write_sheet(ws2, pd_headers, pd_results, [
            "title", "total_100", "grade",
            lambda r: r.get("scores", {}).get("demand_signal", 0),
            lambda r: r.get("scores", {}).get("price_advantage", 0),
            lambda r: r.get("scores", {}).get("seller_validation", 0),
            lambda r: r.get("scores", {}).get("freshness", 0),
            lambda r: r.get("scores", {}).get("supply_attr", 0),
            lambda r: r.get("scores", {}).get("item_quality", 0),
            "price",
        ])

        wb.save(file_path)

    @staticmethod
    def _write_sheet(ws, headers: list, rows: list, cols: list):
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
        body_font = Font(name="Microsoft YaHei", size=10)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                if callable(col):
                    val = col(row)
                else:
                    val = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")

        # Auto-width
        for ci in range(1, len(headers) + 1):
            max_w = len(str(headers[ci - 1])) + 4
            for ri in range(2, len(rows) + 2):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_w = max(max_w, len(str(val)) + 2)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w, 40)
