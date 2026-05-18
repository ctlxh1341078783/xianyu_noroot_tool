"""
数据导入导出模块：JSON/Excel 读写、格式标准化
"""
from __future__ import annotations
import json
import os
from datetime import datetime
from typing import Dict, List, Optional, Any


class DataIO:
    """统一的数据读写接口"""

    @staticmethod
    def load_json(path: str) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def save_json(data: dict, path: str):
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    @staticmethod
    def load_settings(path: str) -> dict:
        if not os.path.exists(path):
            return {}
        return DataIO.load_json(path)

    @staticmethod
    def save_settings(settings: dict, path: str):
        DataIO.save_json(settings, path)

    @staticmethod
    def load_model_config(path: str) -> dict:
        if not os.path.exists(path):
            raise FileNotFoundError(f"模型配置文件不存在: {path}")
        return DataIO.load_json(path)

    @staticmethod
    def keywords_from_excel(path: str, col: str = "关键词") -> List[str]:
        """从 Excel 读取关键词列表，使用 openpyxl 避免 pandas 依赖"""
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value == col:
                    col_idx = cell.column
                    break
            if col_idx is None:
                return []
            keywords = []
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
                val = row[0]
                if val and str(val).strip():
                    keywords.append(str(val).strip())
            return list(dict.fromkeys(keywords))  # 去重保序
        finally:
            wb.close()

    @staticmethod
    def keywords_from_text(text: str) -> List[str]:
        return [kw.strip() for kw in text.strip().split("\n") if kw.strip()]

    @staticmethod
    def export_market_excel(market_results: Dict[str, dict], path: str) -> str:
        """导出行情数据到 Excel，返回文件路径"""
        try:
            from openpyxl import Workbook
            wb = Workbook()

            # Sheet1: 行情概况
            ws1 = wb.active
            ws1.title = "行情概况"
            ws1.append(["关键词", "SPU名称", "均价", "涨跌幅", "热度", "缺货标志", "SPU图片", "类目ID", "SPU ID"])
            for kw, data in market_results.items():
                tb = data.get("topbar", {})
                sh = tb.get("spuHeader", {}) or {}
                ws1.append([kw, sh.get("name", ""), sh.get("avgPrice", ""), sh.get("avgPriceInc", ""),
                           sh.get("hot", ""), sh.get("shortageFlag", ""), sh.get("image", ""),
                           tb.get("categoryId", ""), tb.get("spuId", "")])

            # Sheet2: 成交记录
            ws2 = wb.create_sheet("成交记录")
            ws2.append(["关键词", "历史最高价", "历史最低价", "历史成交量", "商品ID", "标题", "成交价", "发布价", "成交时间", "成交描述", "图片"])
            for kw, data in market_results.items():
                hs = data.get("historysale", {})
                for si in hs.get("itemSaleList", []) or []:
                    ws2.append([kw, hs.get("historyMaxPrice", ""), hs.get("historyMinPrice", ""),
                               hs.get("historyOrder", ""), si.get("itemId", ""), si.get("title", ""),
                               si.get("dealPrice", ""), si.get("publishPrice", ""),
                               si.get("recentSoldTimeDescribe", ""), si.get("salesDescribe", ""),
                               si.get("imageUrl", "")])

            # Sheet3: 价格趋势
            ws3 = wb.create_sheet("价格趋势")
            ws3.append(["关键词", "类型", "AI价格趋势解读", "AI热度解读", "热点指数", "热度等级",
                        "价格看涨比例", "价格看跌比例", "热度看涨比例", "热度看跌比例",
                        "时间", "日均价", "日最高价", "日最低价", "周环比", "搜索UV", "环比增长"])
            for kw, data in market_results.items():
                pt = data.get("pricetrend", {})
                ai = pt.get("aiInterpretationData", {}) or {}
                hot_spot = pt.get("hotSpotIndexData", {}) or {}
                declare = pt.get("declareData", {}) or {}
                ws3.append([kw, "AI解读", ai.get("aiTrendInterpretation", ""), ai.get("aiHotspotInterpretation", ""), "", "", "", "", "", "", "", "", "", "", "", "", ""])
                ws3.append([kw, "热点指数", "", "", hot_spot.get("hotSpotIndex", ""), hot_spot.get("hotSpotLevel", ""), "", "", "", "", "", "", "", "", "", "", ""])
                ws3.append([kw, "多空看板", "", "", "", "", declare.get("bullishRatio", ""), declare.get("bearishRatio", ""), declare.get("hotBullishRatio", ""), declare.get("hotBearishRatio", ""), "", "", "", "", "", "", ""])
                for pp in pt.get("priceTrendList", []) or []:
                    ws3.append([kw, "日均价", "", "", "", "", "", "", "", "", DataIO._ts_str(pp.get("timestamp")), pp.get("dayAvgPrice"), pp.get("dayMaxPrice"), pp.get("dayMinPrice"), pp.get("weekOverWeek"), "", ""])
                htd = pt.get("hotTrendListData", {}) or {}
                for ht in htd.get("hotTrendList", []) or []:
                    ws3.append([kw, "热度趋势", "", "", "", "", "", "", "", "", DataIO._ts_str(ht.get("timestamp")), "", "", "", "", ht.get("historySearchUv"), ht.get("onGrowth")])

            os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
            wb.save(path)
            return path
        except ImportError:
            raise ImportError("openpyxl 未安装，无法导出 Excel")

    @staticmethod
    def export_product_excel(product_results: Dict[str, dict], path: str) -> str:
        """导出选品数据到 Excel"""
        try:
            from openpyxl import Workbook
            wb = Workbook()

            ws1 = wb.active
            ws1.title = "搜索列表"
            ws1.append(["关键词", "序号", "商品ID", "标题", "价格", "卖家昵称", "卖家头像", "卖家身份",
                        "鱼店标签", "地区", "主图URL", "图宽", "图高", "商品标签", "价格标签",
                        "富文本标题", "是否拍卖", "是否广告", "详情页类型", "有视频", "想要", "targetUrl"])
            for kw, data in product_results.items():
                for i, item in enumerate(data.get("search_items", [])):
                    ws1.append([kw, i + 1, item.get("itemId", ""), item.get("title", ""),
                               item.get("price", ""), item.get("userNick", ""), item.get("userAvatarUrl", ""),
                               item.get("userIdentity", ""), item.get("userFishShopLabel", ""), item.get("area", ""),
                               item.get("picUrl", ""), item.get("picWidth", ""), item.get("picHeight", ""),
                               item.get("fishTags", ""), item.get("priceTag", ""), item.get("richTitle", ""),
                               "是" if item.get("isAuction") else "", "是" if item.get("isAliMaMaAD") else "",
                               item.get("detailPageType", ""), "是" if item.get("showVideoIcon") else "",
                               item.get("want", ""), item.get("targetUrl", "")])

            ws2 = wb.create_sheet("商品详情")
            ws2.append(["关键词", "商品ID", "标题", "价格", "商品描述", "类目ID", "发布时间", "商品状态",
                        "可议价", "浏览量", "收藏数", "想要数", "单品已售", "所有图片", "类目属性",
                        "卖家昵称", "卖家头像", "卖家城市", "发货城市", "卖家已售", "好评数", "差评数",
                        "好评率", "回复率", "回复间隔", "平均回复时长(天)", "在售商品数", "注册时间", "最后登录"])
            for kw, data in product_results.items():
                items = {it.get("itemId"): it for it in data.get("search_items", [])}
                for item_id, detail in data.get("details", {}).items():
                    item = items.get(item_id, {})
                    seller = detail.get("sellerDO", {}) or {}
                    item_do = detail.get("itemDO", {}) or {}
                    remark = seller.get("remarkDO", {}) or {}
                    image_urls = " | ".join([img.get("imgUrl", "") or img.get("url", "") for img in (item_do.get("imageInfos", []) or []) if isinstance(img, dict)])
                    cpv_text = "; ".join([f"{lb.get('propertyName','')}:{lb.get('valueName','')}" for lb in (item_do.get("cpvLabels", []) or []) if isinstance(lb, dict)])
                    ws2.append([kw, item_id, (item.get("title", "") or "")[:200], item.get("price", ""),
                               (item_do.get("desc", "") or "")[:800], item_do.get("categoryId", ""),
                               DataIO._ts_str(item_do.get("gmtCreate")), item_do.get("itemStatus", ""),
                               "是" if item_do.get("bargained") == "true" else "", item_do.get("browseCnt", ""),
                               item_do.get("collectCnt", ""), item_do.get("wantCnt", ""), item_do.get("soldCnt", ""),
                               image_urls[:500], cpv_text, seller.get("nick", ""), seller.get("portraitUrl", ""),
                               seller.get("city", ""), seller.get("publishCity", ""), seller.get("hasSoldNumInteger", ""),
                               remark.get("sellerGoodRemarkCnt", ""), remark.get("sellerBadRemarkCnt", ""),
                               seller.get("newGoodRatioRate", ""), seller.get("replyRatio24h", ""),
                               seller.get("replyInterval", ""), seller.get("avgReply30dLong", ""),
                               seller.get("itemCount", ""), DataIO._ts_str(seller.get("registerTime")),
                               seller.get("lastVisitTime", "")])

            ws3 = wb.create_sheet("评论列表")
            ws3.append(["关键词", "商品ID", "评论ID", "评论内容", "评论者", "评论时间", "地区"])
            for kw, data in product_results.items():
                for item_id, comments_data in data.get("comments", {}).items():
                    for c in comments_data.get("items", []) if isinstance(comments_data, dict) else []:
                        rt = c.get("reportTime", "")
                        ws3.append([kw, item_id, c.get("commentId", ""), c.get("content", ""),
                                   c.get("reporterNick", ""), DataIO._ts_str(rt), c.get("ipRegionAddress", "")])

            os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
            wb.save(path)
            return path
        except ImportError:
            raise ImportError("openpyxl 未安装，无法导出 Excel")

    @staticmethod
    def _ts_str(ts):
        if not ts:
            return ""
        try:
            ts_int = int(ts)
            if ts_int > 1e12:
                ts_int = ts_int // 1000
            return datetime.fromtimestamp(ts_int).strftime("%Y-%m-%d %H:%M:%S")
        except:
            return str(ts)
