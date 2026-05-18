"""
货源查找模块 v9.1 - 手机端 uiautomator2 版
==========================================
v9.1 变更（全量AI + 去掉accessory）：
  · [优化1] AI判断从"采样8件"升级为"全量分批"：所有采集到的商品均走AI判断
            每批8件串行调用，批次间间隔2秒防限流，某批失败自动降级到语义相似度
  · [优化2] 去掉 accessory（配件/零件/周边）分类，配件也算可用货源
            提示词只保留 same / alternative / risky / unrelated 四类
  · [优化3] 统一商品处理函数 _process_product_list，代码更简洁

v9.0 变更（提示词升级 + 上架建议表）：
  · [优化1] AI判断改为两阶段：第1步锚点识别(PROMPT_ANALYZE) + 第2步精准匹配(PROMPT_MATCH)
            与test3.py逻辑一致，先理解闲鱼商品的核心交付物/目标用户/核心需求，再做货源比对
            reason字段使用结构化格式，大幅提升匹配准确率
  · [优化2] AI调用频次优化：锚点识别结果缓存复用，两路货源(标题/图搜)共享同一锚点
            每件商品只多消耗1次API（锚点），其余保持原有分批策略不变
  · [优化3] 新增上架建议表，字段：上架天数/想要数/日均想要数/商品描述/商品价格/
            最优同款前3/平替前3/货源详情/上架优先级（S/A/B/C/D级）
  · [优化4] 闲鱼需求数据（已上架天数/想要人数/日均想要数）从采集数据透传到记录中
  · [优化5] 上架建议表可独立导出Excel

v8.1 变更（API超时优化）：
  · 分批调用：标题路和图搜路各自独立调用API，不再合并一次请求
    - 单次token量减半，超时概率大幅降低
    - 任意一路失败只影响那一路，另一路AI结果仍然保留
  · 智能采样：从前N件改为价格分散采样（覆盖低/中/高价区间）
    - 每路最多8件（原来12件），确保单次输入token可控
  · 提示词精简：去掉大段说明和示例，core指令压缩到200字内
    - max_tokens从8000降到2000，输出也更快
  · 超时上限从90s降到55s（单路8件不需要那么久）
  · 分路降级：标题路/图搜路各自有独立AI状态，失败的那路用
    语义相似度兜底，成功的那路继续用AI结果

v8.0 变更：
  · [需求1] 去掉AI清洗标题，改用全标题直接搜索PDD，不截断
  · [需求2] AI提示词升级：识别5种商品类型
            same(同款) / alternative(平替) / accessory(配件) /
            unrelated(无关品) / risky(有风险)
            新增4个细分利润字段：标题同款利润、标题平替利润、图搜同款利润、图搜平替利润
  · [需求3] 双阶段风控（第1阶段清缓存，第2阶段才告警）
  · [需求4] 更新设置后打印所有实际生效值
"""

from __future__ import annotations
import os
import re
import time
import random
import threading
import queue
import requests
import json as _json
import subprocess
import sys
import tempfile
import webbrowser
from typing import List, Dict, Optional, Callable, Set, Tuple
from datetime import datetime, timedelta
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

try:
    import uiautomator2 as u2
    U2_OK = True
except ImportError:
    U2_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

try:
    import jieba
    jieba.setLogLevel(60)
    JIEBA_OK = True
except ImportError:
    JIEBA_OK = False

TEXT2VEC_OK = None  # 懒加载，首次使用时才检查


# ══════════════════════════════════════════════════════════════════════════
# AI 提示词定义 v9.0（两阶段：锚点识别 + 精准匹配）
# ══════════════════════════════════════════════════════════════════════════

# ── 第1步：锚点识别（每件闲鱼商品只调用1次，结果缓存给标题路+图搜路共用）──
PROMPT_ANALYZE = """请分析以下闲鱼商品，输出严格符合格式的JSON对象，不要任何额外文字或Markdown。

闲鱼标题：{xianyu_title}
描述：{xianyu_desc}

要求：根据标题和描述的实际内容，自由提炼，不要使用泛化词汇：
1. core_item：买家实际收到的是什么（具体物品/服务/虚拟权益，精确描述）
2. target_user：谁会购买，越具体越好（例：想快速提升抖音等级的账号运营者）
3. core_need：买家核心诉求/解决问题
4. summary：一句话总结，格式 [谁]需要[什么]来[干什么]

输出纯JSON，不要用```json标记，直接输出：
{{"core_item":"...","target_user":"...","core_need":"...","summary":"..."}}"""

# ── 第2步：货源匹配（标题路/图搜路各调用1次，共享锚点结果）────────────────
PROMPT_MATCH = """你是货源匹配专家。根据闲鱼需求锚点判断每条货源。

【闲鱼需求锚点】
- 交付物：{core_item}
- 用户：{target_user}
- 需求：{core_need}

判断标准：
- same：交付物完全一致，可直接发货
- alternative：有差异但能满足同一用户同一需求，可作平替
- risky：能满足但有风险（二手/瑕疵/来源不明）
- unrelated：无法满足该用户的需求

【{source_type}列表（第{batch_no}批，共{batch_total}批，idx从{idx_start}开始）】
{product_list}

输出JSON数组，idx必须从{idx_start}开始连续编号，每条包含idx、product_type、reason。
reason格式："货源是【核心品类】，用户需要【核心需求】，因为【判断理由】，所以【结论】"
注意：货源描述不要复制原标题，要提炼核心品类

注意：product_type 必须是以下值之一：same、alternative、risky、unrelated（不要使用中文）

示例：
{{"idx":{idx_start},"product_type":"unrelated","reason":"货源是充电宝，用户需要220V户外电源，因为无法提供220V输出，所以无关"}}

只输出JSON数组，不要任何其他文字："""

# 兼容旧引用
PROMPT_COMPARE_SINGLE = PROMPT_MATCH
PROMPT_COMPARE_BATCH = PROMPT_MATCH

# ══════════════════════════════════════════════════════════════════════════
# 匹配等级枚举
# ══════════════════════════════════════════════════════════════════════════

class TitleCleanerAI:
    """
    使用智谱 API 处理两类任务（两次独立调用，时序分离）：
      · 第一次：clean_title()  —— 在搜索前调用，清洗闲鱼标题 → 得到PDD搜索词
      · 第二次：compare_batch() —— 采集完成后调用，批量判断标题搜+图搜商品是否同款
    降级链：API失败 → 语义模型(SameProductMatcher) → 字符相似度
    
    v6.1 增强特性：
      · 熔断保护机制（连续失败自动暂停）
      · 智能超时计算（基于历史性能）
      · 网络质量检测
      · 更完善的重试策略
      · 优化提示词，增加详细理由输出
    """

    API_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"

    def __init__(self, api_key: str = "", log_cb: Callable = None):
        self.api_key = api_key
        self.log = log_cb or print
        self.enabled = bool(api_key)
        self._stats = {
            "clean_calls": 0, "clean_ok": 0, "clean_fail": 0,
            "compare_calls": 0, "compare_ok": 0, "compare_fail": 0,
        }
        
        # ========== 新增：熔断器相关 ==========
        self._consecutive_failures = 0          # 连续失败次数
        self._circuit_breaker_open = False      # 熔断开关
        self._circuit_breaker_open_time = 0     # 熔断开启时间
        self._circuit_breaker_threshold = 3     # 连续3次失败触发熔断
        self._circuit_breaker_cooldown = 90     # 熔断90秒
        
        # ========== 新增：网络质量追踪 ==========
        self._api_call_history = []             # 记录最近10次API耗时（秒）
        self._last_api_duration = 0             # 最近一次API耗时
        self._last_api_time = 0                 # 最近一次API调用时间
        
        # ========== 新增：限流保护 ==========
        self._rate_limit_counter = 0             # 限流计数
        self._rate_limit_reset_time = 0          # 限流重置时间
        self._request_interval = 1.0             # 请求间隔（秒）

    def set_api_key(self, api_key: str):
        self.api_key = api_key
        self.enabled = bool(api_key)
        if not api_key:
            self.log("⚠️ API Key 已清空，将使用本地清洗")
        else:
            # 重置熔断状态
            self._reset_circuit_breaker()
            self.log("✅ API Key 已设置，熔断器已重置")

    def get_stats(self) -> Dict:
        return self._stats.copy()

    def reset_stats(self):
        self._stats = {
            "clean_calls": 0, "clean_ok": 0, "clean_fail": 0,
            "compare_calls": 0, "compare_ok": 0, "compare_fail": 0,
        }
    
    def _reset_circuit_breaker(self):
        """重置熔断器状态"""
        self._consecutive_failures = 0
        self._circuit_breaker_open = False
        self._circuit_breaker_open_time = 0
        if self.log:
            self.log("  🔓 熔断器已重置")

    def _should_use_circuit_breaker(self) -> bool:
        """
        检查是否应该使用熔断保护
        返回 True 表示应该跳过API调用
        """
        if not self._circuit_breaker_open:
            return False
        
        # 检查冷却时间是否已过
        elapsed = time.time() - self._circuit_breaker_open_time
        if elapsed >= self._circuit_breaker_cooldown:
            self._circuit_breaker_open = False
            self._consecutive_failures = 0
            if self.log:
                self.log("  🔓 熔断器冷却完成，恢复API调用")
            return False
        
        # 还在熔断期
        remaining = int(self._circuit_breaker_cooldown - elapsed)
        if self.log and remaining % 30 == 0:  # 每30秒提示一次
            self.log(f"  🔒 熔断保护中，剩余{remaining}秒")
        return True
    
    def _check_rate_limit(self) -> bool:
        """
        检查限流状态
        返回 True 表示应该等待
        """
        now = time.time()
        
        # 每60秒重置计数器
        if now - self._rate_limit_reset_time > 60:
            self._rate_limit_counter = 0
            self._rate_limit_reset_time = now
        
        # 每分钟超过20次请求，触发限流等待
        if self._rate_limit_counter >= 20:
            wait_time = 60 - (now - self._rate_limit_reset_time)
            if wait_time > 0:
                if self.log:
                    self.log(f"  ⚠️ 达到限流阈值，等待{wait_time:.0f}秒...")
                time.sleep(wait_time)
                self._rate_limit_counter = 0
                self._rate_limit_reset_time = time.time()
            return True
        
        return False
    
    def _record_api_call(self, duration: float, success: bool, is_timeout: bool = False):
        """
        记录API调用结果，用于熔断和性能分析
        """
        self._last_api_duration = duration
        self._last_api_time = time.time()
        
        # 记录历史耗时（只记录成功的，避免超时数据污染）
        if success:
            self._api_call_history.append(duration)
            if len(self._api_call_history) > 10:
                self._api_call_history.pop(0)
        
        # 更新熔断状态
        if not success:
            self._consecutive_failures += 1
            if self._consecutive_failures >= self._circuit_breaker_threshold:
                self._circuit_breaker_open = True
                self._circuit_breaker_open_time = time.time()
                if self.log:
                    self.log(f"  🔥 连续{self._consecutive_failures}次失败，触发熔断保护（{self._circuit_breaker_cooldown}秒）")
        else:
            # 成功后重置失败计数
            self._consecutive_failures = 0
            
            # 如果之前熔断过，记录恢复
            if self._circuit_breaker_open:
                self._circuit_breaker_open = False
                if self.log:
                    self.log("  ✅ API恢复成功，熔断器关闭")
    
    def _calculate_timeout(self, base_timeout: int, total_items: int = 0) -> int:
        """
        智能计算超时时间（优化版：分批后单路最多8件，无需超过55s）
        """
        # 1. 根据商品数量调整（单路最多8件）
        if total_items > 0:
            # 基础40s + 每件+1.5s，上限55s
            timeout = min(55, 40 + int(total_items * 1.5))
        else:
            timeout = min(45, base_timeout + 10)
        
        # 2. 根据历史性能调整
        if self._api_call_history:
            avg_duration = sum(self._api_call_history) / len(self._api_call_history)
            if avg_duration > 25:
                timeout = min(55, int(timeout * 1.2))
                if self.log and timeout > base_timeout:
                    self.log(f"  📊 网络较慢(均{avg_duration:.0f}s)，超时→{timeout}s")
        
        # 3. 上次接近超时，小幅追加
        if self._last_api_duration >= base_timeout * 0.85:
            timeout = min(55, int(timeout * 1.1))
        
        return timeout
    
    def _enforce_request_interval(self):
        """强制请求间隔，避免限流"""
        now = time.time()
        elapsed = now - getattr(self, '_last_request_time', 0)
        if elapsed < self._request_interval:
            sleep_time = self._request_interval - elapsed
            if sleep_time > 0:
                time.sleep(sleep_time)
        self._last_request_time = time.time()
    
    def test_connection(self) -> Tuple[bool, str]:
        if not self.api_key:
            return False, "未配置 API Key"
        try:
            headers = {"Content-Type": "application/json",
                       "Authorization": f"Bearer {self.api_key}"}
            data = {"model": "glm-4-flash",
                    "messages": [{"role": "user", "content": "测试"}],
                    "max_tokens": 10}
            resp = requests.post(self.API_URL, headers=headers, json=data, timeout=15)
            if resp.status_code == 200:
                # 测试成功，重置熔断器
                self._reset_circuit_breaker()
                return True, "连接成功"
            return False, f"HTTP {resp.status_code}: {resp.text[:100]}"
        except Exception as e:
            return False, str(e)

    def _call_api(self, prompt: str, max_tokens: int = 500,
                  timeout: int = 30, temperature: float = 0.1,
                  retry_count: int = 0, total_items: int = 0) -> Optional[str]:
        """
        底层 HTTP 调用，支持自动重试和递增超时
        增强版：熔断保护 + 智能超时 + 限流控制
        
        参数:
            prompt: 提示词
            max_tokens: 最大token数
            timeout: 超时时间（秒）
            temperature: 温度参数
            retry_count: 当前重试次数（内部使用）
            total_items: 商品总数（用于动态调整超时）
        
        返回:
            成功返回文本内容，失败返回 None
        """
        if not self.enabled or not self.api_key:
            return None
        
        # 熔断检查
        if self._should_use_circuit_breaker():
            return None
        
        # 限流检查
        self._check_rate_limit()
        
        # 强制请求间隔
        self._enforce_request_interval()
        
        max_retries = 2  # 最多重试2次
        
        # 智能超时计算
        if total_items > 0:
            current_timeout = self._calculate_timeout(timeout, total_items)
        else:
            current_timeout = self._calculate_timeout(timeout)
        
        # 重试时增加超时
        if retry_count > 0:
            current_timeout = min(90, current_timeout + (retry_count * 10))
        
        start_time = time.time()
        
        try:
            headers = {"Content-Type": "application/json",
                       "Authorization": f"Bearer {self.api_key}"}
            data = {"model": "glm-4-flash",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": temperature,
                    "max_tokens": max_tokens}
            
            if self.log and retry_count > 0:
                self.log(f"  🔄 API重试 ({retry_count}/{max_retries})，超时{current_timeout}s...")
            
            # 记录限流计数
            self._rate_limit_counter += 1
            
            resp = requests.post(self.API_URL, headers=headers, 
                                json=data, timeout=current_timeout)
            
            elapsed = time.time() - start_time
            success = resp.status_code == 200
            
            # 记录调用结果
            self._record_api_call(elapsed, success)
            
            if resp.status_code == 200:
                content = resp.json()['choices'][0]['message']['content'].strip()
                if self.log and elapsed > 30:
                    self.log(f"  ⏱️ API耗时{elapsed:.1f}s（较慢）")
                return content
                
            elif resp.status_code == 429:  # 限流
                if self.log:
                    self.log(f"  ⚠️ API限流 (HTTP 429)")
                
                # 限流时增加冷却时间
                self._circuit_breaker_cooldown = 120
                
                if retry_count < max_retries:
                    wait_time = 5 * (retry_count + 1)  # 5s, 10s
                    if self.log:
                        self.log(f"  ⏳ 等待{wait_time}秒后重试...")
                    time.sleep(wait_time)
                    return self._call_api(prompt, max_tokens, timeout, 
                                        temperature, retry_count + 1, total_items)
                return None
                
            elif resp.status_code == 500 or resp.status_code == 502 or resp.status_code == 503:
                # 服务器错误，可重试
                if self.log:
                    self.log(f"  ⚠️ 服务器错误 (HTTP {resp.status_code})")
                
                if retry_count < max_retries:
                    wait_time = 3 * (retry_count + 1)
                    time.sleep(wait_time)
                    return self._call_api(prompt, max_tokens, timeout,
                                        temperature, retry_count + 1, total_items)
                return None
                
            else:
                if self.log:
                    self.log(f"  ⚠️ API HTTP {resp.status_code}")
                return None
                
        except requests.exceptions.Timeout:
            elapsed = time.time() - start_time
            self._record_api_call(elapsed, False, is_timeout=True)
            
            if retry_count < max_retries:
                wait_time = 3 * (retry_count + 1)
                if self.log:
                    self.log(f"  ⏱️ API超时({current_timeout}s，实际{elapsed:.1f}s)，"
                            f"{wait_time}秒后重试 ({retry_count + 1}/{max_retries})...")
                time.sleep(wait_time)
                return self._call_api(prompt, max_tokens, timeout, 
                                    temperature, retry_count + 1, total_items)
            else:
                if self.log:
                    self.log(f"  ⏱️ API超时，已重试{max_retries}次，放弃")
                return None
                
        except requests.exceptions.ConnectionError as e:
            elapsed = time.time() - start_time
            self._record_api_call(elapsed, False)
            
            if retry_count < max_retries:
                wait_time = 5 * (retry_count + 1)
                if self.log:
                    self.log(f"  🔌 连接错误，{wait_time}秒后重试 ({retry_count + 1}/{max_retries})...")
                time.sleep(wait_time)
                return self._call_api(prompt, max_tokens, timeout,
                                    temperature, retry_count + 1, total_items)
            return None
            
        except Exception as e:
            elapsed = time.time() - start_time
            self._record_api_call(elapsed, False)
            if self.log:
                self.log(f"  ⚠️ API 异常({elapsed:.1f}s): {e}")
            return None

    # ── 第一次调用：标题清洗 ─────────────────────────────────────────
    def clean_title(self, raw_title: str, raw_desc: str = "") -> Optional[str]:
        """
        【第一次API调用】在搜索前执行。
        将闲鱼标题+描述清洗为拼多多搜索词，失败返回 None（降级到本地清洗）。
        """
        if not raw_title or len(raw_title) < 3:
            return None

        self._stats["clean_calls"] += 1

        prompt = PROMPT_CLEAN_TITLE.format(raw_title=raw_title, raw_desc=raw_desc)
        
        # 清洗请求：超时35秒，最多重试2次，智能调整
        result = self._call_api(prompt, max_tokens=60, timeout=35, 
                               temperature=0.1, total_items=0)
        
        if result:
            # 只取第一行，去掉可能的引号/标点
            cleaned = result.split('\n')[0].strip().strip('"""\'')[:40]
            if len(cleaned) >= 2:
                self._stats["clean_ok"] += 1
                if self.log:
                    self.log(f"  🤖 [第1次API] 清洗: {raw_title[:28]}... → {cleaned}")
                return cleaned

        self._stats["clean_fail"] += 1
        return None

    def analyze_anchor(self, xianyu_title: str, xianyu_desc: str) -> Optional[Dict]:
        """
        【第1步】锚点识别：理解闲鱼商品的核心交付物/目标用户/核心需求。
        结果供标题路和图搜路共用，每件商品只额外多消耗1次API。
        失败返回None，调用方可降级到直接使用标题。
        """
        if not self.enabled or not self.api_key:
            return None

        prompt = PROMPT_ANALYZE.format(
            xianyu_title=xianyu_title[:120],
            xianyu_desc=(xianyu_desc or '').replace('\n', ' ')[:200],
        )

        raw = self._call_api(prompt, max_tokens=300, timeout=30, temperature=0.1)
        if not raw:
            return None

        try:
            # 清理Markdown代码块标记
            clean = raw.strip()
            clean = re.sub(r'^```json\s*', '', clean)
            clean = re.sub(r'^```\s*', '', clean)
            clean = re.sub(r'\s*```$', '', clean)
            clean = re.sub(r',\s*}', '}', clean)
            # 提取JSON对象
            s = clean.find('{')
            e = clean.rfind('}')
            if s == -1 or e == -1:
                return None
            obj = _json.loads(clean[s:e+1])
            if obj.get('core_item'):
                return obj
        except Exception as ex:
            if self.log:
                self.log(f"  ⚠️ 锚点解析失败: {ex} | 原文: {raw[:150]}")
        return None

    def compare_batch(
        self,
        xianyu_title: str,
        xianyu_desc: str,
        search_keyword: str,
        title_products: List[Dict],
        img_products: List[Dict],
    ) -> Optional[Dict]:
        """
        【v9.0 两阶段版】
        第1步：锚点识别（1次API，标题路+图搜路共用）
        第2步：分路匹配（标题1次 + 图搜1次，最多8件，价格分散采样）
        相比v8.1仅多1次轻量API（约300token），准确率大幅提升。
        任意一路失败都独立降级，不影响另一路。
        """
        if not self.enabled or not self.api_key:
            return None
        if not title_products and not img_products:
            return None

        self._stats["compare_calls"] += 1

        # ── 第1步：锚点识别（1次，两路共用）──────────────────────────────
        if self.log:
            self.log(f"  🔍 [第1步] 识别闲鱼需求锚点...")
        anchor = self.analyze_anchor(xianyu_title, xianyu_desc)
        if anchor:
            if self.log:
                self.log(f"  📦 交付物: {anchor.get('core_item', '')[:30]}"
                         f" | 用户: {anchor.get('target_user', '')[:25]}")
            core_item   = anchor.get('core_item', xianyu_title[:30])
            target_user = anchor.get('target_user', '购买者')
            core_need   = anchor.get('core_need', '满足需求')
        else:
            # 锚点识别失败，降级：用标题作为core_item
            if self.log:
                self.log(f"  ⚠️ 锚点识别失败，降级使用标题作为锚点")
            core_item   = xianyu_title[:50]
            target_user = '购买者'
            core_need   = '满足实际需求'

        # ── 分批配置：每批8件，批次间间隔2秒避免限流 ──────────────────
        BATCH_SIZE      = 8     # 每批最多8件（单次token可控）
        BATCH_INTERVAL  = 2.0   # 批次间等待秒数（防限流）

        def _fmt_list(products: List[Dict], idx_start: int = 1) -> str:
            if not products:
                return "（无采集结果）"
            lines = []
            for i, p in enumerate(products, idx_start):
                price_str = f"¥{p['price']}" if p.get('price') else "价格未知"
                title = p.get('title', '')[:40]
                
                # 调试：打印原始标题
                print(f"【调试】商品{i} 原始标题: {repr(title)}")
                
                # 检查标题里是否有 { 或 }
                if '{' in title or '}' in title:
                    print(f"【调试】⚠️ 标题包含花括号！原始: {repr(title)}")
                
                lines.append(f"{i}.[{price_str}]{title}")
            return '\n'.join(lines)

        def _extract_json_array(raw: str) -> Optional[list]:
            """健壮的JSON数组提取，支持多种格式"""
            if not raw:
                return None
            clean = raw.strip()
            clean = re.sub(r'^```json\s*', '', clean)
            clean = re.sub(r'^```\s*', '', clean)
            clean = re.sub(r'\s*```$', '', clean)
            clean = re.sub(r',\s*]', ']', clean)
            clean = re.sub(r',\s*}', '}', clean)
            # 方法1：直接解析
            try:
                result = _json.loads(clean)
                if isinstance(result, list):
                    return result
            except Exception:
                pass
            # 方法2：逐个数组提取合并（兼容分段输出）
            all_items = []
            depth, start, in_str, escape = 0, -1, False, False
            i = 0
            while i < len(clean):
                ch = clean[i]
                if escape:
                    escape = False; i += 1; continue
                if ch == '\\':
                    escape = True; i += 1; continue
                if ch == '"' and not escape:
                    in_str = not in_str; i += 1; continue
                if not in_str:
                    if ch == '[':
                        if depth == 0: start = i
                        depth += 1
                    elif ch == ']':
                        depth -= 1
                        if depth == 0 and start != -1:
                            try:
                                arr = _json.loads(clean[start:i+1])
                                if isinstance(arr, list):
                                    all_items.extend(arr)
                            except Exception:
                                pass
                            start = -1
                i += 1
            if all_items:
                return all_items
            # 方法3：兼容 {"matches":[...]} 格式
            try:
                s = clean.find('{')
                e = clean.rfind('}')
                if s != -1 and e != -1:
                    obj = _json.loads(clean[s:e+1])
                    for key in ('matches', 'results', 'items'):
                        if key in obj and isinstance(obj[key], list):
                            return obj[key]
            except Exception:
                pass
            return None

        def _call_match_all(products: List[Dict], source_type: str) -> tuple:
           
            """
            全量分批匹配：把 products 按 BATCH_SIZE 切片，每批独立调用API，
            结果按原始 idx 合并，返回 (all_matches, any_ok)。
            批次间等待 BATCH_INTERVAL 秒，防止限流。
            """
            if not products:
                return [], True

            # 切批
            batches = [products[i:i+BATCH_SIZE]
                       for i in range(0, len(products), BATCH_SIZE)]
            total_batches = len(batches)

            all_matches: List[Dict] = []
            any_ok = False

            for batch_no, batch in enumerate(batches, 1):
                idx_start = (batch_no - 1) * BATCH_SIZE + 1
                if self.log:
                    self.log(f"  🤖 [{source_type}] 第{batch_no}/{total_batches}批 "
                             f"({len(batch)}件, idx {idx_start}~{idx_start+len(batch)-1})...")

    

               
                prompt = PROMPT_MATCH.format(
                    core_item=core_item,
                    target_user=target_user,
                    core_need=core_need,
                    source_type=source_type,
                    batch_no=batch_no,
                    batch_total=total_batches,
                    idx_start=idx_start,
                    product_list=_fmt_list(batch, idx_start),
                )
                raw = self._call_api(prompt, max_tokens=2000,
                                     timeout=55, temperature=0.1,
                                     total_items=len(batch))
                if raw:
                    parsed = _extract_json_array(raw)
                    if parsed:
                        any_ok = True
                        all_matches.extend(parsed)
                    else:
                        if self.log:
                            self.log(f"  ⚠️ [{source_type}]第{batch_no}批解析失败 | 原文: {raw[:150]}")
                        # 该批失败：插入占位（让 _enrich 对这批商品降级）
                        for j in range(len(batch)):
                            all_matches.append({'idx': idx_start + j, 'product_type': '_fallback'})
                else:
                    if self.log:
                        self.log(f"  ⚠️ [{source_type}]第{batch_no}批API失败，该批降级")
                    for j in range(len(batch)):
                        all_matches.append({'idx': idx_start + j, 'product_type': '_fallback'})

                # 批次间隔（最后一批不等）
                if batch_no < total_batches:
                    time.sleep(BATCH_INTERVAL)

            return all_matches, any_ok

        # ── 第2步：全量分批匹配 ────────────────────────────────────────────
        n_title = len(title_products)
        n_img   = len(img_products)
        n_title_batches = max(1, (n_title + BATCH_SIZE - 1) // BATCH_SIZE)
        n_img_batches   = max(1, (n_img   + BATCH_SIZE - 1) // BATCH_SIZE)

        if self.log:
            self.log(f"  🤖 [第2步] 标题全量匹配: {n_title}件 → {n_title_batches}批")
        title_matches_raw, title_ok = _call_match_all(title_products, "标题搜索")

        if self.log:
            self.log(f"  🤖 [第2步] 图搜全量匹配: {n_img}件 → {n_img_batches}批")
        img_matches_raw, img_ok = _call_match_all(img_products, "以图搜款")

        if not title_ok and not img_ok:
            self._stats["compare_fail"] += 1
            if self.log:
                self.log("  ⚠️ 两路AI比对均失败，降级到语义/字符相似度")
            return None

        if not title_ok:
            if self.log:
                self.log("  ⚠️ 标题比对失败，标题路降级；图搜路AI结果保留")
        if not img_ok:
            if self.log:
                self.log("  ⚠️ 图搜比对失败，图搜路降级；标题路AI结果保留")

        def _enrich(products: List[Dict], matches: list) -> List[Dict]:
            """
            把 AI 返回的 matches（按 idx 定位）与原始 products 对齐。
            product_type == '_fallback' 表示该批API失败，该条商品降级
            （调用方会用语义相似度兜底，这里返回 is_same=False）。
            """
            idx_map: Dict[int, Dict] = {}
            for m in (matches or []):
                try:
                    idx_val = m.get('idx')
                    if idx_val is not None:
                        idx_map[int(idx_val)] = m
                except (ValueError, TypeError):
                    pass

            result = []
            for i, prod in enumerate(products, 1):
                m  = idx_map.get(i, {})
                pt = m.get('product_type', 'unrelated')

                # _fallback：该批API失败，标记为 ai_failed，调用方降级
                ai_failed = (pt == '_fallback')
                if ai_failed:
                    pt = 'unrelated'

                is_same        = (pt == 'same')
                is_alternative = (pt == 'alternative')
                can_be_supply  = is_same or is_alternative

                score_raw = m.get('score', 0.0)
                try:
                    score_val = max(0.0, min(1.0, float(score_raw)))
                except (ValueError, TypeError):
                    score_val = 0.85 if is_same else (0.7 if is_alternative else 0.1)

                result.append({
                    'title':          prod.get('title', ''),
                    'price':          prod.get('price', 0),
                    'sales':          prod.get('sales', ''),
                    'is_same':        is_same,
                    'is_alternative': is_alternative,
                    'is_accessory':   False,   # v9.0: 去掉accessory分类
                    'product_type':   pt,
                    'score':          score_val,
                    'reason':         m.get('reason', ''),
                    'can_be_supply':  can_be_supply,
                    'ai_failed':      ai_failed,  # 标记该条是否需要降级
                })
            return result

        title_matches = _enrich(title_products, title_matches_raw)
        img_matches   = _enrich(img_products,   img_matches_raw)

        self._stats["compare_ok"] += 1

        if self.log:
            exact_t   = sum(1 for m in title_matches if m['is_same'])
            alt_t     = sum(1 for m in title_matches if m['is_alternative'])
            fallback_t= sum(1 for m in title_matches if m.get('ai_failed'))
            exact_i   = sum(1 for m in img_matches if m['is_same'])
            alt_i     = sum(1 for m in img_matches if m['is_alternative'])
            fallback_i= sum(1 for m in img_matches if m.get('ai_failed'))
            t_status  = "✅AI" if title_ok else "⚠️降级"
            i_status  = "✅AI" if img_ok   else "⚠️降级"
            fb_t = f"/降级{fallback_t}" if fallback_t else ""
            fb_i = f"/降级{fallback_i}" if fallback_i else ""
            self.log(f"  🤖 比对完成: 标题[{t_status}]同款{exact_t}/平替{alt_t}{fb_t}"
                     f" | 图搜[{i_status}]同款{exact_i}/平替{alt_i}{fb_i}")

        return {
            'title_matches': title_matches,
            'img_matches': img_matches,
            'title_ai_ok': title_ok,
            'img_ai_ok': img_ok,
        }


# ══════════════════════════════════════════════════════════════════════════
# 标题清洗（本地版本 - 从外部文件读取排除词）
# ══════════════════════════════════════════════════════════════════════════

_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_EXCLUDE_FILE = os.path.join(_PROJECT_DIR, "闲鱼标题剔除词.txt")

def load_exclude_words(file_path: str = None) -> Set[str]:
    if file_path is None:
        file_path = DEFAULT_EXCLUDE_FILE

    default_set = {
        '二手', '闲置', '全新', '转让', '低价', '自用', '包邮', '顺丰',
        '闲鱼', '到手', '原价', '可刀', '议价', '已出', '清仓', '几乎全新',
        '急出', '忍痛出', '含泪出', '分手', '怀孕', '生娃', '没时间',
        '毕业', '带不走', '退圈', '随缘', '有缘人', '赔钱', '回笼资金',
        '撤柜', '仓库到期', '特价', '亏本', '白菜价', '骨折价', '捡漏',
        '现货', '秒发', '手慢无', '先到先得', '闭眼入', '神器', '天花板',
        '爆款', '网红', '必入', '超值', '绝了', '全国包邮', '顺丰包邮',
        '支持验货', '支持退换', '七天无理由', '送配件', '赠送', '全套',
        '质保', '保修', '假一罚十', '不正包退', '代工厂', '外贸', '剪标',
    }

    if not os.path.exists(file_path):
        print(f"⚠️ 排除词文件不存在: {file_path}")
        print(f"   使用默认内置排除词集，共 {len(default_set)} 个")
        return default_set

    words_set = set()
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                words_set.add(line)
    return words_set


_EXCLUDE_WORDS = None

def get_exclude_words() -> Set[str]:
    global _EXCLUDE_WORDS
    if _EXCLUDE_WORDS is None:
        _EXCLUDE_WORDS = load_exclude_words()
    return _EXCLUDE_WORDS


def refresh_exclude_words() -> None:
    global _EXCLUDE_WORDS
    _EXCLUDE_WORDS = load_exclude_words()
    print("🔄 排除词已刷新")


def clean_title_local(raw_title: str) -> str:
    exclude_words = get_exclude_words()
    cleaned = raw_title
    for w in exclude_words:
        cleaned = cleaned.replace(w, ' ')
    cleaned = re.sub(r'[^\u4e00-\u9fffA-Za-z0-9]', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    if len(cleaned) > 20:
        cleaned = cleaned[:20]
    return cleaned if len(cleaned) >= 3 else raw_title[:15]


_AI_CLEANER: Optional[TitleCleanerAI] = None

def get_ai_cleaner() -> TitleCleanerAI:
    global _AI_CLEANER
    if _AI_CLEANER is None:
        _AI_CLEANER = TitleCleanerAI()
    return _AI_CLEANER


def set_ai_api_key(api_key: str, log_cb: Callable = None):
    cleaner = get_ai_cleaner()
    if log_cb:
        cleaner.log = log_cb
    cleaner.set_api_key(api_key)


def clean_title_with_ai_fallback(raw_title: str, use_ai: bool = True) -> str:
    if not raw_title:
        return ""
    if use_ai:
        cleaner = get_ai_cleaner()
        ai_result = cleaner.clean_title(raw_title)
        if ai_result:
            return ai_result
    return clean_title_local(raw_title)


# ══════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════

def rand_sleep(lo: float, hi: float) -> None:
    time.sleep(random.uniform(lo, hi))


def extract_products_from_xml(xml_content: str) -> List[Dict]:
    products, seen = [], set()
    lines = xml_content.split('\n')
    title_indices = [i for i, ln in enumerate(lines)
                     if 'com.xunmeng.pinduoduo:id/tv_title' in ln and 'resource-id' in ln]

    for idx, tln in enumerate(title_indices):
        line = lines[tln]
        m = re.search(r'content-desc="([^"]*)"', line)
        title = m.group(1).replace('&#10;', ' ').strip() if m else ''
        if not title:
            m = re.search(r'text="([^"]*)"', line)
            title = m.group(1).strip() if m else ''
        if not title or len(title) < 3 or title in seen:
            continue

        end = title_indices[idx + 1] if idx + 1 < len(title_indices) else len(lines)
        card = '\n'.join(lines[tln:end])

        price = None
        for pat in [r'text="(\d+\.\d+)"',
                    r'content-desc="[^"]*[¥￥](\d+\.?\d*)[^"]*"',
                    r'[¥￥]\s*(\d+\.?\d*)']:
            m = re.search(pat, card)
            if m:
                price = float(m.group(1))
                break

        sales = ''
        m = re.search(r'(已拼|已售)[\s]*([\d.]+[万千]?\+?件?)', card)
        if m:
            sales = m.group(0)

        if price:
            seen.add(title)
            products.append({'title': title, 'price': price, 'sales': sales})

    return products


# ══════════════════════════════════════════════════════════════════════════
# 拟人化操作
# ══════════════════════════════════════════════════════════════════════════

class HumanBehavior:
    def __init__(self, device, screen_w: int, screen_h: int, log: Callable):
        self.d = device
        self.w = screen_w
        self.h = screen_h
        self.log = log

    def type_text(self, field, text: str):
        field.clear_text()
        rand_sleep(0.2, 0.6)
        field.click()
        rand_sleep(0.2, 0.5)

        i = 0
        while i < len(text):
            if i > 2 and random.random() < 0.06:
                self.d.send_keys(random.choice('的了是在'), clear=False)
                rand_sleep(0.15, 0.35)
                self.d.press("del")
                rand_sleep(0.1, 0.25)

            chunk_len = random.randint(1, min(3, len(text) - i))
            chunk = text[i:i + chunk_len]
            self.d.send_keys(chunk, clear=False)
            if any('\u4e00' <= c <= '\u9fff' for c in chunk):
                time.sleep(random.uniform(0.1, 0.28))
            else:
                time.sleep(random.uniform(0.05, 0.15))
            i += chunk_len

    def scroll_down(self, intensity: str = 'normal'):
        cx = self.w // 2 + random.randint(-40, 40)
        ratio = {'short': random.uniform(0.18, 0.30),
                 'long': random.uniform(0.55, 0.72)}.get(intensity, random.uniform(0.32, 0.52))
        sy = int(self.h * random.uniform(0.60, 0.75))
        ey = max(100, sy - int(self.h * ratio))
        ex = cx + random.randint(-25, 25)
        dur = random.choice([random.uniform(0.15, 0.25),
                             random.uniform(0.28, 0.50),
                             random.uniform(0.55, 0.85)])
        self.d.swipe(cx, sy, ex, ey, duration=dur)

        if random.random() < 0.05:
            rand_sleep(0.4, 1.0)
        if random.random() < 0.15:
            rand_sleep(0.3, 0.8)
            rb = random.randint(60, 180)
            self.d.swipe(cx, ey, cx, ey + rb, duration=random.uniform(0.12, 0.25))
            rand_sleep(0.2, 0.6)

        wait = max(0.8, min(5.0, random.gauss(2.0, 0.8)))
        time.sleep(wait)

    def random_idle(self):
        if random.random() > 0.05:
            return
        rand_sleep(1.5, 4.0)


# ══════════════════════════════════════════════════════════════════════════
# 拼多多控制器
# ══════════════════════════════════════════════════════════════════════════

class PinduoduoMobileController:
    def __init__(self, log_cb: Callable = None):
        self.log = log_cb or print
        self.device = None
        self._connected = False
        self.screen_width = 1080
        self.screen_height = 2400
        self.hb: Optional[HumanBehavior] = None

    def connect(self, serial: str = None) -> bool:
        if not U2_OK:
            self.log("❌ uiautomator2 未安装")
            return False
        try:
            if serial:
                self.log(f"📱 连接设备: {serial}...")
                self.device = u2.connect(serial)
            else:
                self.log("📱 自动连接设备...")
                self.device = u2.connect()
            try:
                info = self.device.info
                self.screen_width = info.get('displayWidth', 1080)
                self.screen_height = info.get('displayHeight', 2400)
                self.log(f"✅ 连接成功: {info.get('productName', 'Unknown')} "
                         f"{self.screen_width}x{self.screen_height}")
                self.hb = HumanBehavior(self.device, self.screen_width, self.screen_height, self.log)
                self._connected = True
                return True
            except (AttributeError, TypeError) as e2:
                self.log(f"❌ 设备信息获取失败 (可能多设备未指定serial): {e2}")
                self.device = None
                return False
        except Exception as e:
            self.log(f"❌ 连接失败: {e}")
            return False

    def is_connected(self) -> bool:
        return self._connected and self.device is not None

    def launch_pinduoduo(self) -> bool:
        try:
            self.log("🚀 启动拼多多...")
            self.device.app_start("com.xunmeng.pinduoduo")
            rand_sleep(2.5, 4.5)
            self.log("✅ 拼多多已启动")
            return True
        except Exception as e:
            self.log(f"❌ 启动失败: {e}")
            return False

    def go_to_home(self):
        try:
            for _ in range(3):
                self.device.press("back")
                rand_sleep(0.4, 0.9)
            rand_sleep(0.5, 1.2)
        except Exception:
            pass

    def take_screenshot(self, filename: str = None) -> Optional[str]:
        """截图并保存到本地"""
        try:
            if not self.device:
                return None
            if filename is None:
                filename = os.path.expanduser(
                    f"~/Desktop/pdd_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
            self.device.screenshot(filename)
            self.log(f"📸 截图已保存: {filename}")
            return filename
        except Exception as e:
            self.log(f"❌ 截图失败: {e}")
            return None

    def _click_search_box(self) -> bool:
        for sel in [{"description": "搜索"}, {"text": "搜索"},
                    {"resourceId": "com.xunmeng.pinduoduo:id/search_hint"},
                    {"resourceId": "com.xunmeng.pinduoduo:id/tv_search_hint"}]:
            elem = self.device(**sel)
            if elem.exists:
                rand_sleep(0.3, 0.9)
                elem.click()
                rand_sleep(0.4, 0.9)
                return True
        return False

    def search_and_collect(self, keyword: str, scroll_pages: int = 5,
                           max_items: int = 20) -> List[Dict]:
        self.log(f"  🔍 搜索: 【{keyword}】")

        if not self._click_search_box():
            self.log("    ✗ 找不到搜索框")
            return []

        input_field = self.device(className="android.widget.EditText")
        if not input_field.exists:
            self.log("    ✗ 找不到输入框")
            return []

        self.hb.type_text(input_field, keyword)
        rand_sleep(0.5, 1.5)

        search_btn = self.device(text="搜索", className="android.widget.TextView")
        if not search_btn.exists:
            search_btn = self.device(text="搜索")
        if search_btn.exists:
            search_btn.click()
        else:
            self.device.press("search")

        wait = max(2.5, min(7.0, random.gauss(4.0, 0.8)))
        self.log(f"    ⏳ 等待加载 {wait:.1f}s...")
        time.sleep(wait)

        all_products: List[Dict] = []
        seen: Set[str] = set()

        for page in range(scroll_pages):
            if len(all_products) >= max_items:
                break

            self.log(f"    📄 第 {page+1}/{scroll_pages} 页...")

            try:
                xml = self.device.dump_hierarchy()
                new = [p for p in extract_products_from_xml(xml) if p['title'] not in seen]
            except Exception as e:
                self.log(f"    ⚠️ 获取页面失败: {e}")
                new = []

            for p in new:
                seen.add(p['title'])
                all_products.append(p)
            self.log(f"      新增{len(new)}件，累计{len(all_products)}件")

            if len(all_products) >= max_items:
                break
            if page < scroll_pages - 1:
                self.hb.random_idle()
                self.hb.scroll_down(random.choice(['short', 'normal', 'normal', 'long']))

        return all_products[:max_items]

    def back_to_search_start(self):
        try:
            for _ in range(2):
                self.device.press("back")
                rand_sleep(0.4, 0.8)
        except Exception:
            pass

    def image_search_from_result_page(self, pic_url: str, scroll_pages: int = 3,
                                       max_items: int = 20) -> List[Dict]:
        if not self.device:
            self.log("  [图搜] ❌ 设备未连接")
            return []

        if not pic_url:
            self.log("  [图搜] ⚠️ 无图片URL，跳过图搜")
            return []

        w = self.screen_width
        h = self.screen_height

        local_tmp = os.path.join(tempfile.gettempdir(), "pdd_supply_imgsearch.jpg")
        try:
            self.log(f"  [图搜] 📥 下载图片...")
            resp = requests.get(pic_url, timeout=15, headers={
                'User-Agent': 'Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36'
            })
            if resp.status_code != 200 or len(resp.content) < 1000:
                self.log(f"  [图搜] ❌ 图片下载失败 ({resp.status_code})")
                return []
            with open(local_tmp, 'wb') as f:
                f.write(resp.content)
        except Exception as e:
            self.log(f"  [图搜] ❌ 下载异常: {e}")
            return []

        phone_img = "/sdcard/pdd_supply_imgsearch.jpg"
        try:
            self.log(f"  [图搜] 📲 推送图片到手机...")
            result = subprocess.run(
                ["adb", "push", local_tmp, phone_img],
                timeout=20, capture_output=True, text=True,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
            )
            if result.returncode != 0:
                self.log(f"  [图搜] ❌ adb push 失败: {result.stderr[:80]}")
                return []
            self.device.shell(f"am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE "
                              f"-d file://{phone_img}")
            rand_sleep(1.0, 2.0)
        except Exception as e:
            self.log(f"  [图搜] ❌ 推送异常: {e}")
            return []
        finally:
            try:
                os.remove(local_tmp)
            except Exception:
                pass

        self.log("  [图搜] 📷 尝试点击拍照搜索按钮...")
        camera = None

        try:
            self.log("  [图搜]    方式1: 搜索结果页下滑后查找...")
            self.device.swipe(self.screen_width // 2, 200,
                            self.screen_width // 2, 500, duration=0.02)
            time.sleep(0.2)

            camera = self.device(description="拍照搜索")
            if not camera.exists(timeout=1):
                camera = self.device(text="拍照搜索")

            if camera.exists(timeout=1):
                self.log("  [图搜]    ✅ 搜索结果页找到拍照搜索按钮")
        except Exception as e:
            self.log(f"  [图搜]    ⚠️ 方式1异常: {e}")

        if not camera or not camera.exists:
            self.log("  [图搜]    方式2: 返回搜索页查找拍照搜索按钮...")
            try:
                self.device.press("back")
                rand_sleep(0.5, 1.0)

                camera = self.device(description="拍照搜索")
                if not camera.exists(timeout=1.5):
                    camera = self.device(text="拍照搜索")

                if camera.exists(timeout=1):
                    self.log("  [图搜]    ✅ 搜索页找到拍照搜索按钮")
                else:
                    self.log("  [图搜]    ❌ 搜索页也未找到拍照搜索按钮")
            except Exception as e:
                self.log(f"  [图搜]    ⚠️ 方式2异常: {e}")

        if camera and camera.exists:
            camera.click()
            self.log("  [图搜]    ✅ 已点击拍照搜索按钮")
            rand_sleep(1.0, 1.8)
        else:
            self.log("  [图搜] 🔧 坐标兜底点击拍照搜索")
            click_x = self.screen_width - 140 if self.screen_width else 900
            click_y = 150
            self.device.click(click_x, click_y)
            rand_sleep(1.0, 1.8)

        album_clicked = False
        for kw in ["我的相册", "相册", "从相册选图", "图库", "图片"]:
            elem = self.device(textContains=kw)
            if elem.exists(timeout=2):
                elem.click()
                self.log(f"  [图搜] ✅ 点击相册: {kw}")
                album_clicked = True
                rand_sleep(1.5, 2.5)
                break
        if not album_clicked:
            self.log("  [图搜] ⚠️ 未找到相册按钮，继续尝试...")
            rand_sleep(1.5, 2.0)

        self.log("  [图搜] 🖼️ 选择图片...")
        rv = self.device(resourceId="com.xunmeng.pinduoduo:id/pvd",
                         className="androidx.recyclerview.widget.RecyclerView")
        if not rv.exists(timeout=2):
            rv = self.device(resourceId="com.xunmeng.pinduoduo:id/pdd",
                             className="android.support.v7.widget.RecyclerView")
        if rv.exists(timeout=2):
            bounds = rv.info.get('bounds', {})
            if bounds.get('top', 0) > 0:
                x = bounds['left'] + 3 + 133
                y = bounds['top'] + 3 + 134
                self.device.click(x, y)
                self.log(f"  [图搜] ✅ 动态选图: ({x}, {y})")
                rand_sleep(0.8, 1.2)
        else:
            self.device.click(136, int(h * 0.83))
            self.log(f"  [图搜] 🔧 坐标兜底选图")
            rand_sleep(0.8, 1.2)

        for kw in ["使用", "确定", "确认", "搜索相似", "搜图片同款"]:
            btn = self.device(text=kw)
            if btn.exists(timeout=1.5):
                btn.click()
                self.log(f"  [图搜] ✅ 确认: {kw}")
                rand_sleep(0.8, 1.2)
                break

        self.log("  [图搜] ⏳ 等待图搜结果...")
        for _ in range(15):
            xml_check = self.device.dump_hierarchy()
            if 'tv_title' in xml_check or '搜图片同款' in xml_check:
                self.log("  [图搜] ✅ 图搜结果已加载")
                break
            time.sleep(1.0)

        all_products: List[Dict] = []
        seen_titles: Set[str] = set()

        for page in range(scroll_pages):
            if len(all_products) >= max_items:
                break
            self.log(f"  [图搜] 📄 第 {page+1}/{scroll_pages} 页...")
            try:
                xml = self.device.dump_hierarchy()
                new_items = _parse_imgsearch_xml(xml)
                added = 0
                for p in new_items:
                    if p['title'] not in seen_titles:
                        seen_titles.add(p['title'])
                        all_products.append(p)
                        added += 1
                self.log(f"  [图搜]    新增{added}件，累计{len(all_products)}件")
            except Exception as e:
                self.log(f"  [图搜]    ⚠️ 解析失败: {e}")

            if len(all_products) >= max_items:
                break
            if page < scroll_pages - 1:
                cx = w // 2
                sy = int(h * 0.72)
                ey = int(h * 0.28)
                self.device.swipe(cx, sy, cx, ey, duration=0.4)
                time.sleep(2.0)

        try:
            self.device.shell(f"rm -f {phone_img}")
        except Exception:
            pass

        self.log(f"  [图搜] ✅ 图搜完成，共{len(all_products)}件")
        return all_products[:max_items]


def _parse_imgsearch_xml(xml: str) -> list:
    products = []
    seen = set()
    price_nodes = re.findall(
        r'text=\"(\d+\.?\d+)\" resource-id=\"com\.xunmeng\.pinduoduo:id/pdd\"[^>]*'
        r'bounds=\"\[(\d+),(\d+)\]\[(\d+),(\d+)\]\"',
        xml
    )
    price_list = [(float(v), int(x1), int(y1))
                  for v, x1, y1, x2, y2 in price_nodes if 5.0 < float(v) < 99999.0]
    title_nodes = re.findall(
        r'resource-id=\"com\.xunmeng\.pinduoduo:id/tv_title\"[^>]*'
        r'content-desc=\"([^\"]+)\"[^>]*'
        r'bounds=\"\[(\d+),(\d+)\]\[(\d+),(\d+)\]\"',
        xml
    )
    for raw_title, x1, y1, x2, y2 in title_nodes:
        title = raw_title.replace('&#10;', '').strip()
        if not title or title in seen:
            continue
        seen.add(title)
        tx1, ty2 = int(x1), int(y2)
        t_side = 'left' if tx1 < 537 else 'right'
        best_price, best_dist = None, 9999
        for pval, px1, py1 in price_list:
            p_side = 'left' if px1 < 537 else 'right'
            dist = py1 - ty2
            if p_side == t_side and 30 <= dist <= 350 and dist < best_dist:
                best_dist = dist
                best_price = pval
        products.append({'title': title, 'price': best_price, 'sales': ''})
    return products


# ══════════════════════════════════════════════════════════════════════════
# 语义匹配
# ══════════════════════════════════════════════════════════════════════════

class SameProductMatcher:
    _model = None
    _lock = threading.Lock()
    _loaded = False
    _error = None

    def __init__(self, threshold: float = 0.8):
        self.threshold = max(0.0, min(1.0, threshold))

    @classmethod
    def load_model(cls, log_cb: Callable = None) -> bool:
        if cls._loaded:
            return True
        with cls._lock:
            if cls._loaded:
                return True
            global TEXT2VEC_OK
            if TEXT2VEC_OK is None:
                try:
                    from text2vec import Similarity as _Sim
                    TEXT2VEC_OK = True
                except ImportError:
                    TEXT2VEC_OK = False
            if not TEXT2VEC_OK:
                cls._error = "text2vec 未安装"
                if log_cb:
                    log_cb("⚠️ text2vec 未安装，将使用字符相似度兜底")
                return False
            try:
                from text2vec import Similarity
                if log_cb:
                    log_cb("📦 后台加载语义模型...")
                # Use HF mirror for China + local cache
                import os
                os.environ.setdefault("HF_ENDPOINT", "https://hf-mirror.com")
                cls._model = Similarity()
                cls._loaded = True
                if log_cb:
                    log_cb("✅ 语义模型加载完成")
                return True
            except Exception as e:
                cls._error = str(e)
                if log_cb:
                    log_cb(f"⚠️ 语义模型不可用: {e}，将使用字符相似度兜底")
                return False

    def is_same_product(self, source: str, pdd_title: str,
                        threshold: float = None) -> Tuple[bool, float]:
        th = threshold if threshold is not None else self.threshold
        if not source or not pdd_title:
            return False, 0.0
        try:
            score = self.__class__._model.get_score(source, pdd_title) if self.__class__._model else self._char_sim(source, pdd_title)
        except Exception:
            score = self._char_sim(source, pdd_title)
        return score >= th, round(score, 4)

    @staticmethod
    def _char_sim(a: str, b: str) -> float:
        a = re.sub(r'[^\w\u4e00-\u9fff]', '', a.lower())
        b = re.sub(r'[^\w\u4e00-\u9fff]', '', b.lower())
        if not a or not b:
            return 0.0
        sa, sb = set(a), set(b)
        return len(sa & sb) / len(sa | sb)


# ══════════════════════════════════════════════════════════════════════════
# 利润分析（优化版，支持平替评分）
# ══════════════════════════════════════════════════════════════════════════

class ProfitAnalyzer:
    @staticmethod
    def _parse_sales(text: str) -> int:
        if not text:
            return 0
        m = re.search(r'([\d.]+)\s*([万千百]?)', text)
        if not m:
            return 0
        n = float(m.group(1))
        return int(n * {'万': 10000, '千': 1000, '百': 100}.get(m.group(2), 1))

    def analyze(self, xianyu_price: float, items: list) -> list:
        if not items:
            return []
        # 修复：过滤掉价格为 None 的商品，避免排序异常
        valid_items = [it for it in items if it.get('pdd_price_yuan') is not None]
        if not valid_items:
            return []
        
        prices = [it.get('pdd_price_yuan', 99999) for it in valid_items]
        min_p, max_p = min(prices), max(prices)

        results = []
        for it in valid_items:
            p = it.get('pdd_price_yuan', 0)
            sim = it.get('sim_score', 0)
            sales = self._parse_sales(it.get('sales_tip', ''))
            profit = xianyu_price - p
            rate = (profit / p * 100) if p > 0 else 0

            match_level = it.get('match_level', 'mismatch')
            is_alternative = it.get('is_alternative', False)
            reason = it.get('match_reason', '')
            match_src = it.get('match_src', '')

            if match_level == 'exact':
                profit_weight = 40
            elif match_level == 'good':
                profit_weight = 35
            else:
                profit_weight = 30

            sp = profit_weight if rate >= 50 else (
                30 if rate >= 30 else (20 if rate >= 20 else (10 if rate >= 10 else 0)))

            alternative_bonus = 5 if is_alternative and rate >= 20 else 0

            ss = 30 if sales >= 100000 else (22 if sales >= 10000 else (
                15 if sales >= 1000 else (8 if sales >= 100 else 0)))
            spr = int(20 * (1 - (p - min_p) / (max_p - min_p))) if max_p > min_p else 20
            total = sp + ss + spr + int(sim * 10) + alternative_bonus

            if match_level == 'exact':
                recommend_tag = '⭐ 推荐' if total >= 50 else ('一般' if total >= 30 else '不推荐')
            elif is_alternative:
                recommend_tag = '🔄 平替推荐' if total >= 45 else ('⚠️ 平替参考' if total >= 30 else '不推荐')
            else:
                recommend_tag = '一般' if total >= 30 else '不推荐'

            results.append({
                **it,
                '拼多多进价(元)': round(p, 2),
                '预估利润(元)': round(profit, 2),
                '利润率(%)': round(rate, 1),
                '货源评分': total,
                '是否推荐货源': recommend_tag,
                '匹配类型': '同款' if match_level == 'exact' else ('平替' if is_alternative else '部分匹配'),
                '匹配说明': reason,
                '比对方式': match_src,
            })

        results.sort(key=lambda x: (0 if x.get('is_exact_match', False) else 1, -x['货源评分']))
        return results


# ══════════════════════════════════════════════════════════════════════════
# 四象限综合评估
# ══════════════════════════════════════════════════════════════════════════

def evaluate_supply_quadrant(title_items: list, img_items: list,
                              title_has_data: bool = True,
                              img_has_data: bool = True) -> Dict:
    def _best_profit(items) -> Optional[float]:
        if not items:
            return None
        return items[0].get('预估利润(元)', None)

    def _best_price(items) -> Optional[float]:
        if not items:
            return None
        return items[0].get('拼多多进价(元)', None)

    title_profit = _best_profit(title_items) if title_has_data else None
    img_profit = _best_profit(img_items) if img_has_data else None
    title_price = _best_price(title_items) if title_has_data else None
    img_price = _best_price(img_items) if img_has_data else None

    title_pos = title_profit is not None and title_profit > 0
    img_pos = img_profit is not None and img_profit > 0
    title_has_match = title_profit is not None
    img_has_match = img_profit is not None

    if not title_has_match and not img_has_match:
        quadrant = 'Q5'
        quadrant_label = '无有效数据'
        quadrant_emoji = '⛔'
        final_profit = None
        final_price = None
        final_source = 'none'
        recommendation = '两种方式均未找到同款/平替货源，跳过'
    elif title_has_match and img_has_match:
        if title_pos and img_pos:
            quadrant = 'Q1'
            quadrant_label = '双正利润'
            quadrant_emoji = '⭐'
            if title_profit >= img_profit:
                final_profit = title_profit
                final_price = title_price
                final_source = 'title'
            else:
                final_profit = img_profit
                final_price = img_price
                final_source = 'img'
            recommendation = f'最优货源 强推荐（标题¥{title_profit:.1f} / 图搜¥{img_profit:.1f}）'
        elif img_pos and not title_pos:
            quadrant = 'Q2'
            quadrant_label = '图搜优先'
            quadrant_emoji = '📷'
            final_profit = img_profit
            final_price = img_price
            final_source = 'img'
            recommendation = f'以图搜款货源有利润¥{img_profit:.1f}，优先选图搜结果'
        elif title_pos and not img_pos:
            quadrant = 'Q3'
            quadrant_label = '标题参考'
            quadrant_emoji = '📝'
            final_profit = title_profit
            final_price = title_price
            final_source = 'title'
            recommendation = f'标题搜索有利润¥{title_profit:.1f}，图搜同款偏贵，谨慎参考'
        else:
            quadrant = 'Q4'
            quadrant_label = '双负利润'
            quadrant_emoji = '❌'
            final_profit = max(title_profit, img_profit) if title_profit and img_profit else (title_profit or img_profit)
            final_price = title_price if title_profit > img_profit else img_price
            final_source = 'none'
            recommendation = '两种方式均亏损，不推荐'
    elif title_has_match and not img_has_match:
        if title_pos:
            quadrant = 'Q3'
            quadrant_label = '仅标题有数据·正'
            quadrant_emoji = '📝'
            final_profit = title_profit
            final_price = title_price
            final_source = 'title'
            recommendation = f'图搜无数据，标题货源有利润¥{title_profit:.1f}，参考'
        else:
            quadrant = 'Q4'
            quadrant_label = '仅标题有数据·负'
            quadrant_emoji = '❌'
            final_profit = title_profit
            final_price = title_price
            final_source = 'none'
            recommendation = '图搜无数据，标题货源亏损，不推荐'
    else:
        if img_pos:
            quadrant = 'Q2'
            quadrant_label = '仅图搜有数据·正'
            quadrant_emoji = '📷'
            final_profit = img_profit
            final_price = img_price
            final_source = 'img'
            recommendation = f'标题搜无数据，图搜货源有利润¥{img_profit:.1f}'
        else:
            quadrant = 'Q4'
            quadrant_label = '仅图搜有数据·负'
            quadrant_emoji = '❌'
            final_profit = img_profit
            final_price = img_price
            final_source = 'none'
            recommendation = '标题搜无数据，图搜货源亏损，不推荐'

    return {
        'quadrant': quadrant,
        'quadrant_label': quadrant_label,
        'quadrant_emoji': quadrant_emoji,
        'title_profit': round(title_profit, 2) if title_profit is not None else None,
        'img_profit': round(img_profit, 2) if img_profit is not None else None,
        'final_profit': round(final_profit, 2) if final_profit is not None else None,
        'final_price': round(final_price, 2) if final_price is not None else None,
        'final_source': final_source,
        'recommendation': recommendation,
    }


# ══════════════════════════════════════════════════════════════════════════
# 调度器（增强版：风控自动暂停 + 截图 + 发送文件）
# ══════════════════════════════════════════════════════════════════════════

class MobileSupplyScheduler:
    def __init__(self, task_queue: queue.Queue,
                 result_cb: Callable, log_cb: Callable,
                 countdown_cb: Callable = None,
                 score_threshold: int = 80, sim_threshold: float = 0.8,
                 scroll_pages: int = 5, max_items: int = 20,
                 img_scroll_pages: int = 3,
                 use_img_search: bool = True,
                 use_ai_compare: bool = True,
                 delay_between_products: int = 8,
                 pause_every: int = 5, pause_duration: int = 60,
                 use_ai_clean: bool = True,
                 cache_clear_interval_min: int = 120,
                 empty_threshold: int = 3):
        self.task_queue = task_queue
        self.push_result = result_cb
        self.log = log_cb
        self.countdown_cb = countdown_cb
        self.score_threshold = score_threshold
        self.sim_threshold = sim_threshold
        self.scroll_pages = scroll_pages
        self.max_items = max_items
        self.img_scroll_pages = img_scroll_pages
        self.use_img_search = use_img_search
        self.use_ai_compare = use_ai_compare
        self.delay_between_products = delay_between_products
        self.pause_every = pause_every
        self.pause_duration = pause_duration
        self.use_ai_clean = use_ai_clean
        self.cache_clear_interval_min = cache_clear_interval_min
        self._has_processed = False
        self._auto_stop_when_empty = False

        self.matcher = SameProductMatcher(threshold=sim_threshold)
        self.analyzer = ProfitAnalyzer()
        self.controller: Optional[PinduoduoMobileController] = None

        self._running = False
        self._paused = False  # 新增：风控暂停标志
        self._pause_requested = False  # 新增：请求暂停标志
        self._thread: Optional[threading.Thread] = None
        self._processed_count = 0
        self.all_results: list = []
        self._pushed_profit_ids: Set[int] = set()  # 记录已推送的正利润商品ID（用于去重）
        self._wechat_webhook: Optional[str] = None

        self._last_cache_clear_time: float = time.time()
        self._consecutive_empty_count = 0
        self._empty_threshold = empty_threshold
        self._empty_alert_sent = False
        # ── 需求3：双阶段风控 ──────────────────────────────────────────────
        # 第一阶段：触发阈值 → 立即清缓存+重启，重置计数，等待固定时间
        # 第二阶段：清缓存后再次触发阈值 → 才触发风控告警+暂停
        self._cache_cleared_for_empty = False   # 是否已做过一次空采集缓存清理
        self._post_cache_empty_count = 0        # 清缓存后的空采集计数
        self._post_cache_wait_sec = 30          # 清缓存后等待固定时间再采集（秒）

    def start_processing(self, auto_stop_when_empty: bool = False):
        if not self.controller:
            self.log("❌ 请先连接手机")
            return

        if self._running:
            self.log("⚠️ 队列已在处理中")
            return

        self._auto_stop_when_empty = auto_stop_when_empty
        self._running = True
        self._paused = False
        self._pause_requested = False
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        self.log("🚀 开始处理队列...")

    def resume(self):
        """恢复队列处理（风控解除后调用）"""
        if not self._running:
            self.log("⚠️ 队列未运行，请先点击开始处理")
            return
        if self._paused:
            self._paused = False
            self._pause_requested = False
            self.log("▶ 已恢复队列处理，继续处理下一个商品...")
        else:
            self.log("ℹ️ 队列未处于暂停状态")

    def is_paused(self) -> bool:
        return self._paused

    def stop(self, send_report: bool = True):
        self._running = False
        self._paused = False

    def close(self):
        self._running = False
        self._paused = False

    def _emit_countdown(self, remaining: int):
        if self.countdown_cb:
            try:
                self.countdown_cb(remaining)
            except Exception:
                pass

    def _clear_cache_and_restart(self):
        if not self.controller or not self.controller.is_connected():
            self.log("  ⚠️ 设备未连接，跳过缓存清理")
            return

        d = self.controller.device
        cleared = False

        try:
            self.log("  🧹 Step1: 返回桌面...")
            d.press("home")
            time.sleep(1.5)

            self.log("  🧹 Step2: 打开最近任务...")
            d.press("recent")
            time.sleep(2.0)

            try:
                clear_elem = d(resourceId="com.miui.home:id/clearAnimView")
                if clear_elem.exists(timeout=2):
                    self.log("    ✅ 找到清理按钮，执行点击...")
                    clear_elem.click()
                    time.sleep(0.8)
                    cleared = True
            except Exception as e:
                self.log(f"    ⚠️ 清理失败: {e}")

            if not cleared:
                self.log("    ⚠️ 未找到清理按钮，降级：adb force-stop PDD")
                subprocess.run(
                    ["adb", "shell", "am", "force-stop", "com.xunmeng.pinduoduo"],
                    timeout=10, capture_output=True,
                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                )
                time.sleep(1.5)

            self.log("  🧹 Step3: 返回桌面...")
            d.press("home")
            time.sleep(1.5)

            self.log("  🧹 Step4: 重新启动拼多多...")
            self.controller.launch_pinduoduo()
            time.sleep(5)

            self.log("  ✅ 缓存清理完成，拼多多已就绪")
        except Exception as e:
            self.log(f"  ❌ 缓存清理异常: {e}")

    def _trigger_risk_control_pause(self, reason: str = "连续多次采集0件"):
        """触发风控暂停：截图 + 发送文件 + 暂停队列"""
        self.log(f"\n🚨 【风控告警】{reason}")
        self.log(f"🚨 已暂停队列处理，等待人工介入")
        self.log(f"🚨 请检查网络/账号状态后，点击「继续处理」按钮恢复")

        # 1. 自动截图
        screenshot_path = None
        if self.controller and self.controller.is_connected():
            screenshot_path = self.controller.take_screenshot()
        else:
            self.log("  ⚠️ 无法截图：设备未连接")

        # 2. 导出截止到风控前已收集的正利润货源
        profit_records = [r for r in self.all_results if r.get('final_profit') and r['final_profit'] > 0]
        if profit_records:
            excel_path = os.path.expanduser(
                f"~/Desktop/拼多多货源_风控前正利润_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            export_supply_to_excel(profit_records, excel_path, self.log, profit_only=True)
            self.log(f"📁 已导出风控前正利润货源: {excel_path}")
        else:
            excel_path = None
            self.log("📁 暂无正利润货源可导出")

        # 3. 发送企业微信通知（如果配置了 webhook）
        if self._wechat_webhook:
            self._send_risk_alert_via_wechat(reason, screenshot_path, excel_path)

        # 4. 暂停队列
        self._paused = True
        self._pause_requested = True

    def _send_risk_alert_via_wechat(self, reason: str, screenshot_path: Optional[str], excel_path: Optional[str]):
        """发送风控告警到企业微信"""
        if not self._wechat_webhook:
            return

        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 1. 发送文本消息
            content = f"""## 🚨 货源查找风控告警

    **时间：** {timestamp}
    **原因：** {reason}
    **状态：** 队列已暂停，等待人工介入

    **操作建议：**
    1. 检查拼多多账号状态
    2. 检查网络连接
    3. 点击GUI中的「继续处理」按钮恢复队列
    """
            data = {"msgtype": "markdown", "markdown": {"content": content}}
            requests.post(self._wechat_webhook, json=data, timeout=10)
            self.log("  📤 告警文本已发送")

            # 辅助函数：上传文件获取 media_id
            def upload_file(file_path: str, file_type: str = "file") -> Optional[str]:
                """上传文件到企业微信，返回 media_id"""
                if not os.path.exists(file_path):
                    return None
                
                # 构建正确的上传 URL
                # webhook 格式: https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=xxx
                # 上传 URL 格式: https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=xxx&type=file
                base_url = self._wechat_webhook.replace('/send?', '/upload_media?')
                upload_url = f"{base_url}&type={file_type}"
                
                try:
                    with open(file_path, 'rb') as f:
                        files = {'media': f}
                        resp = requests.post(upload_url, files=files, timeout=30)
                        if resp.status_code == 200:
                            result = resp.json()
                            if result.get('errcode') == 0:
                                return result.get('media_id')
                            else:
                                self.log(f"  ⚠️ 上传失败: {result}")
                        return None
                except Exception as e:
                    self.log(f"  ⚠️ 上传异常: {e}")
                    return None

            # 2. 发送截图（type=file）
            if screenshot_path and os.path.exists(screenshot_path):
                media_id = upload_file(screenshot_path, "file")
                if media_id:
                    file_data = {"msgtype": "file", "file": {"media_id": media_id}}
                    requests.post(self._wechat_webhook, json=file_data, timeout=10)
                    self.log("  📸 截图已发送")
                else:
                    self.log("  ⚠️ 截图上传失败")
            else:
                self.log("  ⚠️ 无截图可发送")

            # 3. 发送 Excel 文件（type=file）
            if excel_path and os.path.exists(excel_path):
                media_id = upload_file(excel_path, "file")
                if media_id:
                    file_data = {"msgtype": "file", "file": {"media_id": media_id}}
                    requests.post(self._wechat_webhook, json=file_data, timeout=10)
                    self.log("  📁 Excel文件已发送")
                else:
                    self.log("  ⚠️ Excel上传失败")
            else:
                self.log("  ⚠️ 无Excel文件可发送")

            self.log("  ✅ 风控告警发送完成")

        except Exception as e:
            self.log(f"  ❌ 发送风控告警失败: {e}")

    def send_profit_report_via_wechat(self, webhook_url: str, excel_path: str, new_records: List[Dict] = None):
        """发送正利润货源报告到企业微信"""
        if not webhook_url:
            return

        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            profit_records = new_records if new_records else [r for r in self.all_results if r.get('final_profit') and r['final_profit'] > 0]

            if not profit_records:
                self.log("  ℹ️ 无正利润货源，跳过推送")
                return

            # 构建摘要信息
            summary_lines = []
            for r in profit_records[:10]:  # 最多显示10条
                title = r.get('source_title', '')[:30]
                profit = r.get('final_profit', 0)
                quadrant = r.get('quadrant', '')
                summary_lines.append(f"- {title} → ¥{profit:.2f} ({quadrant})")
            if len(profit_records) > 10:
                summary_lines.append(f"- ... 共{len(profit_records)}条")

            content = f"""## 📦 拼多多货源报告（正利润）

**时间：** {timestamp}
**正利润货源数：** {len(profit_records)}

**详情摘要：**
{chr(10).join(summary_lines)}

**附件：** 完整Excel文件
"""
            data = {
                "msgtype": "markdown",
                "markdown": {"content": content}
            }
            requests.post(webhook_url, json=data, timeout=10)

            # 发送Excel文件
            if excel_path and os.path.exists(excel_path):
                with open(excel_path, 'rb') as f:
                    files = {'media': (os.path.basename(excel_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                    upload_url = webhook_url.replace('send', 'upload_media') + '&type=file'
                    upload_resp = requests.post(upload_url, files=files, timeout=30)
                    if upload_resp.status_code == 200:
                        media_id = upload_resp.json().get('media_id')
                        if media_id:
                            file_data = {
                                "msgtype": "file",
                                "file": {"media_id": media_id}
                            }
                            requests.post(webhook_url, json=file_data, timeout=10)

            self.log(f"  📤 正利润报告已发送到企业微信（{len(profit_records)}条）")
        except Exception as e:
            self.log(f"  ⚠️ 发送微信报告失败: {e}")

    def _loop(self):
        processed_since_last_break = 0

        while self._running:
            # 检查是否处于暂停状态
            if self._paused:
                time.sleep(1)
                continue

            # 检查缓存清理
            if self.cache_clear_interval_min > 0 and self.controller:
                now = time.time()
                if now - self._last_cache_clear_time >= self.cache_clear_interval_min * 60:
                    self.log(f"\n🧹 手机运行 {(now - self._last_cache_clear_time)/60:.0f} 分钟，开始清理缓存...")
                    self._clear_cache_and_restart()
                    self._last_cache_clear_time = now
                    time.sleep(2)

            try:
                item = self.task_queue.get(timeout=1.0)
            except queue.Empty:
                if self._auto_stop_when_empty and self.task_queue.empty():
                    self.log("\n📭 队列已空，停止处理...")
                    break
                time.sleep(0.5)
                continue

            try:
                self._process_item(item)
                self._processed_count += 1
                processed_since_last_break += 1

                if self.pause_every > 0 and processed_since_last_break >= self.pause_every:
                    wait = random.uniform(self.pause_duration * 0.8, self.pause_duration * 1.2)
                    self.log(f"\n☕ 已处理{processed_since_last_break}件，休息 {wait:.0f}s 避免风控...")
                    for remaining in range(int(wait), 0, -1):
                        if not self._running or self._paused:
                            break
                        self._emit_countdown(remaining)
                        time.sleep(1)
                    self.log("▶ 继续采集...")
                    self._emit_countdown(0)
                    processed_since_last_break = 0

            except Exception as e:
                self.log(f"❌ 处理异常: {e}")
            finally:
                self.task_queue.task_done()
                if self._running and not self._paused:
                    wait = max(3.0, random.gauss(self.delay_between_products, 2.0))
                    self.log(f"  ⏸ 下一件等待 {wait:.0f}s...")
                    for remaining in range(int(wait), 0, -1):
                        if not self._running or self._paused:
                            break
                        self._emit_countdown(remaining)
                        time.sleep(1)
                    self._emit_countdown(0)

        self._on_queue_completed()
        self.log("🛑 队列消费已停止")

    def _on_queue_completed(self):
        if not self.all_results:
            self.log("⚠️ 没有处理任何商品")
            return

        self.log("\n" + "=" * 50)
        self.log("✅ 队列已处理完成！")
        self.log(f"   共处理 {len(self.all_results)} 个商品")

        q_counts = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0, 'Q5': 0}
        profit_count = 0
        for r in self.all_results:
            q = r.get('quadrant', '')
            if q in q_counts:
                q_counts[q] += 1
            fp = r.get('final_profit')
            if fp is not None and fp > 0:
                profit_count += 1

        self.log(f"   ⭐Q1双正:{q_counts['Q1']} | 📷Q2图搜优:{q_counts['Q2']} | "
                 f"📝Q3标题参考:{q_counts['Q3']} | ❌Q4双负:{q_counts['Q4']} | ⛔Q5无数据:{q_counts['Q5']}")
        self.log(f"   找到正利润货源: {profit_count} 个")

        if profit_count > 0:
            excel_path = os.path.expanduser(
                f"~/Desktop/拼多多货源_正利润_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            export_supply_to_excel(self.all_results, excel_path, self.log, profit_only=True)
            self.log(f"📁 正利润已导出: {excel_path}")

        self.log("=" * 50)

    @staticmethod
    def _trim_search_keyword(title: str) -> str:
        """智能截断闲鱼标题，去除描述性文字，提取有效搜索词"""
        # 常见描述标记 → 在此处截断
        cut_markers = [
            "购买须知", "使用说明", "产品参数", "注意事项", "发货须知",
            "关于发货", "关于售后", "售后说明", "质保服务", "保修说明",
            "下单须知", "购买必读", "下单必读", "温馨提示", "特别说明",
            "产品信息", "规格参数", "产品详情", "商品详情",
            "一、", "二、", "三、", "四、", "五、",
            "1.", "2.", "3.",
        ]
        for marker in cut_markers:
            idx = title.find(marker)
            if idx > 20:  # 至少保留20个字符
                return title[:idx].strip().rstrip("，。,.")
        # 如果标题太长（>100字符），截断到100
        if len(title) > 100:
            return title[:100].strip().rstrip("，。,.")
        return title

    def _process_item(self, item: dict):
        title = item.get('商品标题', '')
        xianyu_desc = item.get('商品描述', '')
        try:
            xianyu_price = float(str(item.get('商品价格', 0)).replace(',', ''))
        except Exception:
            xianyu_price = 0.0

        if not title or xianyu_price <= 0:
            return

        score = item.get('综合评分', 0)

        # ── v9.0: 透传闲鱼需求数据（来自采集工具）──────────────────────
        try:
            shelf_days  = int(item.get('已上架天数', 0) or 0)
        except Exception:
            shelf_days  = 0
        try:
            want_cnt    = int(item.get('想要人数', 0) or 0)
        except Exception:
            want_cnt    = 0
        try:
            daily_want  = float(item.get('日均想要数', 0) or 0)
        except Exception:
            daily_want  = 0.0
        quadrant_tag = item.get('四象限标签', '')
        priority_tag = item.get('优先级', '')

        sold_cnt = item.get('已售数量', 0) or 0
        self.log(f"\n📦 [{score}分] {title[:40]} ¥{xianyu_price}"
                 f"  📅上架{shelf_days}天 ❤️想要{want_cnt}人 🛒已售{sold_cnt} 📈日均{daily_want:.1f}")

        cleaner = get_ai_cleaner()
        # 智能截断：闲鱼标题可能包含大量描述文字，截取有效产品名
        search_keyword = self._trim_search_keyword(title.strip())
        self.log(f"  🔍 搜索词: {search_keyword[:80]}{'...' if len(search_keyword) > 80 else ''}")

        raw_title_products = self.controller.search_and_collect(
            search_keyword,
            scroll_pages=self.scroll_pages,
            max_items=self.max_items,
        )
        self.log(f"  📦 标题搜采集 {len(raw_title_products)} 件")

        if len(raw_title_products) == 0:
            self._consecutive_empty_count += 1
            self.log(f"  ⚠️ 标题搜采集0件！连续 {self._consecutive_empty_count} 次")

            if self._consecutive_empty_count >= self._empty_threshold:
                if not self._cache_cleared_for_empty:
                    # ── 需求3 第一阶段：疑似手机卡死，先清缓存+重启 ──────────
                    self.log(f"  🧹 [第1阶段] 连续{self._consecutive_empty_count}次为空，"
                             f"疑似内存不足导致卡死，立即清理缓存+重启PDD...")
                    self._clear_cache_and_restart()
                    self._cache_cleared_for_empty = True
                    self._consecutive_empty_count = 0
                    self._post_cache_empty_count = 0
                    self.log(f"  ⏳ [第1阶段] 缓存已清理，等待{self._post_cache_wait_sec}秒后继续采集...")
                    for i in range(self._post_cache_wait_sec, 0, -1):
                        if not self._running:
                            break
                        self._emit_countdown(i)
                        time.sleep(1)
                    self._emit_countdown(0)
                    self.log(f"  ▶ [第1阶段] 恢复采集")
                else:
                    # ── 需求3 第二阶段：清缓存后依然为空 → 真正风控 ──────────
                    self._post_cache_empty_count += 1
                    if self._post_cache_empty_count >= self._empty_threshold and not self._empty_alert_sent:
                        self._empty_alert_sent = True
                        self.log(f"  🚨 [第2阶段] 清缓存后依然连续{self._post_cache_empty_count}次为空，确认为风控！")
                        self._trigger_risk_control_pause(
                            f"清缓存后依然连续{self._post_cache_empty_count}次采集到0件商品（确认风控）")
        else:
            # 恢复正常
            if self._consecutive_empty_count > 0 or self._post_cache_empty_count > 0:
                self.log("  ✅ 恢复采集，重置空采集计数")
            self._consecutive_empty_count = 0
            self._post_cache_empty_count = 0
            self._empty_alert_sent = False
            self._cache_cleared_for_empty = False
        img_attempted = False
        raw_img_products = []  # 初始化，防止无图片URL时引用未赋值变量

        if self.use_img_search:
            pics_raw = item.get('商品图片', '')
            first_pic = pics_raw.split(',')[0].strip() if pics_raw else ''
            if first_pic:
                self.log("  🖼️ 开始以图搜款...")
                raw_img_products = self.controller.image_search_from_result_page(
                    pic_url=first_pic,
                    scroll_pages=self.img_scroll_pages,
                    max_items=self.max_items,
                )
                img_attempted = True
                if not raw_img_products:
                    self.log("  [图搜] ⚠️ 图搜未采集到商品")
            else:
                self.log("  [图搜] ⚠️ 无闲鱼图片URL，跳过图搜")
        else:
            self.log("  [图搜] 已禁用，跳过")

        self.controller.back_to_search_start()

        compare_result = None
        used_ai_compare = False
        title_ai_ok = False
        img_ai_ok = False

        if self.use_ai_compare and cleaner.enabled and (raw_title_products or raw_img_products):
            self.log(f"  🤖 AI同款比对（标题{len(raw_title_products)}件 / 图搜{len(raw_img_products)}件）...")
            compare_result = cleaner.compare_batch(
                xianyu_title=title,
                xianyu_desc=xianyu_desc,
                search_keyword=search_keyword,
                title_products=raw_title_products,
                img_products=raw_img_products,
            )
            if compare_result:
                used_ai_compare = True
                # 分路AI状态：部分成功时各路独立降级
                title_ai_ok = compare_result.get('title_ai_ok', True)
                img_ai_ok = compare_result.get('img_ai_ok', True)
            else:
                title_ai_ok = False
                img_ai_ok = False
                self.log("  ⚠️ AI比对失败，降级到语义/字符相似度")

        def _process_product_list(raw_products, ai_matches_list, use_ai, platform_label, log_prefix=''):
            """
            统一处理标题路/图搜路的商品列表：
            - use_ai=True 且 ai_matches_list 非空：优先用AI结果
            - 单条 ai_failed=True：该条降级到语义相似度
            - use_ai=False：全部用语义相似度
            """
            matched = []
            self.log(f"  {log_prefix}📋 {platform_label}商品（{'AI全量' if use_ai else '相似度模型'}，共{len(raw_products)}件）：")

            for i, p in enumerate(raw_products):
                # 优先用AI结果
                if use_ai and i < len(ai_matches_list):
                    m = ai_matches_list[i]
                    ai_failed_flag = m.get('ai_failed', False)
                else:
                    ai_failed_flag = True
                    m = {}

                if not ai_failed_flag:
                    # ── AI判断路径 ──
                    product_type  = m.get('product_type', 'unrelated')
                    is_same       = (product_type == 'same')
                    is_alternative= (product_type == 'alternative')
                    can_be_supply = is_same or is_alternative
                    sim           = m.get('score', 0.85 if is_same else (0.7 if is_alternative else 0.1))
                    src           = '🤖AI'
                    match_reason  = m.get('reason', '')
                else:
                    # ── 降级：语义相似度 ──
                    product_type   = 'unrelated'
                    is_same, sim   = self.matcher.is_same_product(
                        search_keyword, p['title'], self.sim_threshold)
                    is_alternative = False
                    can_be_supply  = is_same
                    src            = '📐语义'
                    match_reason   = f"相似度{sim:.0%}"

                can_be_supply_final = bool(can_be_supply) and (is_same or is_alternative)

                if is_same:
                    flag        = "✅ 同款"
                    match_level = 'exact'
                elif is_alternative:
                    flag        = "🔄 平替"
                    match_level = 'good'
                else:
                    flag        = "❌ 无关"
                    match_level = 'mismatch'

                reason_short = match_reason[:60] + "..." if len(match_reason) > 60 else match_reason
                self.log(f"      {flag}[{src}|{sim:.2%}] ¥{p['price']} {p['title'][:35]}")
                if match_reason:
                    self.log(f"          📝 {reason_short}")

                if can_be_supply_final:
                    matched.append({
                        'goods_name':     p['title'],
                        'pdd_price_yuan': p['price'],
                        'sales_tip':      p.get('sales', ''),
                        'sim_score':      sim,
                        'platform':       platform_label,
                        'match_src':      src,
                        'match_level':    match_level,
                        'match_reason':   match_reason,
                        'is_exact_match': is_same,
                        'is_alternative': is_alternative,
                        'product_type':   product_type,
                    })
            return matched

        title_matched: List[Dict] = []
        if raw_title_products:
            _use_ai_title   = used_ai_compare and title_ai_ok
            ai_title_list   = compare_result.get('title_matches', []) if compare_result else []
            title_matched   = _process_product_list(
                raw_title_products, ai_title_list, _use_ai_title, '拼多多(标题搜)')
            self.log(f"  ✅ 标题搜有效商品 {len(title_matched)} 件")

        img_matched: List[Dict] = []
        if raw_img_products:
            _use_ai_img   = used_ai_compare and img_ai_ok
            ai_img_list   = compare_result.get('img_matches', []) if compare_result else []
            img_matched   = _process_product_list(
                raw_img_products, ai_img_list, _use_ai_img, '拼多多(图搜)', log_prefix='[图搜] ')
            self.log(f"  [图搜] ✅ 图搜有效商品 {len(img_matched)} 件")

        title_analyzed = self.analyzer.analyze(xianyu_price, title_matched)
        img_analyzed = self.analyzer.analyze(xianyu_price, img_matched)

        # ── 需求2：分别计算同款利润 / 平替利润 ────────────────────────────
        def _best_profit_by_type(items, exact_only: bool):
            """从analyzed列表中找同款或平替的最优利润"""
            filtered = [it for it in items if
                        (it.get('is_exact_match') if exact_only
                         else (it.get('is_alternative') and not it.get('is_exact_match')))]
            if not filtered:
                return None, None
            best = filtered[0]  # analyzer已按评分降序排好
            return best.get('预估利润(元)'), best.get('拼多多进价(元)')

        title_same_profit, title_same_price = _best_profit_by_type(title_analyzed, True)
        title_alt_profit, title_alt_price = _best_profit_by_type(title_analyzed, False)
        img_same_profit, img_same_price = _best_profit_by_type(img_analyzed, True)
        img_alt_profit, img_alt_price = _best_profit_by_type(img_analyzed, False)

        if title_analyzed:
            b = title_analyzed[0]
            self.log(f"  💰 标题最优: {b['goods_name'][:30]} ¥{b['拼多多进价(元)']} "
                     f"→ 利润¥{b['预估利润(元)']} ({b['利润率(%)']:.1f}%) [{b['匹配类型']}]")
            if title_same_profit is not None:
                self.log(f"      ✅ 标题同款利润: ¥{title_same_profit:.2f}  进价¥{title_same_price:.2f}")
            if title_alt_profit is not None:
                self.log(f"      🔄 标题平替利润: ¥{title_alt_profit:.2f}  进价¥{title_alt_price:.2f}")
        if img_analyzed:
            b = img_analyzed[0]
            self.log(f"  [图搜] 💰 图搜最优: {b['goods_name'][:30]} ¥{b['拼多多进价(元)']} "
                     f"→ 利润¥{b['预估利润(元)']} ({b['利润率(%)']:.1f}%) [{b['匹配类型']}]")
            if img_same_profit is not None:
                self.log(f"      ✅ 图搜同款利润: ¥{img_same_profit:.2f}  进价¥{img_same_price:.2f}")
            if img_alt_profit is not None:
                self.log(f"      🔄 图搜平替利润: ¥{img_alt_profit:.2f}  进价¥{img_alt_price:.2f}")

        quadrant_result = evaluate_supply_quadrant(
            title_items=title_analyzed,
            img_items=img_analyzed,
            title_has_data=(len(raw_title_products) > 0),
            img_has_data=(img_attempted and len(raw_img_products) > 0),
        )
        q = quadrant_result
        self.log(f"  {q['quadrant_emoji']} 象限={q['quadrant']} {q['quadrant_label']} | {q['recommendation']}")

        parts = []
        if title_matched:
            exact_cnt = sum(1 for it in title_matched if it.get('is_exact_match'))
            alt_cnt = sum(1 for it in title_matched if it.get('is_alternative') and not it.get('is_exact_match'))
            parts.append(f"标题(同款{exact_cnt}/平替{alt_cnt})")
        if img_matched:
            exact_cnt = sum(1 for it in img_matched if it.get('is_exact_match'))
            alt_cnt = sum(1 for it in img_matched if it.get('is_alternative') and not it.get('is_exact_match'))
            parts.append(f"图搜(同款{exact_cnt}/平替{alt_cnt})")
        status = f"{q['quadrant_emoji']}{q['quadrant']} {'/'.join(parts) or '无有效货源'}"

        record = {
            'source_title': title,
            'xianyu_price': xianyu_price,
            'xianyu_desc': xianyu_desc,
            'xianyu_score': score,
            'xianyu_link': item.get('商品链接', ''),
            'xianyu_pics': item.get('商品图片', ''),
            'xianyu_video': item.get('商品视频', ''),
            # ── v9.0: 闲鱼需求数据 ──
            'shelf_days': shelf_days,
            'want_cnt': want_cnt,
            'daily_want': daily_want,
            'xianyu_sold_count': item.get('已售数量', item.get('sold_count', item.get('soldCount', ''))),
            # ── 卖家信息 ──
            'seller_nick': item.get('卖家昵称', ''),
            'seller_sold': item.get('卖家已售', 0),
            'seller_good_rate': item.get('卖家好评率', ''),
            'seller_reply_rate': item.get('卖家回复率', ''),
            # ── 商品详情字段 ──
            'xianyu_desc_full': item.get('商品描述', ''),
            'collect_cnt': item.get('收藏数', 0),
            'browse_cnt': item.get('浏览数', 0),
            'quadrant_tag': quadrant_tag,
            'priority_tag': priority_tag,
            'search_keyword': search_keyword,
            'pdd_items': title_analyzed,
            'img_pdd_items': img_analyzed,
            'quadrant': q['quadrant'],
            'quadrant_label': q['quadrant_label'],
            'quadrant_emoji': q['quadrant_emoji'],
            'title_profit': q['title_profit'],
            'img_profit': q['img_profit'],
            'final_profit': q['final_profit'],
            'final_price': q['final_price'],
            'final_source': q['final_source'],
            'recommendation': q['recommendation'],
            'status': status,
            # ── 需求2新增：4维细分利润 ──
            'title_same_profit': round(title_same_profit, 2) if title_same_profit is not None else None,
            'title_same_price': round(title_same_price, 2) if title_same_price is not None else None,
            'title_alt_profit': round(title_alt_profit, 2) if title_alt_profit is not None else None,
            'title_alt_price': round(title_alt_price, 2) if title_alt_price is not None else None,
            'img_same_profit': round(img_same_profit, 2) if img_same_profit is not None else None,
            'img_same_price': round(img_same_price, 2) if img_same_price is not None else None,
            'img_alt_profit': round(img_alt_profit, 2) if img_alt_profit is not None else None,
            'img_alt_price': round(img_alt_price, 2) if img_alt_price is not None else None,
        }
        self.all_results.append(record)
        self.push_result(record)


# ══════════════════════════════════════════════════════════════════════════
# Excel 导出
# ══════════════════════════════════════════════════════════════════════════

def export_supply_to_excel(all_results: list, excel_path: str,
                           log_cb: Callable = None, profit_only: bool = False):
    if not PANDAS_OK:
        if log_cb:
            log_cb("❌ pandas 未安装")
        return
    try:
        summary_rows = []
        title_detail_rows = []
        img_detail_rows = []

        for r in all_results:
            final_profit = r.get('final_profit')

            if profit_only:
                if final_profit is None or final_profit <= 0:
                    continue

            all_pics = r.get('xianyu_pics', '')
            first_pic = all_pics.split(',')[0].strip() if all_pics else ''

            title_items = r.get('pdd_items', [])
            img_items = r.get('img_pdd_items', [])
            title_best = title_items[0] if title_items else None
            img_best = img_items[0] if img_items else None

            summary_rows.append({
                '闲鱼商品标题': r['source_title'],
                '闲鱼价格(元)': r['xianyu_price'],
                '综合评分': r.get('xianyu_score', ''),
                '已售数量': r.get('xianyu_sold_count', ''),
                '已上架天数': r.get('shelf_days', ''),
                '想要人数': r.get('want_cnt', ''),
                '日均想要数': r.get('daily_want', ''),
                '收藏数': r.get('collect_cnt', 0),
                '浏览数': r.get('browse_cnt', 0),
                '商品描述': (r.get('xianyu_desc_full', '') or '')[:200],
                '卖家昵称': r.get('seller_nick', ''),
                '卖家已售': r.get('seller_sold', 0),
                '卖家好评率': r.get('seller_good_rate', ''),
                '卖家回复率': r.get('seller_reply_rate', ''),
                '象限': f"{r.get('quadrant_emoji', '')}{r.get('quadrant', '')} {r.get('quadrant_label', '')}",
                '建议': r.get('recommendation', ''),
                '综合最优利润(元)': final_profit if final_profit is not None else '',
                '综合最优进价(元)': r.get('final_price', ''),
                '利润来源': r.get('final_source', ''),
                '标题搜利润(元)': r.get('title_profit', ''),
                '图搜利润(元)': r.get('img_profit', ''),
                '标题同款利润(元)': r.get('title_same_profit', ''),
                '标题同款进价(元)': r.get('title_same_price', ''),
                '标题平替利润(元)': r.get('title_alt_profit', ''),
                '标题平替进价(元)': r.get('title_alt_price', ''),
                '图搜同款利润(元)': r.get('img_same_profit', ''),
                '图搜同款进价(元)': r.get('img_same_price', ''),
                '图搜平替利润(元)': r.get('img_alt_profit', ''),
                '图搜平替进价(元)': r.get('img_alt_price', ''),
                '标题搜有效数': len(title_items),
                '图搜有效数': len(img_items),
                '搜索关键词': r.get('search_keyword', ''),
                '状态': r.get('status', ''),
                '闲鱼链接': r.get('xianyu_link', ''),
                '闲鱼图片(第一张)': first_pic,
                '闲鱼图片(全部)': all_pics,
            })

            for it in title_items:
                title_detail_rows.append({
                    '闲鱼商品标题': r['source_title'],
                    '闲鱼价格(元)': r['xianyu_price'],
                    '象限': r.get('quadrant', ''),
                    '货源来源': '标题搜索',
                    '货源商品名': it.get('goods_name', ''),
                    '货源价格(元)': it.get('拼多多进价(元)', ''),
                    '预估利润(元)': it.get('预估利润(元)', ''),
                    '利润率(%)': it.get('利润率(%)', ''),
                    '货源评分': it.get('货源评分', ''),
                    '是否推荐': it.get('是否推荐货源', ''),
                    '匹配类型': it.get('匹配类型', ''),
                    '匹配说明': it.get('匹配说明', ''),
                    '销量': it.get('sales_tip', ''),
                    '相似度': it.get('sim_score', 0),
                    '比对方式': it.get('match_src', ''),
                })

            for it in img_items:
                img_detail_rows.append({
                    '闲鱼商品标题': r['source_title'],
                    '闲鱼价格(元)': r['xianyu_price'],
                    '象限': r.get('quadrant', ''),
                    '货源来源': '以图搜款',
                    '货源商品名': it.get('goods_name', ''),
                    '货源价格(元)': it.get('拼多多进价(元)', ''),
                    '预估利润(元)': it.get('预估利润(元)', ''),
                    '利润率(%)': it.get('利润率(%)', ''),
                    '货源评分': it.get('货源评分', ''),
                    '是否推荐': it.get('是否推荐货源', ''),
                    '匹配类型': it.get('匹配类型', ''),
                    '匹配说明': it.get('匹配说明', ''),
                    '销量': it.get('sales_tip', ''),
                    '相似度': it.get('sim_score', 0),
                    '比对方式': it.get('match_src', ''),
                })

        if not summary_rows:
            if log_cb:
                log_cb("⚠️ 没有符合条件的数据可导出")
            return

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            pd.DataFrame(summary_rows).to_excel(
                writer, sheet_name='货源汇总', index=False)
            if title_detail_rows:
                pd.DataFrame(title_detail_rows).to_excel(
                    writer, sheet_name='标题搜索详情', index=False)
            if img_detail_rows:
                pd.DataFrame(img_detail_rows).to_excel(
                    writer, sheet_name='图搜货源详情', index=False)

        if log_cb:
            label = "（仅正利润）" if profit_only else ""
            log_cb(f"✅ 已导出{label}: {excel_path}  共{len(summary_rows)}件 "
                   f"（标题详情{len(title_detail_rows)}条/图搜详情{len(img_detail_rows)}条）")
    except Exception as e:
        if log_cb:
            log_cb(f"❌ 导出失败: {e}")


# ══════════════════════════════════════════════════════════════════════════
# 上架建议表导出 v9.0
# ══════════════════════════════════════════════════════════════════════════

def _calc_listing_priority(r: dict) -> str:
    """
    综合上架优先级判断 v4.0（最终修正版）
    
    核心逻辑：
    1. 日均想要数 = 最核心指标（反映真实且持续的热度）
    2. 上架天数 → 不直接减分，而是作为稳定性参考
       - 日均高 + 天数长 = 稳定爆款（S级）
       - 日均高 + 天数短 = 潜力爆款（S级）
       - 日均低 + 天数长 = 老品冷门（降级）
    3. 绝对想要数作为辅助验证
    """
    
    # ========== 1. 提取数据 ==========
    shelf_days = r.get('shelf_days', 0) or 0
    want_cnt = r.get('want_cnt', 0) or 0
    daily_want = r.get('daily_want', 0.0) or 0.0
    
    # 利润相关
    final_profit = r.get('final_profit') or 0
    title_same_p = r.get('title_same_profit') or 0
    img_same_p = r.get('img_same_profit') or 0
    title_alt_p = r.get('title_alt_profit') or 0
    img_alt_p = r.get('img_alt_profit') or 0
    
    best_same = max(title_same_p, img_same_p)
    best_alt = max(title_alt_p, img_alt_p)
    best_profit = max(best_same, best_alt, final_profit)
    
    has_profit_same = best_same > 0
    has_profit_alt = best_alt > 0
    has_any_profit = best_profit > 0
    
    # ========== 2. 计算需求分（0-100）==========
    # 日均想要分（权重 70%）- 绝对核心
    if daily_want >= 10:
        daily_score = 100   # 日均10+，超级爆款
    elif daily_want >= 5:
        daily_score = 90    # 日均5-10，火爆
    elif daily_want >= 3:
        daily_score = 80    # 日均3-5，很热
    elif daily_want >= 2:
        daily_score = 70    # 日均2-3，不错
    elif daily_want >= 1:
        daily_score = 55    # 日均1-2，正常
    elif daily_want >= 0.5:
        daily_score = 35    # 日均0.5-1，一般
    elif daily_want > 0:
        daily_score = 20    # 日均<0.5，偏冷
    else:
        daily_score = 0
    
    # 绝对想要数分（权重 30%）- 辅助验证
    if want_cnt >= 500:
        want_score = 100
    elif want_cnt >= 200:
        want_score = 85
    elif want_cnt >= 100:
        want_score = 70
    elif want_cnt >= 50:
        want_score = 55
    elif want_cnt >= 20:
        want_score = 40
    elif want_cnt >= 10:
        want_score = 25
    elif want_cnt >= 5:
        want_score = 15
    elif want_cnt > 0:
        want_score = 8
    else:
        want_score = 0
    
    # 需求分 = 日均主导，绝对想要辅助
    demand_score = daily_score * 0.70 + want_score * 0.30
    
    # ========== 3. 上架天数越短 + 日均想要越高 = 潜力爆款 ==========
    stability_bonus = 0

    # 新品且高热度 → 潜力爆款（大幅加分）
    if shelf_days <= 7 and daily_want >= 5:
        stability_bonus = 15   # 🏆 潜力爆款，刚上架就爆
    elif shelf_days <= 7 and daily_want >= 2:
        stability_bonus = 10   # 📈 新品有热度，潜力大
    elif shelf_days <= 14 and daily_want >= 5:
        stability_bonus = 8    # 两周内爆款
    elif shelf_days <= 14 and daily_want >= 2:
        stability_bonus = 5    # 两周内有热度

    # 老品但高热度 → 经典爆款（小幅加分，但不如新品）
    elif shelf_days > 30 and daily_want >= 5:
        stability_bonus = 3    # 老爆款，仍有价值
    elif shelf_days > 60 and daily_want >= 3:
        stability_bonus = 1    # 稳定老品，可作为参考

    # 老品且低热度 → 僵尸品（扣分）
    if shelf_days > 30 and daily_want < 0.5:
        stability_bonus = -10
    elif shelf_days > 60 and daily_want < 1:
        stability_bonus = -15
    elif shelf_days > 90:
        stability_bonus = -20  # 上架太久，已过时
    
    # ========== 4. 计算利润分 ==========
    if best_profit >= 30:
        profit_score = 100
    elif best_profit >= 20:
        profit_score = 85
    elif best_profit >= 10:
        profit_score = 70
    elif best_profit >= 5:
        profit_score = 50
    elif best_profit > 0:
        profit_score = 30
    else:
        profit_score = 0
    
    if has_profit_same:
        profit_score = min(100, profit_score + 5)
    elif has_profit_alt:
        profit_score = max(0, profit_score - 10)
    
    # ========== 5. 综合评分 ==========
    total_score = demand_score * 0.60 + profit_score * 0.40 + stability_bonus
    
    # 无利润来源
    if not has_any_profit:
        if demand_score >= 65:
            return 'C级'
        else:
            return 'D级'
    
    # ========== 6. 确定等级 ==========
    if total_score >= 80:
        return 'S级'
    elif total_score >= 65:
        return 'A级'
    elif total_score >= 50:
        return 'B级'
    elif total_score >= 35:
        return 'C级'
    else:
        return 'D级'


def _fmt_supply_top3(items: list, type_filter: str, top_n: int = 3) -> str:
    """格式化同款或平替前N货源摘要"""
    if not items:
        return '—'
    filtered = [it for it in items if it.get('匹配类型', '') == type_filter]
    if not filtered:
        return '—'
    parts = []
    for it in filtered[:top_n]:
        name  = it.get('goods_name', '')[:20]
        price = it.get('拼多多进价(元)', '')
        profit= it.get('预估利润(元)', '')
        parts.append(f"{name} ¥{price}(利润¥{profit})")
    return ' | '.join(parts)


def export_listing_advice_to_excel(all_results: list, excel_path: str,
                                   log_cb: Callable = None):
    """
    导出上架建议表（独立Sheet）。
    字段：商品标题 / 商品描述 / 商品价格 / 已上架天数 / 想要人数 / 日均想要数 /
         闲鱼综合评分 / 四象限标签 / 最优同款前3 / 平替前3 / 货源简述 /
         最优货源价格 / 最优利润 / 上架优先级 / 建议说明 / 闲鱼链接
    """
    if not PANDAS_OK:
        if log_cb:
            log_cb("❌ pandas 未安装，无法导出上架建议表")
        return
    try:
        rows = []
        for r in all_results:
            title_items = r.get('pdd_items', []) or []
            img_items   = r.get('img_pdd_items', []) or []
            all_items   = title_items + img_items

            # 最优货源（按货源评分降序，已由analyzer排好）
            best = title_items[0] if title_items else (img_items[0] if img_items else None)

            # 上架优先级
            listing_priority = _calc_listing_priority(r)

            # 货源简述（最优货源名称）
            supply_brief = best.get('goods_name', '—')[:30] if best else '暂无货源'

            # 建议说明
            fp = r.get('final_profit')
            shelf_days = r.get('shelf_days', 0) or 0
            daily_want = r.get('daily_want', 0.0) or 0.0
            want_cnt   = r.get('want_cnt', 0) or 0

            advice_parts = []
            if listing_priority == 'S级':
                advice_parts.append('🏆 需求旺盛+有利润，强烈建议优先上架')
            elif listing_priority == 'A级':
                advice_parts.append('✅ 有可用货源+正利润，建议上架')
            elif listing_priority == 'B级':
                advice_parts.append('🔄 货源可用，利润一般，可尝试上架')
            elif listing_priority == 'C级':
                advice_parts.append('⚠️ 货源存在但利润偏低，谨慎上架')
            else:
                advice_parts.append('❌ 暂无有效货源，暂不建议上架')

            if daily_want >= 2.0:
                advice_parts.append(f'日均想要{daily_want:.1f}人（需求旺）')
            if shelf_days > 60:
                advice_parts.append(f'已上架{shelf_days}天（需关注滞销）')
            if fp is not None and fp > 0:
                advice_parts.append(f'最高利润¥{fp:.1f}')

            rows.append({
                '上架优先级':   listing_priority,
                '商品标题':     r.get('source_title', ''),
                '商品描述':     (r.get('xianyu_desc_full', '') or '')[:200],
                '商品价格(元)': r.get('xianyu_price', ''),
                '已售数量':     r.get('xianyu_sold_count', ''),
                '已上架天数':   shelf_days if shelf_days else '—',
                '想要人数':     want_cnt if want_cnt else '—',
                '日均想要数':   f"{daily_want:.2f}" if daily_want else '—',
                '收藏数':       r.get('collect_cnt', 0),
                '浏览数':       r.get('browse_cnt', 0),
                '卖家昵称':     r.get('seller_nick', ''),
                '卖家已售':     r.get('seller_sold', 0),
                '卖家好评率':   r.get('seller_good_rate', ''),
                '闲鱼综合评分': r.get('xianyu_score', ''),
                '四象限标签':   r.get('quadrant_tag', r.get('quadrant_label', '')),
                '最优同款前3':  _fmt_supply_top3(all_items, '同款', 3),
                '平替前3':      _fmt_supply_top3(all_items, '平替', 3),
                '货源简述':     supply_brief,
                '最优货源价格(元)': best.get('拼多多进价(元)', '—') if best else '—',
                '最优利润(元)': f"{fp:.2f}" if fp is not None else '—',
                '建议说明':     ' | '.join(advice_parts),
                '闲鱼链接':     r.get('xianyu_link', ''),
            })

        if not rows:
            if log_cb:
                log_cb("⚠️ 上架建议表：暂无数据")
            return

        df = pd.DataFrame(rows)

        # 按优先级排序：S > A > B > C > D
        priority_order = {'S级': 0, 'A级': 1, 'B级': 2, 'C级': 3, 'D级': 4}
        df['_sort'] = df['上架优先级'].map(priority_order).fillna(9)
        df = df.sort_values('_sort').drop(columns=['_sort'])

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='上架建议', index=False)

            # 自动调列宽 + 条件着色
            try:
                from openpyxl.styles import PatternFill, Font
                ws = writer.sheets['上架建议']

                # 列宽: A=优先级 B=标题 C=描述 D=价格 E=已售 F=天数 G=想要 H=日均
                #       I=收藏 J=浏览 K=卖家昵称 L=卖家已售 M=好评率 N=评分 O=象限
                #       P=同款前3 Q=平替前3 R=货源简述 S=货源价 T=利润 U=建议 V=链接
                col_widths = {
                    'A': 8, 'B': 30, 'C': 30, 'D': 10, 'E': 10,
                    'F': 10, 'G': 10, 'H': 10, 'I': 8, 'J': 8,
                    'K': 14, 'L': 10, 'M': 10, 'N': 12, 'O': 14,
                    'P': 45, 'Q': 45, 'R': 20, 'S': 12, 'T': 10,
                    'U': 50, 'V': 40,
                }
                for col_letter, width in col_widths.items():
                    ws.column_dimensions[col_letter].width = width

                # 优先级着色
                color_map = {
                    'S级': ('FFD700', 'FF0000'),  # 金底红字
                    'A级': ('E8F8EF', '27AE60'),  # 绿底绿字
                    'B级': ('E8F0FB', '185FA5'),  # 蓝底蓝字
                    'C级': ('FFF3CD', '856404'),  # 黄底棕字
                    'D级': ('FEF0F0', 'C0392B'),  # 红底红字
                }
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell_a = row[0]
                    priority = cell_a.value or ''
                    if priority in color_map:
                        bg, fg = color_map[priority]
                        fill = PatternFill(fill_type='solid', fgColor=bg)
                        font = Font(color=fg, bold=True)
                        cell_a.fill = fill
                        cell_a.font = font
            except Exception:
                pass

        if log_cb:
            s_cnt = sum(1 for row in rows if row['上架优先级'] == 'S级')
            a_cnt = sum(1 for row in rows if row['上架优先级'] == 'A级')
            log_cb(f"✅ 上架建议表已导出: {excel_path}  "
                   f"共{len(rows)}件 (S级{s_cnt}/A级{a_cnt})")

    except Exception as e:
        if log_cb:
            log_cb(f"❌ 上架建议表导出失败: {e}")


# ══════════════════════════════════════════════════════════════════════════
# 可排序 Treeview 辅助
# ══════════════════════════════════════════════════════════════════════════

def make_sortable(tree: ttk.Treeview, numeric_cols: Set[str]):
    tree._sort_state: Dict[str, bool] = {}

    def sort_col(col):
        reverse = tree._sort_state.get(col, False)
        data = [(tree.set(k, col), k) for k in tree.get_children('')]

        def key_fn(t):
            v = t[0]
            if col in numeric_cols:
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return -999999
            return v

        data.sort(key=key_fn, reverse=reverse)
        for i, (_, k) in enumerate(data):
            tree.move(k, '', i)

        tree._sort_state[col] = not reverse
        for c in tree['columns']:
            arrow = ''
            if c == col:
                arrow = ' ▼' if reverse else ' ▲'
            tree.heading(c, text=c.rstrip(' ▲▼') + arrow,
                         command=lambda _c=c: sort_col(_c))

    for col in tree['columns']:
        tree.heading(col, text=col, command=lambda c=col: sort_col(c))


# ══════════════════════════════════════════════════════════════════════════
# 货源详情窗口
# ══════════════════════════════════════════════════════════════════════════

class SourceDetailWindow:
    """双击详情行弹出的单条货源纵向阅读卡片"""
    def __init__(self, parent, title: str, source_items: list, source_type: str):
        self.window = tk.Toplevel(parent)
        self.window.title(f"{source_type} — 货源详情")
        self.window.geometry("700x540")
        self.window.configure(bg="#F5F6FA")
        self.window.transient(parent)

        # 顶部标题栏
        header = tk.Frame(self.window, bg="#FF6B35")
        header.pack(fill=tk.X)
        tk.Label(header, text=f"📦 {source_type} 货源详情",
                 bg="#FF6B35", fg="white",
                 font=("Microsoft YaHei", 11, "bold")).pack(side=tk.LEFT, padx=14, pady=10)

        # 滚动区域
        outer = tk.Frame(self.window, bg="#F5F6FA")
        outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

        canvas = tk.Canvas(outer, bg="#F5F6FA", highlightthickness=0)
        vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner = tk.Frame(canvas, bg="#F5F6FA")
        iwin  = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(iwin, width=e.width))

        def _wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _wheel)
        canvas.bind("<Button-4>",   lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>",   lambda e: canvas.yview_scroll(1,  "units"))
        inner.bind("<MouseWheel>",  _wheel)

        COLORS = {
            '同款': ('#E8F8EF', '#27AE60', '✅ 同款'),
            '平替': ('#E8F0FB', '#185FA5', '🔄 平替'),
        }
        FG   = "#3D4152"
        FG_M = "#8A90A8"
        SURF = "#FFFFFF"
        BRD  = "#E2E5F0"

        if not source_items:
            tk.Label(inner, text="暂无货源数据", bg="#F5F6FA",
                     fg=FG_M, font=("Microsoft YaHei", 11)).pack(pady=30)
        else:
            for rank, it in enumerate(source_items, 1):
                match_type = it.get('匹配类型', '')
                bg_c, acc_c, badge = COLORS.get(match_type, ('#F8F9FD', '#8A90A8', '⚪ 未知'))
                profit = it.get('预估利润(元)', 0) or 0
                if profit < 0:
                    bg_c, acc_c = '#FEF0F0', '#E74C3C'

                card = tk.Frame(inner, bg=SURF,
                                highlightbackground=BRD, highlightthickness=1)
                card.pack(fill=tk.X, padx=2, pady=4)

                bar = tk.Frame(card, bg=acc_c, width=5)
                bar.pack(side=tk.LEFT, fill=tk.Y)

                body = tk.Frame(card, bg=SURF)
                body.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=8)

                # 商品名 + 匹配标签
                r1 = tk.Frame(body, bg=SURF)
                r1.pack(fill=tk.X)
                tk.Label(r1, text=f"#{rank}", bg=SURF, fg=FG_M,
                         font=("Microsoft YaHei", 8), width=3, anchor='w').pack(side=tk.LEFT)
                tk.Label(r1, text=it.get('goods_name', ''),
                         bg=SURF, fg=FG,
                         font=("Microsoft YaHei", 10, "bold"),
                         anchor='w', wraplength=520,
                         justify=tk.LEFT).pack(side=tk.LEFT, fill=tk.X, expand=True)
                tk.Label(r1, text=badge, bg=bg_c, fg=acc_c,
                         font=("Microsoft YaHei", 8),
                         padx=6, pady=2).pack(side=tk.RIGHT)

                # 利润数字行
                r2 = tk.Frame(body, bg=SURF)
                r2.pack(fill=tk.X, pady=(4, 0))
                pcolor = "#27AE60" if profit > 0 else "#E74C3C"
                tk.Label(r2, text=f"¥{profit:.2f}", bg=SURF, fg=pcolor,
                         font=("Microsoft YaHei", 15, "bold")).pack(side=tk.LEFT)
                tk.Label(r2, text=" 利润", bg=SURF, fg=FG_M,
                         font=("Microsoft YaHei", 8)).pack(side=tk.LEFT, pady=(4, 0))

                price_v = it.get('拼多多进价(元)', 0) or 0
                rate_v  = it.get('利润率(%)', 0) or 0
                score_v = it.get('货源评分', 0) or 0
                for label, val in [
                    ("进价", f"¥{price_v:.2f}"),
                    ("利润率", f"{rate_v:.1f}%"),
                    ("评分", str(score_v)),
                ]:
                    tk.Label(r2, text=f"  {label} ", bg=SURF, fg=FG_M,
                             font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)
                    tk.Label(r2, text=val, bg=SURF, fg=FG,
                             font=("Microsoft YaHei", 9, "bold")).pack(side=tk.LEFT)

                rec = it.get('是否推荐货源', '')
                if rec:
                    tk.Label(r2, text=rec, bg=SURF, fg=FG_M,
                             font=("Microsoft YaHei", 8)).pack(side=tk.RIGHT)

                # 元数据行
                r3 = tk.Frame(body, bg=SURF)
                r3.pack(fill=tk.X, pady=(3, 0))
                sales = it.get('sales_tip', '') or ''
                sim   = it.get('sim_score', 0) or 0
                src   = it.get('match_src', '') or ''
                meta_parts = []
                if sales: meta_parts.append(f"销量: {sales}")
                if src:   meta_parts.append(f"比对: {src}")
                meta_parts.append(f"相似度: {sim:.0%}")
                tk.Label(r3, text="  ".join(meta_parts),
                         bg=SURF, fg=FG_M,
                         font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)

                # AI理由
                reason = it.get('匹配说明', '') or it.get('reason', '') or ''
                if reason:
                    r4 = tk.Frame(body, bg="#F0F4FF",
                                  highlightbackground="#C8D4F0", highlightthickness=1)
                    r4.pack(fill=tk.X, pady=(6, 0))
                    tk.Label(r4, text=f"💬 {reason}",
                             bg="#F0F4FF", fg="#4A5580",
                             font=("Microsoft YaHei", 8),
                             wraplength=620, justify=tk.LEFT,
                             anchor='w').pack(fill=tk.X, padx=8, pady=5)

        # 底部关闭按钮
        bf = tk.Frame(self.window, bg="#F5F6FA")
        bf.pack(fill=tk.X, pady=(0, 10))
        tk.Button(bf, text="关闭",
                  command=self.window.destroy,
                  bg="#EEF0F8", fg="#3D4152",
                  padx=24, pady=6,
                  font=("Microsoft YaHei", 9),
                  relief=tk.FLAT, cursor="hand2").pack()


# ══════════════════════════════════════════════════════════════════════════
# GUI Tab（更新：增加继续处理按钮）
# ══════════════════════════════════════════════════════════════════════════

class SupplyFinderTab:
    BG = "#F5F6FA"
    SURF = "#FFFFFF"
    RAISED = "#EEF0F8"
    INP = "#F8F9FD"
    ACC = "#FF6B35"
    FG = "#3D4152"
    FG_M = "#8A90A8"
    BRD = "#E2E5F0"
    SUCC = "#27AE60"
    DANGER = "#E74C3C"

    def __init__(self, notebook: ttk.Notebook, main_app):
        self.nb = notebook
        self.app = main_app
        self.frame = tk.Frame(notebook, bg=self.BG)
        self.scheduler: Optional[MobileSupplyScheduler] = None
        self.task_queue: queue.Queue = queue.Queue()
        self._result_records: list = []
        self.controller: Optional[PinduoduoMobileController] = None

        self.use_ai_var = tk.BooleanVar(value=True)
        self.api_key_var = tk.StringVar()
        self._api_testing = False

        self._pushed_ids: set = set()
        self._scheduled_push_thread: Optional[threading.Thread] = None
        self._scheduled_push_running = False

        self._build()
        self._load_api_key()

    def _load_api_key(self):
        config_path = os.path.expanduser("~/pdd_supply_config.json")
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = _json.load(f)
                    api_key = config.get('zhipu_api_key', '')
                    if api_key:
                        self.api_key_var.set(api_key)
                        set_ai_api_key(api_key, self._log)
                        self._log("✅ 已加载保存的 API Key")
                        self._update_api_status(True)
        except Exception as e:
            self._log(f"⚠️ 加载配置失败: {e}")

    def _save_api_key(self):
        config_path = os.path.expanduser("~/pdd_supply_config.json")
        try:
            config = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = _json.load(f)
            config['zhipu_api_key'] = self.api_key_var.get()
            with open(config_path, 'w', encoding='utf-8') as f:
                _json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self._log(f"⚠️ 保存配置失败: {e}")
            return False

    def mount(self):
        self.nb.add(self.frame, text="  🏪 货源查找(手机)  ")
        threading.Thread(target=self._auto_init, daemon=True).start()
        self.frame.after(500, self._start_scheduled_push_thread)
        # 500ms 后注册全局滚轮（等控件全部渲染完）
        self.frame.after(600, self._bind_global_mousewheel)

    def _bind_global_mousewheel(self):
        """
        全局鼠标滚轮绑定策略：
        - 在顶层 frame 上监听 <MouseWheel>（Windows/macOS）和 <Button-4/5>（Linux）
        - 根据鼠标当前位置，把滚轮事件路由给它下面最近的可滚动控件
        - 支持上下滚动（普通滚轮）和左右滚动（Shift+滚轮）
        """
        root = self.frame.winfo_toplevel()

        def _find_scrollable(widget, horizontal=False):
            """向上遍历父链，找到第一个可滚动控件"""
            w = widget
            while w:
                try:
                    if horizontal:
                        # 有 xview 方法且不是顶层窗口
                        if hasattr(w, 'xview') and isinstance(
                                w, (ttk.Treeview, tk.Canvas, tk.Text, scrolledtext.ScrolledText)):
                            return w
                    else:
                        if hasattr(w, 'yview') and isinstance(
                                w, (ttk.Treeview, tk.Canvas, tk.Text, scrolledtext.ScrolledText)):
                            return w
                except Exception:
                    pass
                try:
                    w = w.master
                except Exception:
                    break
            return None

        def _on_global_wheel(event):
            # 找到鼠标下方的控件
            try:
                widget = event.widget
            except Exception:
                return
            target = _find_scrollable(widget, horizontal=False)
            if target:
                target.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _on_global_wheel_linux_up(event):
            try:
                widget = event.widget
            except Exception:
                return
            target = _find_scrollable(widget, horizontal=False)
            if target:
                target.yview_scroll(-1, "units")

        def _on_global_wheel_linux_down(event):
            try:
                widget = event.widget
            except Exception:
                return
            target = _find_scrollable(widget, horizontal=False)
            if target:
                target.yview_scroll(1, "units")

        def _on_global_hwheel(event):
            try:
                widget = event.widget
            except Exception:
                return
            target = _find_scrollable(widget, horizontal=True)
            if target:
                target.xview_scroll(int(-1 * (event.delta / 120)), "units")

        # 绑定到根窗口，所有子控件都会冒泡上来
        root.bind_all("<MouseWheel>",         _on_global_wheel,        add="+")
        root.bind_all("<Button-4>",           _on_global_wheel_linux_up,   add="+")
        root.bind_all("<Button-5>",           _on_global_wheel_linux_down, add="+")
        root.bind_all("<Shift-MouseWheel>",   _on_global_hwheel,       add="+")

    def _auto_init(self):
        SameProductMatcher.load_model(self._log)

        def update_model_status():
            try:
                if SameProductMatcher._loaded:
                    self.stats_label.config(text="最高利润率: — | 模型: ✅ 已加载")
                    self._log("✅ 语义模型已就绪")
                elif SameProductMatcher._error:
                    self.stats_label.config(text="最高利润率: — | 模型: ⚠️ 降级模式")
                    self._log(f"⚠️ 语义模型不可用: {SameProductMatcher._error}")
                else:
                    self.stats_label.config(text="最高利润率: — | 模型: ⏳ 加载中")
            except:
                pass

        self.frame.after(0, update_model_status)
        rand_sleep(0.5, 1.0)
        ctrl = PinduoduoMobileController(self._log)
        if ctrl.connect():
            self.controller = ctrl
            try:
                self.device_status.config(text="✅ 已连接", fg=self.SUCC)
                self.launch_btn.config(state=tk.NORMAL)
                self._start_processing()
            except Exception as e:
                self._log(f"⚠️ 自动启动失败: {e}")

    def _btn(self, p, text, cmd, bg=None, width=None, **kw):
        return tk.Button(p, text=text, command=cmd,
                         bg=bg or self.RAISED, fg=self.FG,
                         activebackground=self.BRD, relief=tk.FLAT, cursor="hand2",
                         font=("Microsoft YaHei", 9), padx=10, pady=4,
                         borderwidth=0, width=width, **kw)

    def _entry(self, p, width=8, default=''):
        e = tk.Entry(p, width=width, bg=self.INP, fg=self.FG,
                     insertbackground=self.ACC, relief=tk.FLAT,
                     highlightbackground=self.BRD, highlightthickness=1,
                     font=("Microsoft YaHei", 9))
        if default:
            e.insert(0, default)
        return e

    def _build(self):
        main = tk.PanedWindow(self.frame, orient=tk.HORIZONTAL, bg=self.BG,
                              sashwidth=4, sashrelief=tk.FLAT)
        main.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        left = tk.Frame(main, bg=self.BG, width=380)
        main.add(left, minsize=360, width=380)

        left_container = tk.Frame(left, bg=self.BG)
        left_container.pack(fill=tk.BOTH, expand=True)

        left_canvas = tk.Canvas(left_container, bg=self.BG, highlightthickness=0)
        left_scrollbar = ttk.Scrollbar(left_container, orient="vertical", command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_scrollbar.set)

        # 先 pack scrollbar 再 pack canvas，确保滚动条始终可见
        left_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        left_inner = tk.Frame(left_canvas, bg=self.BG)
        left_canvas_window = left_canvas.create_window((0, 0), window=left_inner, anchor="nw")

        def _configure_inner(event):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
            left_canvas.itemconfig(left_canvas_window, width=event.width)

        def _configure_canvas(event):
            left_canvas.itemconfig(left_canvas_window, width=event.width)

        left_inner.bind("<Configure>", _configure_inner)
        left_canvas.bind("<Configure>", _configure_canvas)

        def _on_mousewheel(event):
            left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_mousewheel_linux_up(event):
            left_canvas.yview_scroll(-1, "units")
        def _on_mousewheel_linux_down(event):
            left_canvas.yview_scroll(1, "units")

        # 绑定到 canvas 本身及 inner frame，确保鼠标悬停在内容上也能滚动
        for widget in (left_canvas, left_inner):
            widget.bind("<MouseWheel>",    _on_mousewheel)
            widget.bind("<Button-4>",      _on_mousewheel_linux_up)
            widget.bind("<Button-5>",      _on_mousewheel_linux_down)

        # 保存引用，mount 后用于全局绑定
        self._left_canvas = left_canvas
        self._left_inner  = left_inner

        # 统计卡片
        stats_frame = tk.LabelFrame(left_inner, text="📊 实时统计", bg=self.SURF, fg=self.FG_M,
                                     font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                     highlightbackground=self.BRD, highlightthickness=1)
        stats_frame.pack(fill=tk.X, padx=8, pady=(6, 4))

        stats_row = tk.Frame(stats_frame, bg=self.SURF)
        stats_row.pack(fill=tk.X, padx=8, pady=6)

        def stat_block(parent, label, init='0', color=self.ACC):
            blk = tk.Frame(parent, bg=self.SURF)
            blk.pack(side=tk.LEFT, expand=True)
            num_lbl = tk.Label(blk, text=init, bg=self.SURF, fg=color,
                               font=("Microsoft YaHei", 18, "bold"))
            num_lbl.pack()
            tk.Label(blk, text=label, bg=self.SURF, fg=self.FG_M,
                     font=("Microsoft YaHei", 8)).pack()
            return num_lbl

        self._stat_queue = stat_block(stats_row, "队列待处理", "0", self.DANGER)
        self._stat_done = stat_block(stats_row, "已处理", "0", self.SUCC)
        self._stat_profit = stat_block(stats_row, "有利润", "0", self.ACC)

        info_row = tk.Frame(stats_frame, bg=self.SURF)
        info_row.pack(fill=tk.X, padx=8, pady=(0, 6))
        self.countdown_label = tk.Label(info_row, text="⏱ 等待中", bg=self.SURF, fg=self.FG_M,
                                         font=("Microsoft YaHei", 8))
        self.countdown_label.pack(side=tk.LEFT)
        self.stats_label = tk.Label(info_row, text="最高利润率: — | 模型: ⏳加载中",
                                     bg=self.SURF, fg=self.FG_M, font=("Microsoft YaHei", 8))
        self.stats_label.pack(side=tk.RIGHT)

        # AI 设置卡片
        ai_frame = tk.LabelFrame(left_inner, text="🤖 AI 标题清洗（智谱 GLM-4-Flash）", bg=self.SURF, fg=self.FG_M,
                                  font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                  highlightbackground=self.BRD, highlightthickness=1)
        ai_frame.pack(fill=tk.X, padx=8, pady=(4, 4))

        ai_inner = tk.Frame(ai_frame, bg=self.SURF)
        ai_inner.pack(fill=tk.X, padx=6, pady=6)

        ai_check = tk.Checkbutton(ai_inner, text="启用 AI 清洗（API 失效时自动降级）",
                                   variable=self.use_ai_var, bg=self.SURF, fg=self.FG,
                                   selectcolor=self.SURF, activebackground=self.SURF)
        ai_check.pack(anchor="w", pady=(0, 5))

        api_row = tk.Frame(ai_inner, bg=self.SURF)
        api_row.pack(fill=tk.X, pady=3)
        tk.Label(api_row, text="API Key:", bg=self.SURF, fg=self.FG,
                 width=8, anchor='w').pack(side=tk.LEFT)
        self.api_entry = tk.Entry(api_row, bg=self.INP, fg=self.FG,
                                   show="*", relief=tk.FLAT,
                                   highlightbackground=self.BRD, highlightthickness=1,
                                   font=("Microsoft YaHei", 9))
        self.api_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.api_entry.bind('<KeyRelease>', lambda e: self.api_key_var.set(self.api_entry.get()))

        self.show_pwd = False
        self.toggle_pwd_btn = self._btn(api_row, "👁", self._toggle_password, bg=self.RAISED, width=3)
        self.toggle_pwd_btn.pack(side=tk.RIGHT)

        btn_row = tk.Frame(ai_inner, bg=self.SURF)
        btn_row.pack(fill=tk.X, pady=5)
        self.test_api_btn = self._btn(btn_row, "🔌 测试连接", self._test_api, width=9)
        self.test_api_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.save_api_btn = self._btn(btn_row, "💾 保存 Key", self._save_api_key_ui, width=9)
        self.save_api_btn.pack(side=tk.LEFT)

        self.api_status = tk.Label(ai_inner, text="⚪ 未配置 API Key", bg=self.SURF,
                                    fg=self.FG_M, font=("Microsoft YaHei", 8))
        self.api_status.pack(anchor="w", pady=(5, 0))

        self.api_stats_label = tk.Label(ai_inner, text="API 调用: 0次 | 成功: 0次", bg=self.SURF,
                                         fg=self.FG_M, font=("Microsoft YaHei", 8))
        self.api_stats_label.pack(anchor="w", pady=(2, 0))

        # 设备
        device_frame = tk.LabelFrame(left_inner, text="📱 设备", bg=self.SURF, fg=self.FG_M,
                                      font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                      highlightbackground=self.BRD, highlightthickness=1)
        device_frame.pack(fill=tk.X, padx=8, pady=(4, 4))

        dev_row = tk.Frame(device_frame, bg=self.SURF)
        dev_row.pack(fill=tk.X, padx=6, pady=4)
        self.launch_btn = self._btn(dev_row, "▶ 启动拼多多", self._launch_app, state=tk.DISABLED, width=12)
        self.launch_btn.pack(side=tk.LEFT)
        self.device_status = tk.Label(dev_row, text="⏳ 自动连接中...", bg=self.SURF, fg=self.FG_M,
                                       font=("Microsoft YaHei", 8))
        self.device_status.pack(side=tk.LEFT, padx=(10, 0))

        dev_row2 = tk.Frame(device_frame, bg=self.SURF)
        dev_row2.pack(fill=tk.X, padx=6, pady=(0, 6))
        self._btn(dev_row2, "🔄 重连设备", self._reconnect_device, width=10).pack(side=tk.LEFT)
        self._btn(dev_row2, "🧠 重载模型", self._reload_model, width=10).pack(side=tk.LEFT, padx=(6, 0))

        # 参数
        param_frame = tk.LabelFrame(left_inner, text="⚙️ 参数设置", bg=self.SURF, fg=self.FG_M,
                                     font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                     highlightbackground=self.BRD, highlightthickness=1)
        param_frame.pack(fill=tk.X, padx=8, pady=(4, 4))

        param_grid = tk.Frame(param_frame, bg=self.SURF)
        param_grid.pack(fill=tk.X, padx=6, pady=6)

        params = [
            ("相似阈值:", "sim_thresh_e", "0.8", 6),
            ("标题翻页:", "scroll_pages_e", "5", 6),
            ("最多采集:", "max_items_e", "20", 6),
            ("图搜翻页:", "img_scroll_pages_e", "3", 6),
            ("间隔(秒):", "delay_between_e", "8", 6),
            ("推送评分:", "score_thresh_e", "80", 6),
            ("每N件休息:", "pause_every_e", "5", 6),
            ("休息(秒):", "pause_dur_e", "60", 6),
        ]

        for i, (label, attr, default, width) in enumerate(params):
            row, col = i // 2, i % 2
            frame = tk.Frame(param_grid, bg=self.SURF)
            frame.grid(row=row, column=col, sticky="ew", padx=4, pady=3)
            tk.Label(frame, text=label, bg=self.SURF, fg=self.FG,
                     font=("Microsoft YaHei", 9), width=9, anchor='w').pack(side=tk.LEFT)
            entry = self._entry(frame, width=width, default=default)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            setattr(self, attr, entry)

        param_grid.columnconfigure(0, weight=1)
        param_grid.columnconfigure(1, weight=1)

        toggle_row = tk.Frame(param_frame, bg=self.SURF)
        toggle_row.pack(fill=tk.X, padx=6, pady=(2, 4))
        self.use_img_search_var = tk.BooleanVar(value=True)
        self.use_ai_compare_var = tk.BooleanVar(value=True)
        tk.Checkbutton(toggle_row, text="📷 启用以图搜款",
                       variable=self.use_img_search_var,
                       bg=self.SURF, fg=self.FG, font=("Microsoft YaHei", 9),
                       selectcolor=self.SURF).pack(side=tk.LEFT)
        tk.Checkbutton(toggle_row, text="🤖 AI批量同款比对",
                       variable=self.use_ai_compare_var,
                       bg=self.SURF, fg=self.FG, font=("Microsoft YaHei", 9),
                       selectcolor=self.SURF).pack(side=tk.LEFT, padx=(12, 0))

        update_btn_frame = tk.Frame(param_frame, bg=self.SURF)
        update_btn_frame.pack(fill=tk.X, padx=6, pady=(0, 6))

        self.update_settings_btn = self._btn(update_btn_frame, "🔄 更新设置",
                                              self._update_scheduler_settings,
                                              bg=self.ACC, width=12)
        self.update_settings_btn.config(fg="white")
        self.update_settings_btn.pack(side=tk.RIGHT)
        tk.Label(update_btn_frame, text="修改参数后点击更新，立即生效", bg=self.SURF, fg=self.FG_M,
                 font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)

        # 定时推送
        sched_frame = tk.LabelFrame(left_inner, text="⏰ 定时推送 & 保护设置", bg=self.SURF, fg=self.FG_M,
                                     font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                     highlightbackground=self.BRD, highlightthickness=1)
        sched_frame.pack(fill=tk.X, padx=8, pady=(4, 4))

        sched_inner = tk.Frame(sched_frame, bg=self.SURF)
        sched_inner.pack(fill=tk.X, padx=6, pady=6)

        r_sched = tk.Frame(sched_inner, bg=self.SURF)
        r_sched.pack(fill=tk.X, pady=2)
        self.enable_scheduled_push_var = tk.BooleanVar(value=True)
        tk.Checkbutton(r_sched, text="每天", variable=self.enable_scheduled_push_var,
                       bg=self.SURF, fg=self.FG, font=("Microsoft YaHei", 9),
                       selectcolor=self.SURF).pack(side=tk.LEFT)
        self.sched_hour_e = self._entry(r_sched, width=4, default="8")
        self.sched_hour_e.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(r_sched, text="时", bg=self.SURF, fg=self.FG,
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)
        self.sched_min_e = self._entry(r_sched, width=4, default="0")
        self.sched_min_e.pack(side=tk.LEFT, padx=(2, 2))
        tk.Label(r_sched, text="分 自动推送正利润货源", bg=self.SURF, fg=self.FG_M,
                 font=("Microsoft YaHei", 8)).pack(side=tk.LEFT, padx=(2, 0))

        self.sched_status_lbl = tk.Label(sched_inner, text="⏳ 定时推送线程待启动",
                                          bg=self.SURF, fg=self.FG_M, font=("Microsoft YaHei", 8))
        self.sched_status_lbl.pack(anchor="w", pady=(2, 4))

        r_empty = tk.Frame(sched_inner, bg=self.SURF)
        r_empty.pack(fill=tk.X, pady=2)
        tk.Label(r_empty, text="连续", bg=self.SURF, fg=self.FG,
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)
        self.empty_threshold_e = self._entry(r_empty, width=4, default="3")
        self.empty_threshold_e.pack(side=tk.LEFT, padx=(4, 4))
        tk.Label(r_empty, text="次采集到0个商品 → 触发风控告警+截图+暂停",
                 bg=self.SURF, fg=self.FG_M, font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)

        r_cache = tk.Frame(sched_inner, bg=self.SURF)
        r_cache.pack(fill=tk.X, pady=2)
        tk.Label(r_cache, text="运行超过", bg=self.SURF, fg=self.FG,
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)
        self.cache_clear_interval_e = self._entry(r_cache, width=4, default="120")
        self.cache_clear_interval_e.pack(side=tk.LEFT, padx=(4, 4))
        tk.Label(r_cache, text="分钟 → 清理后台并重启PDD", bg=self.SURF, fg=self.FG_M,
                 font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)

        # 操作按钮（增加继续处理按钮）
        action_frame = tk.LabelFrame(left_inner, text="🎮 操作", bg=self.SURF, fg=self.FG_M,
                                      font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                      highlightbackground=self.BRD, highlightthickness=1)
        action_frame.pack(fill=tk.X, padx=8, pady=(4, 4))

        row1 = tk.Frame(action_frame, bg=self.SURF)
        row1.pack(fill=tk.X, padx=6, pady=(4, 2))
        self.auto_push_var = tk.BooleanVar(value=True)
        tk.Checkbutton(row1, text="自动推送高分商品", variable=self.auto_push_var,
                       bg=self.SURF, fg=self.FG, font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)
        self._btn(row1, "📥 推送当前商品", self._push_current_items, width=12).pack(side=tk.RIGHT)

        row2 = tk.Frame(action_frame, bg=self.SURF)
        row2.pack(fill=tk.X, padx=6, pady=(2, 4))
        self.start_btn = self._btn(row2, "▶ 开始处理", self._start_processing,
                                   bg=self.SUCC, state=tk.DISABLED, width=10)
        self.start_btn.config(fg="white")
        self.start_btn.pack(side=tk.LEFT, padx=(0, 4))
        
        self.stop_btn = self._btn(row2, "⏹ 停止", self._stop_processing,
                                  bg=self.DANGER, state=tk.DISABLED, width=6)
        self.stop_btn.config(fg="white")
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 4))
        
        # 新增：继续处理按钮（用于风控恢复）
        self.resume_btn = self._btn(row2, "▶ 继续处理", self._resume_processing,
                                    bg=self.ACC, state=tk.DISABLED, width=10)
        self.resume_btn.config(fg="white")
        self.resume_btn.pack(side=tk.LEFT, padx=(0, 4))

        export_frame = tk.Frame(row2, bg=self.SURF)
        export_frame.pack(side=tk.RIGHT)
        self._btn(export_frame, "📊 导出全部", self._export_all, width=9).pack(side=tk.LEFT, padx=(0, 2))
        self._btn(export_frame, "💰 导出正利润", self._export_profit_only,
                  bg="#FFF3CD", width=10).pack(side=tk.LEFT, padx=(0, 2))
        self._btn(export_frame, "📋 导出上架建议", self._export_listing_advice,
                  bg="#E8F0FB", width=11).pack(side=tk.LEFT)

        # 日志区
        log_frame = tk.LabelFrame(left_inner, text="📋 运行日志", bg=self.SURF, fg=self.FG_M,
                                   font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                   highlightbackground=self.BRD, highlightthickness=1)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 6))

        self.log_text = scrolledtext.ScrolledText(
            log_frame, bg="#1E1E2E", fg="#CDD6F4",
            font=("Consolas", 8), relief=tk.FLAT,
            insertbackground="white", wrap=tk.WORD, height=12)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # 右侧区域
        right = tk.Frame(main, bg=self.BG)
        main.add(right, minsize=500)

        right_pane = tk.PanedWindow(right, orient=tk.VERTICAL, bg=self.BG,
                                    sashwidth=4, sashrelief=tk.FLAT)
        right_pane.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        t1_frame = tk.Frame(right_pane, bg=self.BG)
        right_pane.add(t1_frame, minsize=180)
        tk.Label(t1_frame, text="📋 货源结果（双击商品行查看详情）",
                 bg=self.BG, fg=self.FG_M, font=("Microsoft YaHei", 8)).pack(anchor='w')

        t1_cols = ('闲鱼标题', '售价', '闲鱼评分', '象限', '综合利润', '标题同款利润', '标题平替利润', '图搜同款利润', '图搜平替利润', '最优货源', '进价', '利润率%', '建议', '闲鱼链接', '图片链接')
        t1_widths = (120, 46, 50, 72, 62, 72, 72, 72, 72, 150, 46, 56, 80, 180, 160)
        self.tree = self._make_tree(t1_frame, t1_cols, t1_widths, height=10)
        self.tree.tag_configure('q1', background='#E8F8EF')
        self.tree.tag_configure('q2', background='#E8F0FB')
        self.tree.tag_configure('q3', background='#FFF9E6')
        self.tree.tag_configure('q4', background='#FDECEC')
        self.tree.tag_configure('best', background='#FFF3CD')
        self.tree.tag_configure('profit', background='#EAF7EF')
        make_sortable(self.tree, {'售价', '闲鱼评分', '进价', '综合利润', '标题同款利润', '标题平替利润', '图搜同款利润', '图搜平替利润', '利润率%'})
        self.tree.bind('<<TreeviewSelect>>', self._on_select)
        self.tree.bind('<Double-Button-1>', self._on_tree_double_click)

        t2_frame = tk.Frame(right_pane, bg=self.BG)
        right_pane.add(t2_frame, minsize=240)

        tk.Label(t2_frame, text="📋 标题搜索同款/平替（双击商品行查看详情）",
                 bg=self.BG, fg=self.FG_M, font=("Microsoft YaHei", 8)).pack(anchor='w')
        t2_cols = ('#', '货源商品名', '匹配类型', '进价', '利润', '利润率%', '货源评分', '销量', '相似度', '比对', '推荐', '理由')
        t2_widths = (28, 150, 70, 52, 58, 62, 58, 76, 56, 50, 56, 180)
        self.detail_tree = self._make_tree(t2_frame, t2_cols, t2_widths, height=4)
        self.detail_tree.tag_configure('best_src', background='#FFF3CD')
        self.detail_tree.tag_configure('good_src', background='#EAF7EF')
        make_sortable(self.detail_tree, {'进价', '利润', '利润率%', '货源评分', '相似度'})
        self.detail_tree.bind('<Double-Button-1>', self._on_detail_double_click)

        tk.Label(t2_frame, text="📷 以图搜款同款/平替（双击商品行查看详情）",
                 bg=self.BG, fg=self.FG_M, font=("Microsoft YaHei", 8)).pack(anchor='w', pady=(6, 0))
        img_cols = ('#', '货源商品名', '匹配类型', '进价', '利润', '利润率%', '货源评分', '销量', '相似度', '比对', '推荐', '理由')
        img_widths = (28, 150, 70, 52, 58, 62, 58, 76, 56, 50, 56, 180)
        self.img_detail_tree = self._make_tree(t2_frame, img_cols, img_widths, height=4)
        self.img_detail_tree.tag_configure('best_src', background='#E8F0FB')
        self.img_detail_tree.tag_configure('good_src', background='#E8F8EF')
        make_sortable(self.img_detail_tree, {'进价', '利润', '利润率%', '货源评分', '相似度'})
        self.img_detail_tree.bind('<Double-Button-1>', self._on_img_detail_double_click)

        media_frame = tk.LabelFrame(t2_frame, text="🖼 闲鱼商品媒体信息",
                                     bg=self.SURF, fg=self.FG_M,
                                     font=("Microsoft YaHei", 8), relief=tk.FLAT,
                                     highlightbackground=self.BRD, highlightthickness=1)
        media_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=(4, 2))

        media_canvas = tk.Canvas(media_frame, bg=self.SURF, highlightthickness=0)
        media_scroll = ttk.Scrollbar(media_frame, orient="vertical", command=media_canvas.yview)
        media_canvas.configure(yscrollcommand=media_scroll.set)
        media_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        media_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._media_inner = tk.Frame(media_canvas, bg=self.SURF)
        self._media_window = media_canvas.create_window((0, 0), window=self._media_inner, anchor="nw")

        def _on_media_configure(event):
            media_canvas.configure(scrollregion=media_canvas.bbox("all"))
            media_canvas.itemconfig(self._media_window, width=event.width)

        media_canvas.bind("<Configure>", _on_media_configure)
        self._media_inner.bind("<Configure>", lambda e: media_canvas.configure(
            scrollregion=media_canvas.bbox("all")))

        # 鼠标滚轮支持
        def _media_wheel(e):
            media_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        media_canvas.bind("<MouseWheel>", _media_wheel)
        media_canvas.bind("<Button-4>", lambda e: media_canvas.yview_scroll(-1, "units"))
        media_canvas.bind("<Button-5>", lambda e: media_canvas.yview_scroll(1, "units"))
        self._media_canvas = media_canvas

        self._media_placeholder = tk.Label(self._media_inner,
            text="← 点击左侧货源行，查看闲鱼商品链接和图片",
            bg=self.SURF, fg=self.FG_M, font=("Microsoft YaHei", 9))
        self._media_placeholder.pack(pady=10)

        self._update_api_stats_display()

    def _make_tree(self, parent, cols, widths, height=8) -> ttk.Treeview:
        tf = tk.Frame(parent, bg=self.SURF, highlightbackground=self.BRD, highlightthickness=1)
        tf.pack(fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(tf, orient="vertical")
        hsb = ttk.Scrollbar(tf, orient="horizontal")
        tree = ttk.Treeview(tf, columns=cols, show='headings', height=height,
                            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        # 滚动条始终显示：先 pack vsb/hsb
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(fill=tk.BOTH, expand=True)
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, minwidth=25, anchor='w')
        # 鼠标滚轮纵向滚动（Windows/macOS），支持直接在 tree 上滚动
        def _wheel(e):
            tree.yview_scroll(int(-1 * (e.delta / 120)), "units")
        def _wheel_linux_up(e):
            tree.yview_scroll(-1, "units")
        def _wheel_linux_down(e):
            tree.yview_scroll(1, "units")
        tree.bind("<MouseWheel>",    _wheel)
        tree.bind("<Button-4>",      _wheel_linux_up)
        tree.bind("<Button-5>",      _wheel_linux_down)
        # shift+滚轮 → 横向滚动
        def _hwheel(e):
            tree.xview_scroll(int(-1 * (e.delta / 120)), "units")
        tree.bind("<Shift-MouseWheel>", _hwheel)
        return tree

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        stamped = f"[{ts}] {msg}"

        def _u():
            self.log_text.insert(tk.END, stamped + "\n")
            self.log_text.see(tk.END)

        try:
            self.frame.after(0, _u)
        except Exception:
            pass

    def _update_countdown(self, remaining: int):
        def _u():
            if remaining > 0:
                self.countdown_label.config(text=f"⏱ 倒计时: {remaining}s", fg=self.DANGER)
            else:
                self.countdown_label.config(text="⏱ 等待中", fg=self.FG_M)

        try:
            self.frame.after(0, _u)
        except Exception:
            pass

    def _toggle_password(self):
        self.show_pwd = not self.show_pwd
        if self.show_pwd:
            self.api_entry.config(show="")
            self.toggle_pwd_btn.config(text="🙈")
        else:
            self.api_entry.config(show="*")
            self.toggle_pwd_btn.config(text="👁")

    def _test_api(self):
        if self._api_testing:
            return

        api_key = self.api_entry.get().strip()
        if not api_key:
            messagebox.showwarning("提示", "请先输入 API Key")
            return

        self._api_testing = True
        self.test_api_btn.config(state=tk.DISABLED, text="⏳ 测试中...")
        self._log("🔌 正在测试 API 连接...")

        def do_test():
            try:
                cleaner = TitleCleanerAI(api_key=api_key, log_cb=self._log)
                success, msg = cleaner.test_connection()

                def update_ui():
                    if success:
                        self._log("✅ API 连接成功！")
                        self.api_status.config(text="✅ API 可用", fg=self.SUCC)
                        set_ai_api_key(api_key, self._log)
                        self.api_key_var.set(api_key)
                    else:
                        self._log(f"❌ API 连接失败: {msg}")
                        self.api_status.config(text=f"❌ {msg[:30]}", fg=self.DANGER)
                    self.test_api_btn.config(state=tk.NORMAL, text="🔌 测试连接")
                    self._api_testing = False

                self.frame.after(0, update_ui)
            except Exception as e:
                self.frame.after(0, lambda: self._log(f"❌ 测试异常: {e}"))
                self.frame.after(0, lambda: self.test_api_btn.config(state=tk.NORMAL, text="🔌 测试连接"))
                self.frame.after(0, lambda: setattr(self, '_api_testing', False))

        threading.Thread(target=do_test, daemon=True).start()

    def _save_api_key_ui(self):
        api_key = self.api_entry.get().strip()
        if not api_key:
            messagebox.showwarning("提示", "API Key 不能为空")
            return

        if self._save_api_key():
            set_ai_api_key(api_key, self._log)
            self.api_key_var.set(api_key)
            self._log("✅ API Key 已保存")
            self._update_api_status(True)
            messagebox.showinfo("成功", "API Key 已保存")
        else:
            messagebox.showerror("错误", "保存失败")

    def _update_api_status(self, is_valid: bool = None):
        if is_valid is None:
            is_valid = bool(self.api_key_var.get())

        if is_valid:
            self.api_status.config(text="✅ API 已配置", fg=self.SUCC)
        else:
            self.api_status.config(text="⚪ 未配置 API Key（将使用本地清洗）", fg=self.FG_M)

    def _update_api_stats_display(self):
        try:
            cleaner = get_ai_cleaner()
            stats = cleaner.get_stats()
            clean_calls = stats.get('clean_calls', 0)
            clean_ok = stats.get('clean_ok', 0)
            compare_calls = stats.get('compare_calls', 0)
            compare_ok = stats.get('compare_ok', 0)
            self.api_stats_label.config(
                text=(f"清洗: {clean_ok}/{clean_calls}次 | "
                      f"比对: {compare_ok}/{compare_calls}次")
            )
        except Exception:
            pass
        try:
            self.frame.after(2000, self._update_api_stats_display)
        except Exception:
            pass

    def _get_params(self):
        try:
            return {
                'score_threshold': int(self.score_thresh_e.get() or 80),
                'sim_threshold': float(self.sim_thresh_e.get() or 0.8),
                'scroll_pages': int(self.scroll_pages_e.get() or 5),
                'max_items': int(self.max_items_e.get() or 20),
                'img_scroll_pages': int(self.img_scroll_pages_e.get() or 3),
                'delay_between_products': int(self.delay_between_e.get() or 8),
                'pause_every': int(self.pause_every_e.get() or 5),
                'pause_duration': int(self.pause_dur_e.get() or 60),
                'cache_clear_interval_min': int(self.cache_clear_interval_e.get() or 120),
                'empty_threshold': int(self.empty_threshold_e.get() or 3),
                'use_img_search': self.use_img_search_var.get(),
                'use_ai_compare': self.use_ai_compare_var.get(),
            }
        except Exception as ex:
            messagebox.showerror("参数错误", str(ex))
            return None

    def _update_scheduler_settings(self):
        if not self.scheduler:
            self._log("⚠️ 调度器未启动，请先点击「开始处理」")
            return

        params = self._get_params()
        if not params:
            return

        empty_threshold = max(1, min(20, params['empty_threshold']))
        cache_clear_interval = max(1, min(360, params['cache_clear_interval_min']))

        self.scheduler.score_threshold = params['score_threshold']
        self.scheduler.sim_threshold = params['sim_threshold']
        self.scheduler.scroll_pages = params['scroll_pages']
        self.scheduler.max_items = params['max_items']
        self.scheduler.img_scroll_pages = params['img_scroll_pages']
        self.scheduler.use_img_search = params['use_img_search']
        self.scheduler.use_ai_compare = params['use_ai_compare']
        self.scheduler.delay_between_products = params['delay_between_products']
        self.scheduler.pause_every = params['pause_every']
        self.scheduler.pause_duration = params['pause_duration']
        self.scheduler.cache_clear_interval_min = cache_clear_interval
        self.scheduler._empty_threshold = empty_threshold
        self.scheduler.use_ai_clean = self.use_ai_var.get()

        if hasattr(self.scheduler, 'matcher'):
            self.scheduler.matcher.threshold = params['sim_threshold']

        api_key = self.api_key_var.get()
        if api_key and self.use_ai_var.get():
            set_ai_api_key(api_key, self._log)

        self._log("=" * 50)
        self._log("🔧 设置更新（以下为实际生效值）：")
        self._log(f"  相似阈值:{self.scheduler.sim_threshold} | "
                  f"标题翻页:{self.scheduler.scroll_pages} | "
                  f"图搜翻页:{self.scheduler.img_scroll_pages} | "
                  f"最多采集:{self.scheduler.max_items}")
        self._log(f"  间隔:{self.scheduler.delay_between_products}s | "
                  f"评分:{self.scheduler.score_threshold} | "
                  f"每N休息:{self.scheduler.pause_every} | "
                  f"休息:{self.scheduler.pause_duration}s")
        self._log(f"  📷图搜:{self.scheduler.use_img_search} | "
                  f"🤖AI比对:{self.scheduler.use_ai_compare} | "
                  f"🤖AI清洗:已禁用(需求1-全标题搜索)")
        self._log(f"  ⚠️风控告警阈值(第1阶段): {self.scheduler._empty_threshold}次 → 清缓存")
        self._log(f"  ⚠️风控告警阈值(第2阶段): {self.scheduler._empty_threshold}次 → 告警+暂停")
        self._log(f"  🧹定时缓存清理间隔: {self.scheduler.cache_clear_interval_min}分钟")
        self._log(f"  ⏱️清缓存后等待: {self.scheduler._post_cache_wait_sec}秒")
        self._log("✅ 所有参数已立即生效")
        self._log("=" * 50)

    def _launch_app(self):
        if not self.controller:
            messagebox.showwarning("提示", "手机未连接")
            return

        def do_launch():
            if self.controller.launch_pinduoduo():
                try:
                    self.start_btn.config(state=tk.NORMAL)
                except Exception:
                    pass

        threading.Thread(target=do_launch, daemon=True).start()

    def _reconnect_device(self):
        self._log("🔄 重新连接手机...")
        self.device_status.config(text="⏳ 连接中...", fg=self.FG_M)

        def do_reconnect():
            try:
                ctrl = PinduoduoMobileController(self._log)
                if ctrl.connect():
                    self.controller = ctrl
                    if self.scheduler:
                        self.scheduler.controller = ctrl
                    self.frame.after(0, lambda: self.device_status.config(
                        text="✅ 已连接", fg=self.SUCC))
                    self.frame.after(0, lambda: self.launch_btn.config(state=tk.NORMAL))
                    self._log("✅ 设备重连成功")
                else:
                    self.frame.after(0, lambda: self.device_status.config(
                        text="❌ 连接失败", fg=self.DANGER))
            except Exception as e:
                self._log(f"❌ 重连异常: {e}")
                self.frame.after(0, lambda: self.device_status.config(
                    text="❌ 连接失败", fg=self.DANGER))

        threading.Thread(target=do_reconnect, daemon=True).start()

    def _reload_model(self):
        self._log("🔄 重新加载语义模型...")
        SameProductMatcher._loaded = False
        SameProductMatcher._model = None
        SameProductMatcher._error = None

        def do_reload():
            ok = SameProductMatcher.load_model(self._log)

            def upd():
                try:
                    if ok:
                        self.stats_label.config(
                            text=f"最高利润率: — | 模型: ✅ 已加载")
                    else:
                        self.stats_label.config(
                            text=f"最高利润率: — | 模型: ⚠️ 降级模式")
                except Exception:
                    pass

            self.frame.after(0, upd)

        threading.Thread(target=do_reload, daemon=True).start()

    def _resume_processing(self):
        """继续处理队列（风控恢复后调用）"""
        if not self.scheduler:
            self._log("⚠️ 调度器未启动，请先点击「开始处理」")
            return

        if self.scheduler.is_paused():
            self.scheduler.resume()
            self._log("▶ 已发送恢复指令，队列继续处理")
            # 更新按钮状态
            self.resume_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
        else:
            self._log("ℹ️ 队列未处于暂停状态，无需恢复")

    def _start_processing(self):
        if not self.controller:
            messagebox.showerror("错误", "手机未连接，请点击「重连设备」")
            return

        params = self._get_params()
        if not params:
            return

        empty_threshold = max(1, min(20, params['empty_threshold']))

        api_key = self.api_key_var.get()
        if api_key and self.use_ai_var.get():
            set_ai_api_key(api_key, self._log)

        if self.scheduler and self.scheduler._running:
            self.scheduler.stop()
            time.sleep(0.5)

        self.scheduler = MobileSupplyScheduler(
            task_queue=self.task_queue,
            result_cb=self._on_result,
            log_cb=self._log,
            countdown_cb=self._update_countdown,
            use_ai_clean=self.use_ai_var.get(),
            use_ai_compare=params['use_ai_compare'],
            use_img_search=params['use_img_search'],
            score_threshold=params['score_threshold'],
            sim_threshold=params['sim_threshold'],
            scroll_pages=params['scroll_pages'],
            max_items=params['max_items'],
            img_scroll_pages=params['img_scroll_pages'],
            delay_between_products=params['delay_between_products'],
            pause_every=params['pause_every'],
            pause_duration=params['pause_duration'],
            cache_clear_interval_min=params['cache_clear_interval_min'],
            empty_threshold=empty_threshold,
        )
        self.scheduler.controller = self.controller

        if hasattr(self, '_wechat_webhook') and self._wechat_webhook:
            self.scheduler._wechat_webhook = self._wechat_webhook

        self.scheduler.start_processing()
        self.start_btn.config(text="处理中...", state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.resume_btn.config(state=tk.NORMAL)  # 启用继续按钮
        self._log("🚀 货源查找队列已启动（标题搜索 + 以图搜款双路径，支持同款/平替分级）")

    def _stop_processing(self):
        if self.scheduler:
            self.scheduler.stop()
        self.stop_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.NORMAL)
        self.resume_btn.config(state=tk.DISABLED)
        self._update_countdown(0)

    def _push_current_items(self):
        all_data = getattr(self.app, 'all_data', [])
        thresh = int(self.score_thresh_e.get() or 80)
        pushed = 0
        for item in all_data:
            try:
                if int(item.get('综合评分', 0)) >= thresh:
                    self.task_queue.put(item)
                    pushed += 1
            except Exception:
                pass
        self._log(f"📥 推送 {pushed} 件，队列深度: {self.task_queue.qsize()}")
        self._update_stats()

    def notify_new_item(self, item: dict):
        if not self.auto_push_var.get():
            return

        try:
            score_thresh = int(self.score_thresh_e.get() or 80)
        except:
            score_thresh = 80

        try:
            if int(item.get('综合评分', 0)) >= score_thresh:
                self.task_queue.put(item)
                self._log(f"📥 自动推送到队列: {item.get('商品标题', '')[:30]} ({item.get('综合评分', 0)}分)")
                self._update_stats()
        except Exception as e:
            self._log(f"⚠️ 自动推送失败: {e}")

    def _on_result(self, record: dict):
        self._result_records.append(record)

        def _u():
            title_items = record.get('pdd_items', [])
            img_items = record.get('img_pdd_items', [])
            title_best = title_items[0] if title_items else None
            img_best = img_items[0] if img_items else None

            q = record.get('quadrant', '')
            tag_map = {'Q1': 'q1', 'Q2': 'q2', 'Q3': 'q3', 'Q4': 'q4', 'Q5': 'q4'}
            tag = tag_map.get(q, 'profit' if record.get('final_profit', 0) and record['final_profit'] > 0 else '')

            pics_raw = record.get('xianyu_pics', '')
            first_pic = pics_raw.split(',')[0] if pics_raw else ''

            fp = record.get('final_profit')
            tp = record.get('title_profit')
            ip = record.get('img_profit')
            fnp = record.get('final_price')
            tsp = record.get('title_same_profit')
            tap = record.get('title_alt_profit')
            isp = record.get('img_same_profit')
            iap = record.get('img_alt_profit')

            rate_str = ''
            if fp is not None and fnp and fnp > 0:
                rate_str = f"{fp/fnp*100:.1f}%"

            self.tree.insert('', 0, values=(
                record['source_title'][:18],
                record['xianyu_price'],
                record.get('xianyu_score', ''),
                f"{record.get('quadrant_emoji', '')} {q} {record.get('quadrant_label', '')}",
                f"¥{fp:.1f}" if fp is not None else '—',
                f"¥{tsp:.1f}" if tsp is not None else '—',
                f"¥{tap:.1f}" if tap is not None else '—',
                f"¥{isp:.1f}" if isp is not None else '—',
                f"¥{iap:.1f}" if iap is not None else '—',
                (title_best['goods_name'][:22] if title_best else
                 img_best['goods_name'][:22] if img_best else ''),
                f"¥{fnp:.1f}" if fnp is not None else '—',
                rate_str,
                record.get('recommendation', '')[:30],
                record.get('xianyu_link', ''),
                first_pic,
            ), tags=(tag,))
            self._update_stats()

        self.frame.after(0, _u)

    def _on_select(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], 'values')
        title_prefix = vals[0]

        record = next((r for r in self._result_records
                       if r['source_title'][:18] == title_prefix), None)
        if not record:
            return

        for row in self.detail_tree.get_children():
            self.detail_tree.delete(row)
        for row in self.img_detail_tree.get_children():
            self.img_detail_tree.delete(row)

        def _fill_tree(tree, items, color_best, color_good):
            for rank, it in enumerate(items, 1):
                tag = color_best if rank == 1 else (
                    color_good if (it.get('货源评分', 0) or 0) >= 50 else '')
                match_type = it.get('匹配类型', '')
                match_type_display = match_type
                if match_type == '同款':
                    match_type_display = '✅同款'
                elif match_type == '平替':
                    match_type_display = '🔄平替'
                elif match_type == '部分匹配':
                    match_type_display = '⚠️部分'

                reason = it.get('匹配说明', '')[:80]

                tree.insert('', 'end', values=(
                    f"#{rank}",
                    it.get('goods_name', '')[:35],
                    match_type_display,
                    it.get('拼多多进价(元)', ''),
                    it.get('预估利润(元)', ''),
                    it.get('利润率(%)', ''),
                    it.get('货源评分', ''),
                    it.get('sales_tip', ''),
                    it.get('sim_score', 0),
                    it.get('match_src', ''),
                    it.get('是否推荐货源', ''),
                    reason,
                ), tags=(tag,))

        _fill_tree(self.detail_tree,
                   record.get('pdd_items', []),
                   'best_src', 'good_src')
        _fill_tree(self.img_detail_tree,
                   record.get('img_pdd_items', []),
                   'best_src', 'good_src')

        self._rebuild_media_panel(record)

    def _on_tree_double_click(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], 'values')
        title_prefix = vals[0]

        record = next((r for r in self._result_records
                       if r['source_title'][:18] == title_prefix), None)
        if not record:
            return

        title_items = record.get('pdd_items', [])
        img_items = record.get('img_pdd_items', [])

        if title_items or img_items:
            self._show_record_detail(record)

    def _on_detail_double_click(self, event):
        sel = self.detail_tree.selection()
        if not sel:
            return

        values = self.detail_tree.item(sel[0], 'values')
        if not values:
            return

        tree_sel = self.tree.selection()
        if not tree_sel:
            return
        vals = self.tree.item(tree_sel[0], 'values')
        title_prefix = vals[0]

        record = next((r for r in self._result_records
                       if r['source_title'][:18] == title_prefix), None)
        if not record:
            return

        title_items = record.get('pdd_items', [])
        rank_str = values[0]
        try:
            rank = int(rank_str.replace('#', ''))
            if 1 <= rank <= len(title_items):
                selected_item = [title_items[rank - 1]]
                SourceDetailWindow(self.frame, f"标题搜索货源详情", selected_item, "标题搜索")
        except (ValueError, IndexError):
            pass

    def _on_img_detail_double_click(self, event):
        sel = self.img_detail_tree.selection()
        if not sel:
            return

        values = self.img_detail_tree.item(sel[0], 'values')
        if not values:
            return

        tree_sel = self.tree.selection()
        if not tree_sel:
            return
        vals = self.tree.item(tree_sel[0], 'values')
        title_prefix = vals[0]

        record = next((r for r in self._result_records
                       if r['source_title'][:18] == title_prefix), None)
        if not record:
            return

        img_items = record.get('img_pdd_items', [])
        rank_str = values[0]
        try:
            rank = int(rank_str.replace('#', ''))
            if 1 <= rank <= len(img_items):
                selected_item = [img_items[rank - 1]]
                SourceDetailWindow(self.frame, f"以图搜款货源详情", selected_item, "以图搜款")
        except (ValueError, IndexError):
            pass

    def _show_record_detail(self, record: dict):
        """双击货源行弹出纵向阅读卡片详情窗口"""
        title_items = record.get('pdd_items', [])
        img_items   = record.get('img_pdd_items', [])

        win = tk.Toplevel(self.frame)
        win.title(f"货源详情 — {record['source_title'][:35]}")
        win.geometry("780x680")
        win.configure(bg="#F5F6FA")
        win.transient(self.frame)

        # ── 顶部：闲鱼商品概要 ──────────────────────────────────
        header = tk.Frame(win, bg="#FF6B35")
        header.pack(fill=tk.X)

        tk.Label(header, text="📦 货源详情", bg="#FF6B35", fg="white",
                 font=("Microsoft YaHei", 11, "bold")).pack(side=tk.LEFT, padx=14, pady=10)

        q_emoji = record.get('quadrant_emoji', '')
        q_label = record.get('quadrant_label', '')
        fp      = record.get('final_profit')
        fp_str  = f"综合利润 ¥{fp:.1f}" if fp is not None else "无利润数据"
        tk.Label(header, text=f"{q_emoji} {q_label}  {fp_str}",
                 bg="#FF6B35", fg="white",
                 font=("Microsoft YaHei", 9)).pack(side=tk.RIGHT, padx=14)

        # ── 闲鱼标题 + 象限建议行 ──────────────────────────────
        summary_f = tk.Frame(win, bg="#FFFFFF",
                             highlightbackground="#E2E5F0", highlightthickness=1)
        summary_f.pack(fill=tk.X, padx=10, pady=(8, 4))

        tk.Label(summary_f,
                 text=f"  闲鱼标题：{record['source_title']}",
                 bg="#FFFFFF", fg="#3D4152",
                 font=("Microsoft YaHei", 9, "bold"),
                 anchor='w', wraplength=740).pack(fill=tk.X, padx=6, pady=(6, 2))
        tk.Label(summary_f,
                 text=f"  售价 ¥{record['xianyu_price']}  |  搜索词：{record.get('search_keyword','')}  |  {record.get('recommendation','')}",
                 bg="#FFFFFF", fg="#8A90A8",
                 font=("Microsoft YaHei", 8),
                 anchor='w', wraplength=740).pack(fill=tk.X, padx=6, pady=(0, 6))

        # ── 主内容区：可滚动 canvas ─────────────────────────────
        content_frame = tk.Frame(win, bg="#F5F6FA")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        canvas = tk.Canvas(content_frame, bg="#F5F6FA", highlightthickness=0)
        vsb    = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner = tk.Frame(canvas, bg="#F5F6FA")
        inner_win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_cfg(e):
            canvas.itemconfig(inner_win, width=e.width)
        inner.bind("<Configure>", _on_inner_cfg)
        canvas.bind("<Configure>", _on_canvas_cfg)

        # 滚轮绑定
        def _wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<MouseWheel>",  _wheel)
        canvas.bind("<Button-4>",    lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>",    lambda e: canvas.yview_scroll(1,  "units"))
        inner.bind("<MouseWheel>",   _wheel)

        # ── 渲染货源卡片 ─────────────────────────────────────────
        COLORS = {
            '同款': ('#E8F8EF', '#27AE60', '✅ 同款'),
            '平替': ('#E8F0FB', '#185FA5', '🔄 平替'),
        }
        FG      = "#3D4152"
        FG_M    = "#8A90A8"
        SURF    = "#FFFFFF"
        BRD     = "#E2E5F0"
        FONT_LG = ("Microsoft YaHei", 10, "bold")
        FONT_MD = ("Microsoft YaHei", 9)
        FONT_SM = ("Microsoft YaHei", 8)

        def _section_title(parent, text, icon=""):
            row = tk.Frame(parent, bg="#F5F6FA")
            row.pack(fill=tk.X, pady=(10, 4))
            tk.Label(row, text=f"{icon} {text}", bg="#F5F6FA", fg="#FF6B35",
                     font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT, padx=2)
            tk.Frame(row, bg="#E2E5F0", height=1).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        def _render_items(parent, items, source_tag):
            if not items:
                tk.Label(parent, text=f"  {source_tag} 无货源数据",
                         bg="#F5F6FA", fg=FG_M, font=FONT_SM).pack(anchor='w', padx=4)
                return

            for rank, it in enumerate(items, 1):
                match_type = it.get('匹配类型', '')
                bg_c, acc_c, badge = COLORS.get(match_type, ('#F8F9FD', '#8A90A8', '⚪ 未知'))
                profit = it.get('预估利润(元)', 0) or 0
                if profit < 0:
                    bg_c, acc_c = '#FEF0F0', '#E74C3C'

                # 卡片外框
                card = tk.Frame(parent, bg=SURF,
                                highlightbackground=BRD, highlightthickness=1)
                card.pack(fill=tk.X, padx=4, pady=3)

                # 左侧色条
                bar = tk.Frame(card, bg=acc_c, width=4)
                bar.pack(side=tk.LEFT, fill=tk.Y)

                body = tk.Frame(card, bg=SURF)
                body.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 8), pady=6)

                # 第1行：序号 + 商品名 + 匹配徽章
                row1 = tk.Frame(body, bg=SURF)
                row1.pack(fill=tk.X)
                tk.Label(row1, text=f"#{rank}", bg=SURF, fg=FG_M,
                         font=FONT_SM, width=3, anchor='w').pack(side=tk.LEFT)
                tk.Label(row1, text=it.get('goods_name', ''),
                         bg=SURF, fg=FG,
                         font=FONT_LG, anchor='w', wraplength=500).pack(side=tk.LEFT, fill=tk.X, expand=True)
                badge_lbl = tk.Label(row1, text=badge, bg=bg_c, fg=acc_c,
                                     font=FONT_SM, padx=6, pady=2)
                badge_lbl.pack(side=tk.RIGHT, padx=(4, 0))

                # 第2行：利润数字（大字）
                row2 = tk.Frame(body, bg=SURF)
                row2.pack(fill=tk.X, pady=(3, 0))
                price_v = it.get('拼多多进价(元)', 0) or 0
                rate_v  = it.get('利润率(%)', 0)   or 0
                score_v = it.get('货源评分', 0)     or 0
                profit_color = "#27AE60" if profit > 0 else "#E74C3C"
                tk.Label(row2, text=f"¥{profit:.2f}", bg=SURF, fg=profit_color,
                         font=("Microsoft YaHei", 14, "bold")).pack(side=tk.LEFT)
                tk.Label(row2, text=" 利润", bg=SURF, fg=FG_M, font=FONT_SM).pack(side=tk.LEFT)
                tk.Label(row2, text=f"  进价 ¥{price_v:.2f}", bg=SURF, fg=FG, font=FONT_MD).pack(side=tk.LEFT, padx=(12, 0))
                tk.Label(row2, text=f"  利润率 {rate_v:.1f}%", bg=SURF, fg=FG, font=FONT_MD).pack(side=tk.LEFT, padx=(8, 0))
                tk.Label(row2, text=f"  评分 {score_v}", bg=SURF, fg=FG, font=FONT_MD).pack(side=tk.LEFT, padx=(8, 0))
                rec = it.get('是否推荐货源', '')
                if rec:
                    tk.Label(row2, text=rec, bg=SURF, fg=FG_M, font=FONT_SM).pack(side=tk.RIGHT, padx=4)

                # 第3行：销量 + 比对方式 + 相似度
                row3 = tk.Frame(body, bg=SURF)
                row3.pack(fill=tk.X, pady=(2, 0))
                sales = it.get('sales_tip', '') or ''
                sim   = it.get('sim_score', 0) or 0
                src   = it.get('match_src', '')
                meta  = []
                if sales:  meta.append(f"销量: {sales}")
                if src:    meta.append(f"比对: {src}")
                meta.append(f"相似度: {sim:.0%}")
                tk.Label(row3, text="  ".join(meta), bg=SURF, fg=FG_M, font=FONT_SM).pack(side=tk.LEFT)

                # 第4行：AI理由（如果有）
                reason = it.get('匹配说明', '') or it.get('reason', '')
                if reason:
                    row4 = tk.Frame(body, bg="#F8F9FD",
                                    highlightbackground="#E2E5F0", highlightthickness=1)
                    row4.pack(fill=tk.X, pady=(5, 0))
                    tk.Label(row4, text=f"💬 {reason}",
                             bg="#F8F9FD", fg="#5C6073",
                             font=FONT_SM, wraplength=660, justify=tk.LEFT,
                             anchor='w').pack(fill=tk.X, padx=8, pady=4)

        # 渲染标题搜索货源
        _section_title(inner, f"标题搜索货源（{len(title_items)} 件）", "📋")
        _render_items(inner, title_items, "标题搜索")

        # 渲染以图搜款货源
        _section_title(inner, f"以图搜款货源（{len(img_items)} 件）", "📷")
        _render_items(inner, img_items, "以图搜款")

        # 底部留白
        tk.Frame(inner, bg="#F5F6FA", height=20).pack()

        # ── 关闭按钮 ─────────────────────────────────────────────
        btn_frame = tk.Frame(win, bg="#F5F6FA")
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Button(btn_frame, text="关闭",
                  command=win.destroy,
                  bg="#EEF0F8", fg="#3D4152",
                  padx=24, pady=6,
                  font=("Microsoft YaHei", 9),
                  relief=tk.FLAT, cursor="hand2").pack()

    def _build_detail_table(self, parent, items: list, source_type: str):
        """保留兼容旧调用，实际渲染交给卡片布局（此方法已弃用，仅作空实现）"""
        pass

    def _rebuild_media_panel(self, record: dict):
        for w in self._media_inner.winfo_children():
            w.destroy()

        link = record.get('xianyu_link', '')
        pics = record.get('xianyu_pics', '')
        video = record.get('xianyu_video', '')

        def open_url(url):
            if url and url != '—':
                webbrowser.open(url)

        FONT_LBL = ("Microsoft YaHei", 8, "bold")
        FONT_LINK = ("Microsoft YaHei", 8)
        PAD = dict(padx=6, pady=2)

        row_link = tk.Frame(self._media_inner, bg=self.SURF)
        row_link.pack(fill=tk.X, **PAD)
        tk.Label(row_link, text="🔗 商品链接:", bg=self.SURF, fg=self.FG,
                 font=FONT_LBL, width=9, anchor='w').pack(side=tk.LEFT)
        if link:
            lbl = tk.Label(row_link, text=link[:80] + ('...' if len(link) > 80 else ''),
                           bg=self.SURF, fg="#0055CC", font=FONT_LINK,
                           cursor="hand2", anchor='w')
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)
            lbl.bind("<Button-1>", lambda e, u=link: open_url(u))
            tk.Label(row_link, text="↗ 点击浏览器打开", bg=self.SURF, fg=self.FG_M,
                     font=("Microsoft YaHei", 7)).pack(side=tk.LEFT, padx=(4, 0))
        else:
            tk.Label(row_link, text="—", bg=self.SURF, fg=self.FG_M,
                     font=FONT_LINK).pack(side=tk.LEFT)

        pic_list = [p.strip() for p in pics.split(',') if p.strip()] if pics else []
        hdr_row = tk.Frame(self._media_inner, bg=self.SURF)
        hdr_row.pack(fill=tk.X, padx=6, pady=(4, 0))
        tk.Label(hdr_row, text=f"🖼 图片（共{len(pic_list)}张，点击链接在浏览器打开）:",
                 bg=self.SURF, fg=self.FG, font=FONT_LBL).pack(anchor='w')

        for i, pic_url in enumerate(pic_list, 1):
            pic_row = tk.Frame(self._media_inner, bg=self.SURF)
            pic_row.pack(fill=tk.X, padx=6, pady=1)
            tk.Label(pic_row, text=f"  {i}.", bg=self.SURF, fg=self.FG_M,
                     font=FONT_LINK, width=3).pack(side=tk.LEFT)
            lbl_pic = tk.Label(pic_row,
                               text=pic_url[:100] + ('...' if len(pic_url) > 100 else ''),
                               bg=self.SURF, fg="#0055CC", font=FONT_LINK,
                               cursor="hand2", anchor='w')
            lbl_pic.pack(side=tk.LEFT, fill=tk.X, expand=True)
            lbl_pic.bind("<Button-1>", lambda e, u=pic_url: open_url(u))

        if not pic_list:
            tk.Label(self._media_inner, text="  （无图片数据）", bg=self.SURF,
                     fg=self.FG_M, font=FONT_LINK).pack(anchor='w', padx=6)

        if video:
            row_vid = tk.Frame(self._media_inner, bg=self.SURF)
            row_vid.pack(fill=tk.X, **PAD)
            tk.Label(row_vid, text="🎬 视频:", bg=self.SURF, fg=self.FG,
                     font=FONT_LBL, width=9, anchor='w').pack(side=tk.LEFT)
            lbl_vid = tk.Label(row_vid,
                               text=video[:80] + ('...' if len(video) > 80 else ''),
                               bg=self.SURF, fg="#0055CC", font=FONT_LINK,
                               cursor="hand2", anchor='w')
            lbl_vid.pack(side=tk.LEFT, fill=tk.X, expand=True)
            lbl_vid.bind("<Button-1>", lambda e, u=video: open_url(u))

    def _update_stats(self):
        def _u():
            q = self.task_queue.qsize()
            done = len(self._result_records)
            has_profit = sum(1 for r in self._result_records
                             if (r.get('final_profit') or 0) > 0)
            rates = []
            for r in self._result_records:
                fp = r.get('final_profit')
                fnp = r.get('final_price')
                if fp is not None and fnp and fnp > 0:
                    rates.append(fp / fnp * 100)
            mr = f"{max(rates):.1f}%" if rates else "—"
            model_txt = "✅ 已加载" if SameProductMatcher._loaded else "⏳ 加载中"
            try:
                self._stat_queue.config(text=str(q))
                self._stat_done.config(text=str(done))
                self._stat_profit.config(text=str(has_profit))
                self.stats_label.config(text=f"最高利润率: {mr}  |  模型: {model_txt}")
            except Exception:
                pass

        try:
            self.frame.after(0, _u)
        except Exception:
            pass

    def _get_export_path(self, suffix='') -> str:
        path_entry = getattr(self.app, 'save_path_entry', None)
        base = path_entry.get().strip() if path_entry else ''
        if not base:
            base = os.path.expanduser(f"~/Desktop/闲鱼货源{suffix}.xlsx")
        else:
            if not base.endswith('.xlsx'):
                base += '.xlsx'
            if suffix:
                base = base.replace('.xlsx', f'{suffix}.xlsx')
        return base

    def _export_all(self):
        """导出全部结果"""
        if not self._result_records:
            messagebox.showwarning("提示", "暂无数据")
            return
        
        now = datetime.now()
        date_str = now.strftime('%Y%m%d')
        time_str = now.strftime('%H%M%S')
        count = len(self._result_records)
        
        path = os.path.expanduser(
            f"~/Desktop/全部货源_{date_str}_{time_str}_{count}件.xlsx")
        export_supply_to_excel(self._result_records, path, self._log, profit_only=False)
        messagebox.showinfo("完成", f"已导出全部结果:\n{path}")

    def _export_profit_only(self):
        """导出正利润货源"""
        if not self._result_records:
            messagebox.showwarning("提示", "暂无数据")
            return
        
        profit_records = [r for r in self._result_records if r.get('final_profit', 0) > 0]
        if not profit_records:
            messagebox.showwarning("提示", "暂无正利润货源")
            return
        
        now = datetime.now()
        date_str = now.strftime('%Y%m%d')
        time_str = now.strftime('%H%M%S')
        count = len(profit_records)
        
        path = os.path.expanduser(
            f"~/Desktop/正利润货源_{date_str}_{time_str}_{count}件.xlsx")
        export_supply_to_excel(profit_records, path, self._log, profit_only=True)
        messagebox.showinfo("完成", f"已导出正利润货源:\n{path} 共{count}件")

    def _export_listing_advice(self):
        """导出上架建议表（v9.0新增）- 带时间戳和数量"""
        if not self._result_records:
            messagebox.showwarning("提示", "暂无数据，请先处理商品后再导出")
            return
        
        now = datetime.now()
        date_str = now.strftime('%Y%m%d')
        time_str = now.strftime('%H%M%S')
        count = len(self._result_records)
        
        path = os.path.expanduser(
            f"~/Desktop/上架建议_{date_str}_{time_str}_{count}件.xlsx")
        export_listing_advice_to_excel(self._result_records, path, self._log)
        messagebox.showinfo("完成", f"上架建议表已导出:\n{path}\n\n共{count}件商品，按优先级 S>A>B>C>D 排序。")

    def _start_scheduled_push_thread(self):
        if self._scheduled_push_running:
            return
        self._scheduled_push_running = True
        self._scheduled_push_thread = threading.Thread(
            target=self._scheduled_push_loop, daemon=True)
        self._scheduled_push_thread.start()

    def _scheduled_push_loop(self):
        last_push_date = None
        while self._scheduled_push_running:
            time.sleep(30)
            try:
                try:
                    enabled = self.enable_scheduled_push_var.get()
                    target_h = int(self.sched_hour_e.get() or 8)
                    target_m = int(self.sched_min_e.get() or 0)
                except Exception:
                    continue

                if not enabled:
                    continue

                now = datetime.now()

                if (now.hour == target_h and
                        now.minute == target_m and
                        last_push_date != now.date()):
                    self._log(f"⏰ 定时推送触发：{now.strftime('%H:%M')}，开始推送正利润货源...")
                    self._do_scheduled_push()
                    last_push_date = now.date()
                    try:
                        self.sched_status_lbl.config(
                            text=f"✅ 上次推送: {now.strftime('%Y-%m-%d %H:%M')}",
                            fg=self.SUCC)
                    except Exception:
                        pass
                else:
                    try:
                        next_dt = now.replace(hour=target_h, minute=target_m,
                                              second=0, microsecond=0)
                        if next_dt <= now:
                            next_dt += timedelta(days=1)
                        diff_min = int((next_dt - now).total_seconds() / 60)
                        self.sched_status_lbl.config(
                            text=f"⏰ 下次推送: {next_dt.strftime('%m-%d %H:%M')}（约{diff_min}分后）",
                            fg=self.FG_M)
                    except Exception:
                        pass
            except Exception:
                pass

    def _do_scheduled_push(self):
            """定时推送：正利润Excel + 上架建议Excel + 简报"""
            try:
                print("\n" + "=" * 60)
                print(f"[定时推送] 触发时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"[定时推送] _result_records 数量: {len(self._result_records)}")
                print(f"[定时推送] _pushed_ids 数量: {len(self._pushed_ids)}")
                
                # 打印所有商品
                print("\n[定时推送] 所有商品列表:")
                for i, r in enumerate(self._result_records):
                    fp = r.get('final_profit')
                    title = r.get('source_title', '')[:40]
                    rid = id(r)
                    is_pushed = rid in self._pushed_ids
                    print(f"  [{i}] {title} | profit={fp} | pushed={is_pushed}")
                
                if not hasattr(self, '_wechat_webhook') or not self._wechat_webhook:
                    print("[定时推送] ⚠️ 未配置 Webhook，跳过")
                    self._log("  ⚠️ 未配置 Webhook，定时推送跳过")
                    return

                # 找出新增的正利润商品
                new_profit_records = []
                for r in self._result_records:
                    rid = id(r)
                    if rid in self._pushed_ids:
                        continue
                    if r.get('final_profit') and r['final_profit'] > 0:
                        new_profit_records.append(r)
                
                print(f"\n[定时推送] 新增正利润商品数: {len(new_profit_records)}")
                for r in new_profit_records:
                    print(f"  - {r.get('source_title', '')[:40]} | ¥{r.get('final_profit')}")

                if not new_profit_records:
                    print("[定时推送] 无新增，退出")
                    self._log("  ℹ️ 定时推送：无新增正利润货源，跳过")
                    return

                # 使用更详细的文件名：日期_时分秒_商品数量
                now = datetime.now()
                date_str = now.strftime('%Y%m%d')
                time_str = now.strftime('%H%M%S')
                count = len(new_profit_records)
                
                print(f"[定时推送] 开始导出 {count} 件商品...")
                self._log(f"  📤 定时推送：{count} 个新正利润货源")

                # 1. 导出正利润Excel
                profit_excel_path = os.path.expanduser(
                    f"~/Desktop/正利润货源_{date_str}_{time_str}_{count}件.xlsx")
                export_supply_to_excel(new_profit_records, profit_excel_path, self._log, profit_only=True)
                print(f"[定时推送] 正利润Excel已导出: {profit_excel_path}")

                # 2. 导出上架建议Excel
                listing_excel_path = os.path.expanduser(
                    f"~/Desktop/上架建议_{date_str}_{time_str}_{count}件.xlsx")
                export_listing_advice_to_excel(new_profit_records, listing_excel_path, self._log)
                print(f"[定时推送] 上架建议Excel已导出: {listing_excel_path}")

                # 3. 发送正利润简报 + 上架建议简报
                self._send_profit_summary_via_wechat(new_profit_records, profit_excel_path)
                self._send_listing_advice_summary_via_wechat(new_profit_records, listing_excel_path)

                # 标记已推送
                for r in new_profit_records:
                    self._pushed_ids.add(id(r))

                print(f"[定时推送] ✅ 完成，已标记 {count} 条")
                self._log(f"  ✅ 定时推送完成，已标记 {count} 条为已推送")
                print("=" * 60 + "\n")
                
            except Exception as e:
                print(f"[定时推送] ❌ 异常: {e}")
                import traceback
                traceback.print_exc()
                self._log(f"  ❌ 定时推送异常: {e}")

    def _send_profit_summary_via_wechat(self, records: List[Dict], excel_path: str):
            """发送正利润货源简报到企业微信"""
            if not self._wechat_webhook:
                return

            try:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # 统计利润区间
                profit_0_10 = sum(1 for r in records if 0 < r.get('final_profit', 0) < 10)
                profit_10_30 = sum(1 for r in records if 10 <= r.get('final_profit', 0) < 30)
                profit_30_plus = sum(1 for r in records if r.get('final_profit', 0) >= 30)
                
                # 利润前5
                sorted_records = sorted(records, key=lambda x: x.get('final_profit', 0), reverse=True)
                top5_lines = []
                for r in sorted_records[:5]:
                    title = r.get('source_title', '')[:25]
                    profit = r.get('final_profit', 0)
                    quadrant = r.get('quadrant_label', '')
                    top5_lines.append(f"- {title} → ¥{profit:.1f} ({quadrant})")
                
                content = f"""## 💰 正利润货源简报

    **时间：** {timestamp}
    **新增正利润数：** {len(records)}

    ### 📊 利润分布
    - ¥30以上：{profit_30_plus} 件
    - ¥10~30：{profit_10_30} 件
    - ¥0~10：{profit_0_10} 件

    ### 🏆 利润前5
    {chr(10).join(top5_lines)}

    **附件：** 完整Excel文件
    """
                data = {"msgtype": "markdown", "markdown": {"content": content}}
                requests.post(self._wechat_webhook, json=data, timeout=10)
                self._log(f"  📤 正利润简报已发送（{len(records)}条）")

                # 发送Excel文件
                if excel_path and os.path.exists(excel_path):
                    self._send_file_to_wechat(excel_path, "file")
                    self._log("  📁 正利润Excel已发送")

            except Exception as e:
                self._log(f"  ⚠️ 发送正利润简报失败: {e}")

    def _send_listing_advice_summary_via_wechat(self, records: List[Dict], excel_path: str):
            """发送上架建议简报到企业微信"""
            if not self._wechat_webhook:
                return

            try:
                # 统计各级别数量
                stats = {'S级': 0, 'A级': 0, 'B级': 0, 'C级': 0, 'D级': 0}
                for r in records:
                    priority = _calc_listing_priority(r)
                    stats[priority] = stats.get(priority, 0) + 1

                # 获取S级和A级的商品简要列表
                top_items = []
                for r in records:
                    priority = _calc_listing_priority(r)
                    if priority in ('S级', 'A级'):
                        title = r.get('source_title', '')[:28]
                        profit = r.get('final_profit')
                        profit_str = f"¥{profit:.1f}" if profit else "待计算"
                        top_items.append(f"- {priority} {title} → {profit_str}")

                top_preview = "\n".join(top_items[:8])
                if len(top_items) > 8:
                    top_preview += f"\n... 共{len(top_items)}条"

                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                total = len(records)
                s_cnt = stats.get('S级', 0)
                a_cnt = stats.get('A级', 0)
                recommend_rate = f"{(s_cnt + a_cnt) / total * 100:.0f}%" if total > 0 else "0%"

                content = f"""## 📋 上架建议简报

    **时间：** {timestamp}
    **本次新增货源数：** {total}

    ### 📊 上架优先级分布
    - 🏆 **S级**（强烈推荐）：{s_cnt} 件
    - ✅ **A级**（建议上架）：{a_cnt} 件
    - 🔄 **B级**（可尝试）：{stats.get('B级', 0)} 件
    - ⚠️ **C级**（谨慎）：{stats.get('C级', 0)} 件
    - ❌ **D级**（不建议）：{stats.get('D级', 0)} 件

    **推荐上架率：** {recommend_rate}（S+A级占比）

    ### 🎯 S/A级商品预览
    {top_preview if top_preview else '（无S/A级商品）'}

    ---
    💡 **提示：** 优先处理S/A级商品，利润空间大、需求旺盛

    **附件：** 完整上架建议表
    """
                data = {"msgtype": "markdown", "markdown": {"content": content}}
                requests.post(self._wechat_webhook, json=data, timeout=10)
                self._log(f"  📤 上架建议简报已发送（S级{s_cnt}/A级{a_cnt}）")

                # 发送Excel文件
                if excel_path and os.path.exists(excel_path):
                    self._send_file_to_wechat(excel_path, "file")
                    self._log("  📁 上架建议Excel已发送")

            except Exception as e:
                self._log(f"  ⚠️ 发送上架建议简报失败: {e}")

    def _send_file_to_wechat(self, file_path: str, file_type: str = "file") -> bool:
            """上传文件到企业微信并发送"""
            if not self._wechat_webhook or not os.path.exists(file_path):
                return False
            
            try:
                # 构建上传URL
                upload_url = self._wechat_webhook.replace('/send?', '/upload_media?') + f'&type={file_type}'
                
                with open(file_path, 'rb') as f:
                    files = {'media': (os.path.basename(file_path), f)}
                    resp = requests.post(upload_url, files=files, timeout=30)
                    
                if resp.status_code == 200:
                    result = resp.json()
                    if result.get('errcode') == 0:
                        media_id = result.get('media_id')
                        # 发送文件消息
                        file_data = {"msgtype": "file", "file": {"media_id": media_id}}
                        requests.post(self._wechat_webhook, json=file_data, timeout=10)
                        return True
                    else:
                        self._log(f"  ⚠️ 上传失败: {result}")
                return False
            except Exception as e:
                self._log(f"  ⚠️ 上传文件异常: {e}")
                return False

    def set_wechat_webhook(self, webhook_url: str):
        self._wechat_webhook = webhook_url
        if hasattr(self, 'scheduler') and self.scheduler:
            self.scheduler._wechat_webhook = webhook_url
        self._log(f"📱 企业微信 Webhook 已设置")

    def destroy(self):
        self._scheduled_push_running = False
        if self.scheduler:
            self.scheduler.close()


def inject_supply_tab(notebook: ttk.Notebook, main_app) -> SupplyFinderTab:
    tab = SupplyFinderTab(notebook, main_app)
    tab.mount()
    return tab